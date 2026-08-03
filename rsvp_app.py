"""
rsvp_app.py — MAIN GUI for the Outlook RSVP Tool
=============================================================================
Runs on WINDOWS, requires Outlook desktop installed & signed in.

Install:
    pip install pywin32 openpyxl tkcalendar

Run:
    python rsvp_app.py

All on-screen text (labels, buttons, dialogs) is in ENGLISH, regardless of
which language the OUTGOING EMAIL is composed in — the email language is a
separate setting on the "Compose & Send" tab (English / Japanese /
Vietnamese / Bilingual EN+JP), since recipients may be Japanese and Indian
colleagues.

7 tabs:
    1. Event Setup        — event details incl. expected budget; can reload
                             a past event as a template
    2. Recipients          — load / edit the invite list directly
    3. Compose & Send       — Voting Buttons email (Yes/No/Maybe), choose
                             email language, translate the free-text note
                             via a Copilot copy/paste bridge
    4. Collect Responses    — scan Inbox, read vote results, export report
    5. Attendance & Payment — send a Calendar Invite to everyone who voted
                             Yes or Maybe, then track who actually showed
                             up and how much they paid
    6. Gift Contribution    — track who has contributed to a gift, with
                             per-person amount and running total
    7. Event History        — browse / reuse everything saved so far
"""
import os
import re
import tempfile
import threading
from datetime import datetime, timedelta

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    from tkcalendar import DateEntry
    HAS_TKCALENDAR = True
except ImportError:
    HAS_TKCALENDAR = False

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import history
import db
import outlook_com

APP_TITLE = "Outlook RSVP Tool"

# ══════════════════════════════════════════════════════════════════════════
# Language config
# ══════════════════════════════════════════════════════════════════════════
LANG_LABELS = {
    "en": "English",
    "ja": "Japanese",
    "vi": "Vietnamese",
    "bilingual": "Bilingual (English + Japanese)",
}
LANG_LABEL_TO_CODE = {v: k for k, v in LANG_LABELS.items()}
TRANSLATE_TARGETS = ["English", "Japanese", "Vietnamese", "Bilingual (Japanese + English)"]

BILINGUAL_SEPARATOR = "\n\n――――――――――――――――――――――――――――\n\n"

NOT_TRANSLATED_FLAG = {
    "en": "  [not yet translated — using original text]",
    "ja": "  【未翻訳 — 原文のまま表示しています】",
    "vi": "  [chưa dịch — đang hiển thị bản gốc]",
}


GREETING = {
    "en": "Hello everyone,",
    "ja": "皆様",
    "vi": "Chào các bạn,",
}

# ══════════════════════════════════════════════════════════════════════════
# Copilot translation prompt — SYSTEM DEFAULTS (used unless user customizes
# and saves an override via Tab 3's "Customize Copilot prompt" box).
# Both defaults explicitly ask for emoji icons so translated emails are more
# visually scannable (⏰ time, 📍 location, 💰 cost, 📋 deadline/instructions,
# 👥 attendees). Use "[TARGET_LANGUAGE]" as a placeholder in the SINGLE
# template — it gets swapped for the real language name before copying.
# ══════════════════════════════════════════════════════════════════════════
DEFAULT_PROMPT_SINGLE = (
    "Translate the following internal event email into [TARGET_LANGUAGE]. "
    "Reply with ONLY the final, ready-to-send translated email as continuous text — "
    "do NOT add section labels, headers, quotation marks, or any explanation/commentary "
    "about your process. Keep the same paragraph order as the original. Do not merge, "
    "duplicate, summarize, or repeat any sentence — translate everything exactly once. "
    "Keep a polite, professional tone suitable for a workplace event email, and keep "
    "names, dates, and numbers exactly as written.\n\n"
    "Also make the email more visually scannable by adding ONE relevant emoji icon "
    "inline right before these types of information (do not add a legend/key — just "
    "place the icon naturally in the sentence):\n"
    "  ⏰ before any time / date / deadline\n"
    "  📍 before any location / venue\n"
    "  💰 before any cost / budget / price\n"
    "  📋 before instructions or action items (e.g. what to bring, what to do)\n"
    "  👥 before attendee / participant / headcount information\n"
    "Use each icon at most once per relevant sentence — do not overuse them, and do "
    "not add icons to the Yes/No/Maybe voting-button instructions.\n\n"
    "FORMATTING — plain text only, this matters: do NOT use Markdown syntax anywhere "
    "(no **bold**, no '-' or '*' bullet lists, no '#' headers, no code blocks/backticks). "
    "Write each paragraph as continuous flowing text — do NOT break a single sentence or "
    "short phrase onto its own separate line; keep each paragraph together and only start "
    "a new line at real paragraph boundaries, separated by ONE blank line. (Copying from "
    "some AI chat interfaces can silently drop the line breaks between many short lines, "
    "gluing words together with no space — writing in fewer, longer flowing paragraphs "
    "avoids that problem.) For the Yes/No/Maybe list, keep each bullet ('•' character) on "
    "its own line as in the source."
)

DEFAULT_PROMPT_BILINGUAL = (
    "Translate the following internal event email into BOTH Japanese AND English, "
    "no matter what language the source text below is written in. "
    "Reply with ONE bilingual document structured like this: the complete JAPANESE "
    "version FIRST — starting with '[English below]' as its very first line — "
    "followed by a blank line and a divider line, then the complete ENGLISH version "
    "SECOND. Keep the two language versions COMPLETELY SEPARATE — do not interleave "
    "or mix sentences between languages; each version must be a full, independently "
    "readable translation covering everything in the source. Do not add any other "
    "section labels, headers, or commentary about your process. Do not merge, "
    "duplicate, summarize, or repeat any sentence within a language version. Keep a "
    "polite, professional tone suitable for a workplace event email, and keep names, "
    "dates, and numbers exactly as written.\n\n"
    "Also make BOTH versions more visually scannable by adding ONE relevant emoji icon "
    "inline right before these types of information (do not add a legend/key — just "
    "place the icon naturally in the sentence, in both language versions):\n"
    "  ⏰ before any time / date / deadline\n"
    "  📍 before any location / venue\n"
    "  💰 before any cost / budget / price\n"
    "  📋 before instructions or action items (e.g. what to bring, what to do)\n"
    "  👥 before attendee / participant / headcount information\n"
    "Use each icon at most once per relevant sentence — do not overuse them, and do "
    "not add icons to the Yes/No/Maybe voting-button instructions.\n\n"
    "FORMATTING — plain text only, this matters: do NOT use Markdown syntax anywhere "
    "(no **bold**, no '-' or '*' bullet lists, no '#' headers, no code blocks/backticks). "
    "Write each paragraph as continuous flowing text — do NOT break a single sentence or "
    "short phrase onto its own separate line; keep each paragraph together and only start "
    "a new line at real paragraph boundaries, separated by ONE blank line. (Copying from "
    "some AI chat interfaces can silently drop the line breaks between many short lines, "
    "gluing words together with no space — writing in fewer, longer flowing paragraphs "
    "avoids that problem.) For the Yes/No/Maybe list, keep each bullet ('•' character) on "
    "its own line as in the source."
)

# ══════════════════════════════════════════════════════════════════════════
# Cleanup helper for a known Copilot-copy quirk: pasting a reply copied from
# Copilot's web UI sometimes silently drops line breaks between short lines
# (NOT replaced by a space — just gone), so adjacent lines' words run
# together with zero separator, e.g. "Hello everyone," + "Please respond..."
# becomes "Helloeveryone,Pleaserespond...". This happens in Copilot's own
# copy mechanism (outside this tool's control), so it can't be fixed at the
# source — this is a best-effort, SAFE repair applied after pasting: it only
# touches spots anchored to something recognizable (emoji icons, bullet
# markers, Japanese full stops) and does not attempt to guess-reconstruct
# spaces lost strictly inside a plain sentence with no such anchor nearby.
#
# ALSO observed: Copilot's reply can contain the WHOLE email duplicated
# WITHIN THE SAME RESPONSE — an initial malformed/glued draft immediately
# followed by a self-corrected, properly-spaced version of the identical
# content. This is NOT the user pasting twice; it's literally what Copilot
# returned in one go. The two copies differ in SPACING (one glued, one not),
# so a plain exact-text duplicate check misses it — dedupe below compares a
# WHITESPACE-STRIPPED version of the text so spacing differences don't hide
# the match, then removes everything before the LAST copy (self-correction
# is normally the cleaner one).
# ══════════════════════════════════════════════════════════════════════════
EMOJI_PATTERN = re.compile(
    "["
    "\u2300-\u23FF"          # misc technical (⏰ alarm clock, etc.)
    "\u2600-\u27BF"          # misc symbols & dingbats (✅❌☀️ etc.)
    "\U0001F300-\U0001F5FF"  # misc symbols & pictographs (📍💰📋📅 etc.)
    "\U0001F600-\U0001F64F"  # emoticons
    "\U0001F680-\U0001F6FF"  # transport & map symbols
    "\U0001F900-\U0001F9FF"  # supplemental symbols & pictographs (👥 etc.)
    "\U0001FA70-\U0001FAFF"  # symbols & pictographs extended-A
    "]"
)


def _normalized_to_original_index(text, normalized_index):
    """Map an index into the whitespace-stripped version of `text` back to
    the corresponding index in the ORIGINAL `text` (helper for dedupe below,
    since detection runs on a whitespace-stripped copy but we need to cut
    the ORIGINAL text at the right spot)."""
    count = 0
    for i, ch in enumerate(text):
        if not ch.isspace():
            if count == normalized_index:
                return i
            count += 1
    return len(text)


def dedupe_pasted_translation(text, min_repeat_len=60):
    """If `text` contains the WHOLE email duplicated within itself, keep only
    the LAST copy and discard everything before it. Detection is done on a
    whitespace-STRIPPED copy of the text (see module comment above for why:
    the two copies typically differ only in spacing), then the cut point is
    mapped back to the correct index in the original text.
    Returns `text` unchanged if no duplication is detected."""
    if not text:
        return text
    normalized = re.sub(r"\s+", "", text)
    if len(normalized) < min_repeat_len * 2:
        return text
    head = normalized[:min_repeat_len]
    second_pos = normalized.find(head, min_repeat_len)
    if second_pos == -1:
        return text
    cut_at = _normalized_to_original_index(text, second_pos)
    return text[cut_at:].strip()


def detect_possible_duplicate_paste(text, min_repeat_len=60):
    """True if dedupe_pasted_translation() would meaningfully shorten `text`
    — i.e. the email content appears to be duplicated within itself (either
    from Copilot's own reply containing a malformed-then-corrected repeat, or
    from pasting a new result without clearing the box first)."""
    if not text:
        return False
    deduped = dedupe_pasted_translation(text, min_repeat_len=min_repeat_len)
    return (len(text) - len(deduped)) > 20


def cleanup_pasted_translation(text):
    """Best-effort repair for lost line breaks after pasting a Copilot reply
    (see module-level comment above for why this happens). Fixes, in order:
      1. Strip stray literal '**' markdown bold markers that leaked through
         as plain text.
      2. Ensure a space exists immediately before AND after every emoji icon
         — the icon is usually exactly where a lost line break used to be,
         so this recovers the most visible/common cases.
      3. Ensure a newline appears before each '•' bullet character, so the
         Yes/No/Maybe list renders as a proper list again.
      4. Insert a paragraph break after each Japanese full stop '。' that is
         glued directly to the next character with no space/newline.
    Does NOT deduplicate (see dedupe_pasted_translation() for that — the
    caller decides whether to dedupe first, since it's a more impactful
    change that's worth confirming with the user).
    Does NOT attempt to reconstruct spaces lost strictly inside a plain
    sentence with no icon/bullet/punctuation anchor nearby — please review
    the result before saving/sending."""
    if not text:
        return text

    text = text.replace("**", "")
    text = text.replace("∗∗∗", "").replace("***", "")  # stray emphasis markers

    text = EMOJI_PATTERN.sub(lambda m: f" {m.group(0)} ", text)
    text = re.sub(r"[ \t]{2,}", " ", text)     # collapse doubled spaces just created
    text = re.sub(r"[ \t]+\n", "\n", text)     # trim trailing spaces before a newline
    text = re.sub(r"\n[ \t]+", "\n", text)     # trim leading spaces after a newline

    text = re.sub(r"(?<!\n)[ \t]*•", "\n•", text)

    text = re.sub(r"。(?=[^\s\n])", "。\n\n", text)

    text = re.sub(r"\n{3,}", "\n\n", text)     # collapse 3+ blank lines to 1

    return text.strip()


def build_greeting(lang_code):
    if lang_code == "bilingual":
        return GREETING["en"] + BILINGUAL_SEPARATOR + GREETING["ja"]
    return GREETING.get(lang_code, GREETING["en"])


def build_subject(lang_code, event_id, event_name, is_update=False):
    if is_update:
        subs = {
            "en": f"[UPDATED-{event_id}] {event_name} - Event Details Have Changed",
            "ja": f"【変更のお知らせ-{event_id}】{event_name} - 開催情報が変更されました",
            "vi": f"[CapNhat-{event_id}] {event_name} - Thông tin sự kiện đã thay đổi",
        }
    else:
        subs = {
            "en": f"[Confirm-{event_id}] {event_name} - Please Confirm Attendance",
            "ja": f"【出欠確認-{event_id}】{event_name} - ご出欠確認のお願い",
            "vi": f"[XacNhan-{event_id}] {event_name} - Vui lòng xác nhận tham gia",
        }
    if lang_code == "bilingual":
        return subs["en"] + " / " + subs["ja"]
    return subs.get(lang_code, subs["en"])


UPDATE_NOTICE = {
    "en": ("⚠️ IMPORTANT: The details of this event have been UPDATED since the original "
           "invite. Please review the new information below (date/location/deadline/budget "
           "may have changed) and re-confirm your attendance if anything changed for you.\n"),
    "ja": ("⚠️ 重要: このイベントの詳細が、当初のご案内から変更されました（日時・場所・回答期限・"
           "費用などが変わっている可能性があります）。以下の新しい情報をご確認のうえ、必要に応じて"
           "再度ご回答をお願いいたします。\n"),
    "vi": ("⚠️ QUAN TRỌNG: Thông tin của sự kiện này đã được CẬP NHẬT so với email mời ban đầu "
           "(ngày giờ/địa điểm/hạn phản hồi/chi phí có thể đã thay đổi). Vui lòng xem lại thông "
           "tin mới bên dưới và xác nhận lại nếu có gì thay đổi với bạn.\n"),
}


def build_update_notice(lang_code):
    """Banner cảnh báo chèn ĐẦU email khi gửi ở chế độ 'Send update invite' —
    báo cho người nhận biết thông tin sự kiện đã thay đổi so với lần mời gốc."""
    if lang_code == "bilingual":
        return build_update_notice("en") + BILINGUAL_SEPARATOR + build_update_notice("ja")
    return UPDATE_NOTICE.get(lang_code, UPDATE_NOTICE["en"])


def build_fixed_block(lang_code, event_name, event_date, location, deadline, budget):
    """FIXED part: event details (auto-filled from Tab 1) + voting instructions
    (NO greeting — the greeting is shown separately, before the EDITABLE note).
    This is the SYSTEM DEFAULT wording — the person can hand-edit it in Tab 3 and save
    their own version as the new default (see fixed_overrides in RSVPApp)."""
    if lang_code == "bilingual":
        return (build_fixed_block("en", event_name, event_date, location, deadline, budget)
                + BILINGUAL_SEPARATOR
                + build_fixed_block("ja", event_name, event_date, location, deadline, budget))

    if lang_code == "ja":
        loc = f"（{location}）" if location else ""
        bud = f"予定費用: {budget}\n" if budget else ""
        return (
            f'「{event_name}」を {event_date} に開催いたします{loc}。\n\n'
            f'ご返信期限: {deadline}\n'
            f'{bud}\n'
            '>>> このメール上部にある「Yes / No / Maybe」の投票ボタンのいずれかをクリックしてください <<<\n'
            '（投票ボタンは通常メール本文の上にあるインフォバー、または Outlook デスクトップのリボンの'
            '「返信」欄に表示されます。Outlook Web でも同様にご利用いただけます）\n'
            '📎 ボタンの場所が分からない場合は、添付の画像（クリック方法の説明）をご確認ください。\n\n'
            '  • Yes   = 参加を確認します\n'
            '  • No    = 参加できません\n'
            '  • Maybe = まだ未定です、後日改めてご連絡します\n\n'
            '文字を入力したり件名を変更したりする必要はありません。ボタンをクリックするだけで自動的に記録されます。'
        )

    if lang_code == "vi":
        loc = f", tại {location}" if location else ""
        bud = f"Chi phí dự kiến: {budget}\n" if budget else ""
        return (
            f'Chúng ta tổ chức "{event_name}" vào ngày {event_date}{loc}.\n\n'
            f'Hạn phản hồi: {deadline}\n'
            f'{bud}\n'
            '>>> VUI LÒNG BẤM VÀO 1 TRONG 3 NÚT "Yes / No / Maybe" Ở ĐẦU EMAIL NÀY <<<\n'
            '(Nút vote thường hiện ở thanh InfoBar phía trên nội dung email, hoặc trong Ribbon '
            '"Respond" nếu bạn dùng Outlook desktop — Outlook Web cũng hỗ trợ tương tự)\n'
            '📎 Nếu không tìm thấy nút bấm, xem hình ảnh đính kèm bên dưới để biết vị trí cụ thể.\n\n'
            '  • Yes   = Bạn XÁC NHẬN tham gia\n'
            '  • No    = Bạn KHÔNG tham gia được\n'
            '  • Maybe = Bạn CHƯA CHẮC, sẽ xác nhận sau\n\n'
            'Bạn KHÔNG cần gõ chữ hay đổi tiêu đề — chỉ cần bấm 1 nút, hệ thống sẽ tự ghi nhận.'
        )

    # default / "en"
    loc = f", at {location}" if location else ""
    bud = f"Estimated cost: {budget}\n" if budget else ""
    return (
        f'Our team is organizing "{event_name}" on {event_date}{loc}.\n\n'
        f'Response deadline: {deadline}\n'
        f'{bud}\n'
        '>>> PLEASE CLICK ONE OF THE 3 BUTTONS "Yes / No / Maybe" AT THE TOP OF THIS EMAIL <<<\n'
        '(The voting buttons usually appear in the InfoBar above the message, or under the '
        '"Respond" section of the Ribbon in Outlook desktop — Outlook Web supports this too)\n'
        "📎 Not sure where the button is? See the attached image below for a visual guide.\n\n"
        '  • Yes   = You CONFIRM attendance\n'
        '  • No    = You CANNOT attend\n'
        '  • Maybe = NOT SURE yet, will confirm later\n\n'
        'You do not need to type anything or change the subject line — just click a button '
        'and the system will record it automatically.'
    )


REMINDER_LABELS = {
    "en": {"event": "📋 Event", "date": "⏰ Date", "location": "📍 Location",
           "budget": "💰 Budget", "deadline": "⏰ Please respond by"},
    "ja": {"event": "📋 イベント", "date": "⏰ 日時", "location": "📍 場所",
           "budget": "💰 予算", "deadline": "⏰ 回答期限"},
    "vi": {"event": "📋 Sự kiện", "date": "⏰ Ngày", "location": "📍 Địa điểm",
           "budget": "💰 Ngân sách", "deadline": "⏰ Vui lòng phản hồi trước"},
}


def build_reminder_body(lang_code, event_name, event_date, location, deadline, budget):
    """Nội dung email NHẮC NHỞ (Tab 4 'Collect Responses') — tương tự
    build_fixed_block() nhưng dùng giọng văn ngắn gọn, thân thiện, phù hợp
    cho email nhắc lại (không phải mời lần đầu). Hỗ trợ 4 lựa chọn giống hệt
    Tab 3: English / Japanese / Vietnamese / Bilingual (Japanese + English)
    — bilingual ghép theo đúng convention đã dùng ở nơi khác trong app
    (banner "[English below]" + nội dung tiếng Nhật trước, tiếng Anh sau,
    ngăn cách bằng BILINGUAL_SEPARATOR — xem _refresh_compose_preview())."""
    if lang_code == "bilingual":
        ja_full = build_reminder_body("ja", event_name, event_date, location, deadline, budget)
        en_full = build_reminder_body("en", event_name, event_date, location, deadline, budget)
        return "[English below]\n\n" + ja_full + BILINGUAL_SEPARATOR + en_full

    L = REMINDER_LABELS.get(lang_code, REMINDER_LABELS["en"])
    budget_line = f"{L['budget']}: {budget}\n" if budget else ""

    if lang_code == "ja":
        return (
            "皆様\n\n"
            "こちらはリマインドメールです。まだご回答をいただいておりません。\n\n"
            f"{L['event']}: {event_name}\n"
            f"{L['date']}: {event_date}\n"
            f"{L['location']}: {location}\n"
            f"{budget_line}"
            f"{L['deadline']}: {deadline}\n\n"
            ">>> このメール上部にある「Yes / No / Maybe」の投票ボタンのいずれかをクリックして"
            "ください <<<\n"
            "（投票ボタンは通常メール本文の上にあるインフォバー、または Outlook のリボンの"
            "「応答」欄に表示されます。Outlook Web でも同様にご利用いただけます）\n\n"
            "✅ このリマインドメール自体に直接投票していただいて問題ございません。"
            "（下に添付している最初のご案内メールを開き直す必要はありません。どちらのメールで"
            "投票されても、同じように記録されます。）\n\n"
            "・Yes = 出席\n"
            "・No = 欠席\n"
            "・Maybe = 未定、後日改めてご連絡します\n\n"
            "文字を入力したり件名を変更したりする必要はありません。ボタンをクリックするだけで"
            "自動的に記録されます。\n\n"
            "📎 ボタンの場所が分からない場合は、添付の画像をご確認ください。\n\n"
            "参考までに、元のご案内メールを添付しております。\n\n"
            "よろしくお願いいたします。"
        )

    if lang_code == "vi":
        return (
            "Xin chào,\n\n"
            "Đây là email nhắc nhở thân thiện — chúng tôi vẫn chưa nhận được phản hồi của bạn "
            "cho:\n\n"
            f"{L['event']}: {event_name}\n"
            f"{L['date']}: {event_date}\n"
            f"{L['location']}: {location}\n"
            f"{budget_line}"
            f"{L['deadline']}: {deadline}\n\n"
            ">>> Vui lòng bấm vào một trong các nút Yes / No / Maybe ở phía trên email này <<<\n"
            "(các nút này thường xuất hiện ở thanh thông tin phía trên nội dung email, hoặc "
            "trong mục \"Respond\" trên thanh Ribbon của Outlook — Outlook Web cũng hoạt động "
            "tương tự)\n\n"
            "✅ Bạn có thể bấm vote NGAY TRÊN email nhắc nhở này, không cần mở lại email mời "
            "gốc đính kèm bên dưới — vote trên email nào cũng được ghi nhận như nhau.\n\n"
            "• Yes = xác nhận tham dự\n"
            "• No = không thể tham dự\n"
            "• Maybe = chưa chắc chắn, sẽ phản hồi lại sau\n\n"
            "Không cần gõ gì thêm hay đổi tiêu đề — chỉ cần bấm nút là phản hồi của bạn sẽ được "
            "ghi nhận tự động.\n\n"
            "📎 Chưa biết nút nằm ở đâu? Xem ảnh đính kèm để được hướng dẫn trực quan.\n\n"
            "Email mời gốc được đính kèm để bạn tham khảo.\n\n"
            "Cảm ơn bạn!"
        )

    # default / "en"
    return (
        "Hello,\n\n"
        "This is a friendly reminder — we haven't received your response yet for:\n\n"
        f"{L['event']}: {event_name}\n"
        f"{L['date']}: {event_date}\n"
        f"{L['location']}: {location}\n"
        f"{budget_line}"
        f"{L['deadline']}: {deadline}\n\n"
        ">>> Please click one of the Yes / No / Maybe VOTING BUTTONS at the top of this "
        "email <<<\n"
        "(the buttons usually appear in the info bar above the message body, or in the "
        "\"Respond\" section of the Outlook ribbon — Outlook Web works the same way)\n\n"
        "✅ You can vote directly on THIS reminder email — no need to reopen the original "
        "invite attached below. Voting on either email is recorded the same way.\n\n"
        "  • Yes   = confirming attendance\n"
        "  • No    = cannot attend\n"
        "  • Maybe = not sure yet, will follow up later\n\n"
        "No need to type anything or change the subject — clicking a button records your "
        "response automatically.\n\n"
        "📎 Not sure where the button is? See the attached image for a visual guide.\n\n"
        "The original invitation email is attached for reference.\n\n"
        "Thank you!"
    )


def build_reminder_subject(lang_code, event_id, event_name):
    """Subject của email nhắc nhở — dùng chung cấu trúc [Reminder-{EventID}]
    để cơ chế Scan Inbox (Tab 4) vẫn khớp được Event ID bất kể ngôn ngữ nào
    (Scan chỉ cần Event ID có mặt trong Subject, không quan tâm phần chữ
    còn lại — xem docstring _collect_responses/_lookup_sent_date_hint)."""
    subs = {
        "en": f"[Reminder-{event_id}] {event_name} - Please Confirm Attendance",
        "ja": f"【リマインド-{event_id}】{event_name} - ご回答のお願い",
        "vi": f"[NhacNho-{event_id}] {event_name} - Vui lòng phản hồi",
    }
    if lang_code == "bilingual":
        return subs["ja"] + " / " + subs["en"]
    return subs.get(lang_code, subs["en"])


CALENDAR_LABELS = {
    "en": {"event": "📋 Event", "when": "⏰ Date", "where": "📍 Location", "budget": "💰 Budget"},
    "ja": {"event": "📋 イベント", "when": "⏰ 日時", "where": "📍 場所", "budget": "💰 予算"},
    "vi": {"event": "📋 Sự kiện", "when": "⏰ Ngày", "where": "📍 Địa điểm", "budget": "💰 Ngân sách"},
}


def build_calendar_body(lang_code, event_name="", event_date="", location="", budget=""):
    """Nội dung mặc định của Appointment body (Tab 5 'Attendance & Payment') — lời
    cảm ơn đã vote + nhắc lại thông tin sự kiện (EventName/EventDate/
    Location/Expected event budget, lấy từ Tab 1), gửi kèm Calendar Invite
    chính thức cho những người đã trả lời Yes/Maybe. Hỗ trợ 4 lựa chọn
    giống Tab 3/4: English/Japanese/Vietnamese/Bilingual (Japanese + English,
    ghép theo đúng convention "[English below]" + JA trước + EN sau đã dùng
    ở build_reminder_body()/build_gift_fixed_block()).
    Các tham số event_name/event_date/location/budget để trống ("") vẫn hợp
    lệ (vd lúc mới mở app, Tab 1 chưa điền gì) — chỉ hiện dòng tương ứng
    rỗng, không lỗi."""
    if lang_code == "bilingual":
        ja_full = build_calendar_body("ja", event_name, event_date, location, budget)
        en_full = build_calendar_body("en", event_name, event_date, location, budget)
        return "[English below]\n\n" + ja_full + BILINGUAL_SEPARATOR + en_full

    L = CALENDAR_LABELS.get(lang_code, CALENDAR_LABELS["en"])

    if lang_code == "ja":
        return (
            f"{event_name}への投票、ありがとうございました！\n"
            "正式なカレンダー招待を送付いたしますので、ご確認をお願いいたします。\n\n"
            f"{L['event']}: {event_name}\n"
            f"{L['when']}: {event_date}\n"
            f"{L['where']}: {location}\n"
            f"{L['budget']}: {budget}\n\n"
            "皆様にお会いできるのを楽しみにしております。"
        )

    if lang_code == "vi":
        return (
            f"Cảm ơn bạn đã vote cho sự kiện {event_name}!\n"
            "Tôi xin phép gửi lời mời lịch (Calendar Invite) chính thức như sau đây.\n\n"
            f"{L['event']}: {event_name}\n"
            f"{L['when']}: {event_date}\n"
            f"{L['where']}: {location}\n"
            f"{L['budget']}: {budget}\n\n"
            "Rất mong được gặp mọi người ở buổi tiệc/event!"
        )

    # default / "en"
    return (
        f"Thank you for voting for {event_name}!\n"
        "Please find the official calendar invitation below.\n\n"
        f"{L['event']}: {event_name}\n"
        f"{L['when']}: {event_date}\n"
        f"{L['where']}: {location}\n"
        f"{L['budget']}: {budget}\n\n"
        "Looking forward to seeing everyone there!"
    )


def build_thankyou_subject(lang_code, event_id, event_name):
    """Subject của email cảm ơn sau sự kiện (Tab 5 'Attendance & Payment',
    tính năng mới) — dùng tiền tố RIÊNG '[ThankYou-...]'/'【御礼-...】'/
    '[CamOn-...]' để không lẫn với Invite ('[Confirm-...]'), Reminder
    ('[Reminder-...]') hay Gift Notice ('[Gift-...]') — vẫn giữ event_id
    trong Subject theo đúng convention chung của app (dù email này không
    cần Scan Inbox lại, giữ để tra cứu/lọc mail nhất quán với các loại
    khác)."""
    subs = {
        "en": f"[ThankYou-{event_id}] Thank You for Attending {event_name}",
        "ja": f"【御礼-{event_id}】{event_name}へのご参加ありがとうございました",
        "vi": f"[CamOn-{event_id}] Cảm ơn bạn đã tham gia {event_name}",
    }
    if lang_code == "bilingual":
        return subs["ja"] + " / " + subs["en"]
    return subs.get(lang_code, subs["en"])


THANKYOU_LABELS = {
    "en": {"attend": "👥 Total attendees", "collected": "💰 Total collected",
           "paid": "💸 Amount paid", "remaining": "📊 Remaining amount"},
    "ja": {"attend": "👥 参加人数", "collected": "💰 集金総額",
           "paid": "💸 支払済み金額", "remaining": "📊 残金"},
    "vi": {"attend": "👥 Tổng số người tham gia", "collected": "💰 Tổng tiền thu",
           "paid": "💸 Số tiền đã trả", "remaining": "📊 Số tiền còn lại"},
}


def build_thankyou_body(lang_code, event_name="", event_date="", location="",
                         total_attend="0", total_collected="0", amount_paid="0",
                         remaining_amount="0"):
    """Nội dung mặc định của email cảm ơn sau sự kiện (Tab 5 'Attendance &
    Payment', tính năng mới) — lời cảm ơn mọi người đã tham gia (nội dung
    sự kiện lấy từ Tab 1, giống build_calendar_body()), kèm theo tổng quan
    số liệu Attendance & Payment (Tab 5): tổng số người tham gia thực tế,
    tổng tiền đã thu, số tiền đã trả, và số tiền còn lại — 4 tham số cuối
    LUÔN được tính lại từ dữ liệu MỚI NHẤT ngay trước khi gửi (xem
    _send_thank_you_email()), không phải giá trị lúc soạn mail, giống cách
    _compose_full_body() luôn build lại fixed_text từ Tab 1 thay vì tin vào
    nội dung đang hiển thị. Hỗ trợ 4 lựa chọn ngôn ngữ giống các tab khác."""
    if lang_code == "bilingual":
        ja_full = build_thankyou_body(
            "ja", event_name, event_date, location,
            total_attend, total_collected, amount_paid, remaining_amount)
        en_full = build_thankyou_body(
            "en", event_name, event_date, location,
            total_attend, total_collected, amount_paid, remaining_amount)
        return "[English below]\n\n" + ja_full + BILINGUAL_SEPARATOR + en_full

    L = THANKYOU_LABELS.get(lang_code, THANKYOU_LABELS["en"])

    if lang_code == "ja":
        return (
            f"{event_name}にご参加いただき、誠にありがとうございました！\n"
            "おかげさまで無事に終えることができました。\n\n"
            f"{L['attend']}: {total_attend}\n"
            f"{L['collected']}: {total_collected}\n"
            f"{L['paid']}: {amount_paid}\n"
            f"{L['remaining']}: {remaining_amount}\n\n"
            "参加費用の詳細は添付の集金リストをご確認ください。\n"
            "改めて、ご参加ありがとうございました。"
        )

    if lang_code == "vi":
        return (
            f"Cảm ơn bạn đã tham gia sự kiện {event_name}!\n"
            "Sự kiện đã diễn ra thành công tốt đẹp.\n\n"
            f"{L['attend']}: {total_attend}\n"
            f"{L['collected']}: {total_collected}\n"
            f"{L['paid']}: {amount_paid}\n"
            f"{L['remaining']}: {remaining_amount}\n\n"
            "Vui lòng xem file Excel đính kèm để biết chi tiết về người tham gia và chi phí.\n"
            "Một lần nữa xin cảm ơn mọi người đã tham gia!"
        )

    # default / "en"
    return (
        f"Thank you for attending {event_name}!\n"
        "The event went smoothly thanks to everyone's participation.\n\n"
        f"{L['attend']}: {total_attend}\n"
        f"{L['collected']}: {total_collected}\n"
        f"{L['paid']}: {amount_paid}\n"
        f"{L['remaining']}: {remaining_amount}\n\n"
        "Please see the attached spreadsheet for full attendee and cost details.\n"
        "Thanks again for being there!"
    )


def build_gift_subject(lang_code, event_id, guest_of_honor):
    """Subject email THÔNG BÁO QUYÊN GÓP QUÀ TẶNG (Tab 3 'Send Gift
    Contribution Notice') — dùng tiền tố RIÊNG '[Gift-...]'/'【寄付のお願い-...】'
    /'[QuyenGop-...]' để không lẫn với Invite RSVP thường ('[Confirm-...]')
    hay email nhắc nhở ('[Reminder-...]') — người nhận phân biệt ngay đây
    KHÔNG phải email cần bấm Vote, mà là thông báo kêu gọi đóng góp."""
    subs = {
        "en": f"[Gift-{event_id}] Farewell Gift Contribution for {guest_of_honor}",
        "ja": f"【寄付のお願い-{event_id}】{guest_of_honor}さんへの記念品ご協力のお願い",
        "vi": f"[QuyenGop-{event_id}] Kêu gọi đóng góp quà tặng {guest_of_honor}",
    }
    if lang_code == "bilingual":
        return subs["en"] + " / " + subs["ja"]
    return subs.get(lang_code, subs["en"])


GIFT_LABELS = {
    "en": {"gift_for": "🎁 Gift for", "when": "⏰ When", "where": "📍 Location",
           "contact": "📮 Contact to contribute", "deadline": "⏰ Please contribute by",
           "budget": "💰 Expected gift budget"},
    "ja": {"gift_for": "🎁 贈呈対象", "when": "⏰ 日時", "where": "📍 場所",
           "contact": "📮 ご協力の連絡先", "deadline": "⏰ お申し出期限",
           "budget": "💰 想定予算"},
    "vi": {"gift_for": "🎁 Quà tặng cho", "when": "⏰ Thời gian", "where": "📍 Địa điểm",
           "contact": "📮 Liên hệ đóng góp", "deadline": "⏰ Vui lòng đóng góp trước",
           "budget": "💰 Ngân sách dự kiến"},
}


def build_gift_fixed_block(lang_code, guest_of_honor, start_time, event_date, location,
                            organizer, deadline, gift_budget):
    """FIXED part của email THÔNG BÁO QUYÊN GÓP QUÀ TẶNG — tương tự
    build_fixed_block()/build_reminder_body() nhưng nội dung là kêu gọi
    đóng góp mua quà tặng (KHÔNG phải RSVP, không có Voting Buttons). Nội
    dung theo đúng mẫu người dùng yêu cầu:
    'Chúng tôi sẽ tặng quà cho {GuestOfHonor} vào lúc {StartTime}, ngày
    {EventDate}, ở {Location}. Liên hệ {Organizer} trước {Deadline} nếu
    muốn đóng góp. Expected budget: {GiftBudget}.'
    Hỗ trợ 4 lựa chọn giống các nơi khác: English/Japanese/Vietnamese/
    Bilingual (Japanese + English, ghép theo đúng convention "[English
    below]" + JA trước + EN sau)."""
    if lang_code == "bilingual":
        ja_full = build_gift_fixed_block("ja", guest_of_honor, start_time, event_date, location,
                                          organizer, deadline, gift_budget)
        en_full = build_gift_fixed_block("en", guest_of_honor, start_time, event_date, location,
                                          organizer, deadline, gift_budget)
        return "[English below]\n\n" + ja_full + BILINGUAL_SEPARATOR + en_full

    L = GIFT_LABELS.get(lang_code, GIFT_LABELS["en"])
    when_str = f"{start_time}, {event_date}" if start_time else event_date

    if lang_code == "ja":
        return (
            f"{L['gift_for']}: {guest_of_honor}さん\n"
            f"{L['when']}: {when_str}\n"
            f"{L['where']}: {location}\n\n"
            f"当日、{guest_of_honor}さんへ記念品を贈呈する予定です。\n"
            f"ご協力いただける方は、{L['deadline']}（{deadline}）までに{L['contact']}（{organizer}）までご連絡ください。\n\n"
            f"{L['budget']}: {gift_budget}\n\n"
            "何卒よろしくお願いいたします。"
        )

    if lang_code == "vi":
        return (
            f"{L['gift_for']}: {guest_of_honor}\n"
            f"{L['when']}: {when_str}\n"
            f"{L['where']}: {location}\n\n"
            f"Chúng tôi sẽ tiến hành tặng món quà kỷ niệm cho {guest_of_honor} vào lúc {when_str}, "
            f"tại {location}.\n"
            f"Nếu bạn có ý muốn đóng góp, xin hãy liên hệ {organizer} trước ngày {deadline}.\n\n"
            f"{L['budget']}: {gift_budget}\n\n"
            "Xin chân thành cảm ơn!"
        )

    # default / "en"
    return (
        f"{L['gift_for']}: {guest_of_honor}\n"
        f"{L['when']}: {when_str}\n"
        f"{L['where']}: {location}\n\n"
        f"We will be presenting a farewell gift to {guest_of_honor} at {when_str}, at {location}.\n"
        f"If you would like to contribute, please contact {organizer} before {deadline}.\n\n"
        f"{L['budget']}: {gift_budget}\n\n"
        "Thank you very much!"
    )


def build_gift_reminder_subject(lang_code, event_id, guest_of_honor):
    """Subject email NHẮC NHỞ QUYÊN GÓP QUÀ TẶNG (Tab 6 'Gửi email nhắc nhở
    tới người CHƯA đóng góp') — tiền tố RIÊNG '[Reminder-Gift-...]'/
    '【寄付リマインド-...】'/'[NhacQuyenGop-...]', khác hẳn tiền tố email Gift
    gốc ('[Gift-...]', build_gift_subject()) lẫn tiền tố Reminder RSVP
    thường ('[Reminder-...]', build_reminder_subject()) — để người nhận
    phân biệt ngay đây là NHẮC LẠI lời kêu gọi đóng góp trước đó, không phải
    thông báo mới hay email cần bấm Vote."""
    subs = {
        "en": f"[Reminder-Gift-{event_id}] Friendly Reminder — Gift Contribution for {guest_of_honor}",
        "ja": f"【寄付リマインド-{event_id}】{guest_of_honor}さんへの記念品ご協力のお願い（リマインド）",
        "vi": f"[NhacQuyenGop-{event_id}] Nhắc nhở đóng góp quà tặng {guest_of_honor}",
    }
    if lang_code == "bilingual":
        return subs["en"] + " / " + subs["ja"]
    return subs.get(lang_code, subs["en"])


def _gift_reminder_single_lang_body(lang_code, guest_of_honor, start_time, event_date, location,
                                     organizer, deadline, gift_budget):
    L = GIFT_LABELS.get(lang_code, GIFT_LABELS["en"])
    when_str = f"{start_time}, {event_date}" if start_time else event_date

    if lang_code == "ja":
        return (
            "（このメールはリマインドです。まだご協力の連絡をいただいていない方へお送りしています。）\n\n"
            f"{L['gift_for']}: {guest_of_honor}さん\n"
            f"{L['when']}: {when_str}\n"
            f"{L['where']}: {location}\n\n"
            f"以前ご案内した通り、{guest_of_honor}さんへ記念品を贈呈する予定です。\n"
            f"まだご連絡いただいていない場合は、{L['deadline']}（{deadline}）までに"
            f"{L['contact']}（{organizer}）までご連絡いただけますと幸いです。\n\n"
            f"{L['budget']}: {gift_budget}\n\n"
            "（参考として、以前送付したご案内メールを添付しております。）\n\n"
            "お忙しいところ恐れ入りますが、何卒よろしくお願いいたします。"
        )

    if lang_code == "vi":
        return (
            "(Đây là email nhắc nhở nhẹ — gửi tới những bạn chưa phản hồi đóng góp.)\n\n"
            f"{L['gift_for']}: {guest_of_honor}\n"
            f"{L['when']}: {when_str}\n"
            f"{L['where']}: {location}\n\n"
            f"Như đã thông báo trước đó, chúng tôi sẽ tặng món quà kỷ niệm cho {guest_of_honor} "
            f"vào lúc {when_str}, tại {location}.\n"
            f"Nếu bạn vẫn muốn đóng góp nhưng chưa kịp liên hệ, xin vui lòng liên hệ {organizer} "
            f"trước ngày {deadline} giúp mình nhé.\n\n"
            f"{L['budget']}: {gift_budget}\n\n"
            "(Đính kèm email thông báo gốc để bạn tiện xem lại chi tiết.)\n\n"
            "Không có gì gấp, chỉ là nhắc nhẹ thôi — xin cảm ơn bạn rất nhiều!"
        )

    # default / "en"
    return (
        "(This is a gentle reminder sent only to those who haven't responded yet.)\n\n"
        f"{L['gift_for']}: {guest_of_honor}\n"
        f"{L['when']}: {when_str}\n"
        f"{L['where']}: {location}\n\n"
        f"As announced earlier, we'll be presenting a farewell gift to {guest_of_honor} "
        f"at {when_str}, at {location}.\n"
        f"If you'd still like to contribute but haven't had a chance to reach out yet, "
        f"please contact {organizer} before {deadline}.\n\n"
        f"{L['budget']}: {gift_budget}\n\n"
        "(The original notice email is attached for your reference.)\n\n"
        "No pressure at all — just a friendly nudge. Thank you so much!"
    )


def build_gift_reminder_body(lang_code, guest_of_honor, start_time, event_date, location,
                              organizer, deadline, gift_budget):
    """BODY email nhắc nhở quyên góp — cùng cấu trúc song ngữ với
    build_gift_fixed_block() ('[English below]' + JA trước + EN sau, ghép
    bằng BILINGUAL_SEPARATOR), chỉ khác nội dung là bản NHẮC LẠI thay vì
    thông báo lần đầu."""
    if lang_code == "bilingual":
        ja_full = _gift_reminder_single_lang_body("ja", guest_of_honor, start_time, event_date,
                                                    location, organizer, deadline, gift_budget)
        en_full = _gift_reminder_single_lang_body("en", guest_of_honor, start_time, event_date,
                                                    location, organizer, deadline, gift_budget)
        return "[English below]\n\n" + ja_full + BILINGUAL_SEPARATOR + en_full
    return _gift_reminder_single_lang_body(lang_code, guest_of_honor, start_time, event_date,
                                            location, organizer, deadline, gift_budget)


def build_gift_report_subject(lang_code, event_id, guest_of_honor):
    """Subject email BÁO CÁO số tiền quyên góp đã thu được (Tab 6, tính
    năng mới) — tiền tố RIÊNG '[GiftReport-...]'/'【集金報告-...】'/
    '[BaoCaoQuyenGop-...]' để không lẫn với thông báo kêu gọi đóng góp ban
    đầu ('[Gift-...]', xem build_gift_subject()) hay email nhắc nhở."""
    subs = {
        "en": f"[GiftReport-{event_id}] Contribution Report for {guest_of_honor}",
        "ja": f"【集金報告-{event_id}】{guest_of_honor}さんへの寄付金報告",
        "vi": f"[BaoCaoQuyenGop-{event_id}] Báo cáo quyên góp cho {guest_of_honor}",
    }
    if lang_code == "bilingual":
        return subs["en"] + " / " + subs["ja"]
    return subs.get(lang_code, subs["en"])


GIFT_REPORT_LABELS = {
    "en": {"count": "👥 Total contributors", "total": "💰 Total collected"},
    "ja": {"count": "👥 寄付者数", "total": "💰 集金総額"},
    "vi": {"count": "👥 Tổng số người đóng góp", "total": "💰 Tổng tiền thu được"},
}


def build_gift_report_body(lang_code, guest_of_honor="", event_name="", contributor_count="0",
                            total_amount="0"):
    """Nội dung mặc định của email báo cáo số tiền đã quyên góp (Tab 6) —
    CHỈ là 1 thông báo TỔNG QUAN (đã thu được bao nhiêu người/bao nhiêu
    tiền), KHÔNG liệt kê danh sách từng người trong nội dung mail — danh
    sách chi tiết (No./Name/Email/Amount, chỉ những người ĐÃ đóng góp,
    đánh số lại từ 1, KHÔNG gồm 2 cột checkbox "Send email"/"Contributed")
    nằm trong file Excel ĐÍNH KÈM (xem _build_gift_report_workbook()) —
    người nhận mở file đính kèm để xem chi tiết. contributor_count/
    total_amount LUÔN được tính lại từ dữ liệu MỚI NHẤT ngay trước khi gửi
    (xem _send_gift_report_email()). Hỗ trợ 4 lựa chọn ngôn ngữ giống các
    tab khác."""
    if lang_code == "bilingual":
        ja_full = build_gift_report_body("ja", guest_of_honor, event_name, contributor_count, total_amount)
        en_full = build_gift_report_body("en", guest_of_honor, event_name, contributor_count, total_amount)
        return "[English below]\n\n" + ja_full + BILINGUAL_SEPARATOR + en_full

    L = GIFT_REPORT_LABELS.get(lang_code, GIFT_REPORT_LABELS["en"])

    if lang_code == "ja":
        return (
            f"{guest_of_honor}さんへの記念品にご協力いただき、誠にありがとうございました！\n"
            "現在までの集金状況を下記の通りご報告いたします。\n\n"
            f"{L['count']}: {contributor_count}\n"
            f"{L['total']}: {total_amount}\n\n"
            "詳細（お一人おひとりの内訳）は、添付の集金リストをご確認ください。\n\n"
            "改めまして、ご協力いただき誠にありがとうございました。"
        )

    if lang_code == "vi":
        return (
            f"Cảm ơn mọi người đã đóng góp quà tặng cho {guest_of_honor}!\n"
            "Đây là báo cáo tình hình quyên góp tính đến thời điểm hiện tại.\n\n"
            f"{L['count']}: {contributor_count}\n"
            f"{L['total']}: {total_amount}\n\n"
            "Vui lòng xem file đính kèm để biết chi tiết đóng góp của từng người.\n\n"
            "Xin chân thành cảm ơn sự đóng góp của mọi người!"
        )

    # default / "en"
    return (
        f"Thank you for contributing to the farewell gift for {guest_of_honor}!\n"
        "Here is the contribution report so far.\n\n"
        f"{L['count']}: {contributor_count}\n"
        f"{L['total']}: {total_amount}\n\n"
        "Please see the attached spreadsheet for the per-person breakdown.\n\n"
        "Thank you again for your generosity!"
    )


def build_editable_block(lang_code, note, note_is_translated):
    """EDITABLE part: ONLY the organizer's free-text background note for this event."""
    if lang_code == "bilingual":
        return note  # caller combines per-language notes separately for bilingual
    flag = "" if note_is_translated else NOT_TRANSLATED_FLAG.get(lang_code, "")
    return f"{note}{flag}"


# ══════════════════════════════════════════════════════════════════════════
# Date-picker helper widget
# ══════════════════════════════════════════════════════════════════════════
def make_date_picker(parent, initial=None):
    if HAS_TKCALENDAR:
        w = DateEntry(parent, date_pattern="dd/mm/yyyy", width=14)
        if initial:
            try:
                w.set_date(initial)
            except Exception:
                pass
        return w
    else:
        w = ttk.Entry(parent, width=16)
        w.insert(0, initial.strftime("%d/%m/%Y") if hasattr(initial, "strftime") else (initial or ""))
        return w


def get_date_str(widget):
    if HAS_TKCALENDAR and isinstance(widget, DateEntry):
        return widget.get_date().strftime("%d/%m/%Y")
    return widget.get().strip()


def get_date_obj(widget, default_hour=18, default_minute=0):
    if HAS_TKCALENDAR and isinstance(widget, DateEntry):
        d = widget.get_date()
        return datetime(d.year, d.month, d.day, default_hour, default_minute)
    try:
        d, m, y = widget.get().strip().split("/")
        return datetime(int(y), int(m), int(d), default_hour, default_minute)
    except Exception:
        return datetime.now()


def set_date_str(widget, date_str):
    """Nạp lại giá trị ngày (chuỗi 'dd/mm/yyyy', đọc từ History) vào 1 ô
    date-picker (DateEntry của tkcalendar, hoặc Entry thường nếu không có
    tkcalendar). Bỏ qua an toàn (không làm gì) nếu date_str rỗng hoặc không
    parse được — để không làm hỏng ô đang có sẵn."""
    if not date_str:
        return
    try:
        d, m, y = str(date_str).strip().split("/")
        d, m, y = int(d), int(m), int(y)
        if HAS_TKCALENDAR and isinstance(widget, DateEntry):
            widget.set_date(datetime(y, m, d))
        else:
            widget.delete(0, "end")
            widget.insert(0, f"{d:02d}/{m:02d}/{y}")
    except Exception:
        pass  # không parse được (định dạng lạ) — giữ nguyên giá trị đang có trên ô


def parse_amount_from_text(text):
    """Extracts the first numeric amount found in a free-text string, e.g.
    the "Expected gift budget" field on Tab 1 is usually typed as something
    like "3,000 JPY / person" — this pulls out 3000.0 so Tab 6 can auto-fill
    and total the per-person contribution amount. Returns 0.0 if no number
    is found (e.g. the field is still empty)."""
    if not text:
        return 0.0
    m = re.search(r"[\d][\d,\.]*", str(text))
    if not m:
        return 0.0
    raw = m.group(0).replace(",", "")
    try:
        return float(raw)
    except ValueError:
        return 0.0


def read_gift_contribution_rows(path):
    """Reads a Gift_Contribution_List_*.xlsx file and returns
    list[dict(name, email, checked, amount)], matching columns by their
    HEADER text (row 1) instead of a fixed position — so both older files
    (Name/Email/Contributed only) and newer files (No./Name/Email/Amount/
    Contributed) can be read back correctly."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1), ())
    header = {(str(c.value).strip() if c.value else ""): i for i, c in enumerate(header_row)}
    i_name = header.get("Name")
    i_email = header.get("Email")
    i_amount = header.get("Amount")
    i_contrib = header.get("Contributed")
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        email = row[i_email] if i_email is not None and i_email < len(row) else None
        if not email:
            continue  # skips blank rows and the "TOTAL COLLECTED" summary row
        name = row[i_name] if i_name is not None and i_name < len(row) else None
        contributed = row[i_contrib] if i_contrib is not None and i_contrib < len(row) else None
        amount = row[i_amount] if i_amount is not None and i_amount < len(row) else None
        rows.append({
            "name": str(name).strip() if name else str(email).strip(),
            "email": str(email).strip(),
            "checked": (str(contributed).strip().lower() == "yes") if contributed is not None else False,
            "amount": parse_amount_from_text(amount) if amount is not None else 0.0,
        })
    return rows


# ══════════════════════════════════════════════════════════════════════════
# Scrollable tab container — every tab's real content goes inside `.body`,
# so long tabs get a vertical (and horizontal, if needed) scrollbar instead
# of being cut off on smaller screens / windows.
# ══════════════════════════════════════════════════════════════════════════
class ScrollableFrame(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        # BUG ĐÃ SỬA: tk.Canvas KHÔNG tự lấy màu nền theo theme ttk ("clam")
        # — mặc định nó dùng màu trắng thuần của hệ thống, khác hẳn màu nền
        # xám/xanh nhạt mà mọi ttk.Frame/ttk.Label khác trong app đang dùng.
        # Kết quả: bất cứ khoảng trống nào trong vùng cuộn (canvas rộng hơn
        # nội dung thật bên trong self.body) đều lộ ra 1 mảng trắng lệch
        # tông so với phần nền chứa nút bấm/nhãn thông báo xung quanh. Lấy
        # đúng màu nền mà style ttk "TFrame" đang dùng (fallback về màu xám
        # nhạt tiêu chuẩn của theme "clam" nếu vì lý do gì đó chưa tra được)
        # rồi gán thẳng cho canvas để 2 vùng luôn cùng 1 màu.
        style = ttk.Style(self)
        bg_color = style.lookup("TFrame", "background") or "#dcdad5"
        canvas = tk.Canvas(self, borderwidth=0, highlightthickness=0, background=bg_color)
        vscroll = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        hscroll = ttk.Scrollbar(self, orient="horizontal", command=canvas.xview)
        self.body = ttk.Frame(canvas)

        self.body.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas_window = canvas.create_window((0, 0), window=self.body, anchor="nw")

        # BUG ĐÃ SỬA: self.body trước đây được đặt vào canvas với kích
        # thước CỐ ĐỊNH bằng đúng nội dung của nó lúc tạo — khi cửa sổ app
        # được kéo RỘNG hơn, canvas nới rộng ra theo nhưng self.body (và do
        # đó mọi nhãn thông báo/nút bấm bên trong) KHÔNG nới theo, để lại 1
        # dải trống bên phải/dưới, đúng như hiện tượng "phần thông báo
        # không tự động fix theo" khi resize. Giờ mỗi khi canvas đổi kích
        # thước, ép item cửa sổ bên trong (self.body) rộng bằng đúng canvas
        # — nội dung sẽ luôn lấp đầy chiều ngang, và các nhãn word-wrap
        # (_make_wrapping_label) tính lại đúng bề rộng thật để xuống dòng.
        def _on_canvas_configure(event):
            canvas.itemconfig(canvas_window, width=event.width)
            # MỚI: cũng tính lại wraplength cho MỌI label đã đăng ký vào
            # ĐÚNG canvas này (xem RSVPApp._make_wrapping_label() bên dưới)
            # — dùng TRỰC TIẾP chiều rộng THẬT của canvas, thay vì suy luận
            # gián tiếp từ chiều rộng cửa sổ gốc (self.winfo_width()) như
            # cách cũ. 2 con số đó lệch nhau khá nhiều (thanh cuộn dọc,
            # padding của Notebook, indent riêng của từng frame con...),
            # khiến 1 số label tính sai độ rộng cần xuống dòng và không co
            # lại đúng khi cửa sổ bị thu nhỏ — đây là NGUYÊN NHÂN gốc của
            # lỗi "nội dung hướng dẫn chưa fit theo khi resize". Giờ mỗi
            # canvas tự quản lý danh sách label của riêng nó, luôn khớp
            # đúng chiều rộng thực tế đang hiển thị của ĐÚNG tab đó.
            for lbl, margin in list(getattr(canvas, "wrap_labels", [])):
                try:
                    lbl.configure(wraplength=max(event.width - margin, 250))
                except tk.TclError:
                    pass  # label đã bị huỷ — bỏ qua an toàn
        canvas.bind("<Configure>", _on_canvas_configure)

        canvas.configure(yscrollcommand=vscroll.set, xscrollcommand=hscroll.set)

        canvas.grid(row=0, column=0, sticky="nsew")
        vscroll.grid(row=0, column=1, sticky="ns")
        hscroll.grid(row=1, column=0, sticky="ew")
        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        def _bind_wheel(_e):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)

        def _unbind_wheel(_e):
            canvas.unbind_all("<MouseWheel>")

        canvas.bind("<Enter>", _bind_wheel)
        canvas.bind("<Leave>", _unbind_wheel)
        self.canvas = canvas


def make_scrollable_treeview(parent, columns, height=10, show="headings"):
    """Creates a Treeview WITH its scrollbar, both inside a small container frame.
    Returns (container, tree). Grid/pack the returned CONTAINER — never the tree itself —
    since the tree's actual parent is the container, not the outer `parent` passed in."""
    container = ttk.Frame(parent)
    tree = ttk.Treeview(container, columns=columns, show=show, height=height)
    vscroll = ttk.Scrollbar(container, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=vscroll.set)
    tree.pack(side="left", fill="both", expand=True)
    vscroll.pack(side="right", fill="y")
    return container, tree


def make_scrollable_text(parent, **text_kwargs):
    """Creates a Text widget WITH its scrollbar, both inside a small container frame.
    Returns (container, text_widget). Grid/pack the returned CONTAINER — never the text
    widget itself — since the text widget's actual parent is the container."""
    container = ttk.Frame(parent)
    text_widget = tk.Text(container, **text_kwargs)
    vscroll = ttk.Scrollbar(container, orient="vertical", command=text_widget.yview)
    text_widget.configure(yscrollcommand=vscroll.set, wrap="word")
    text_widget.pack(side="left", fill="both", expand=True)
    vscroll.pack(side="right", fill="y")
    return container, text_widget


# ══════════════════════════════════════════════════════════════════════════
class RSVPApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1020x720")

        # ── shared state across tabs ──
        self.recipients = []
        self.recipient_file = tk.StringVar(value="Danh_sach_nguoi_tham_gia_TEMPLATE.xlsx")
        self.responses = {}
        self._pending_recipients = []  # cập nhật mỗi lần Scan Inbox (Tab 4) — người trong Tab 2 chưa vote
        self._last_scanned_event_id = None  # Event ID của lần Scan Inbox gần nhất — dùng để cảnh báo dữ liệu cũ
        self._last_scan_time = None
        # MỚI: đổi kiến trúc lưu trữ — history_path giờ trỏ tới 1 file SQLite
        # DUY NHẤT (rsvp_data.db, xem db.py) thay vì RSVP_History.xlsx. Toàn
        # bộ Events/Recipients/Gift Contribution/Attendance & Payment/
        # Responded result giờ sống trong DB này; Excel CHỈ còn được tạo ra
        # khi bấm nút "Export to Excel" ở từng tab, không tự động ghi liên
        # tục nữa (giữ tên biến "history_path" để không phải sửa lại quá
        # nhiều nơi — chỉ đổi Ý NGHĨA những gì nó trỏ tới).
        self.history_path = tk.StringVar(value=db.DB_FILE_DEFAULT)
        self.full_translations = {"en": "", "ja": "", "vi": "", "bilingual": ""}  # Copilot-translated full email bodies
        # MỚI: bản dịch Copilot RIÊNG cho mode "Send Gift Contribution Notice"
        # — tách khỏi self.full_translations (dùng cho Invite/Update invite)
        # để tránh 2 loại nội dung email hoàn toàn khác nhau ghi đè lẫn nhau
        # khi chỉ đổi qua lại Send mode trên CÙNG 1 ngôn ngữ.
        self.gift_full_translations = {"en": "", "ja": "", "vi": "", "bilingual": ""}
        self.fixed_overrides = history.load_fixed_overrides()  # user-customized default FIXED wording, persisted to disk
        self.prompt_overrides = history.load_prompt_overrides()  # user-customized Copilot prompt templates, persisted to disk

        # MỚI: nhập 1 LẦN DUY NHẤT dữ liệu từ bộ file Excel CŨ (nếu có, từ
        # trước khi chuyển sang kiến trúc SQLite này) vào rsvp_data.db — an
        # toàn khi gọi lại mỗi lần mở app (tự bỏ qua nếu DB đã có dữ liệu
        # rồi, xem docstring migrate_from_excel_if_needed()). Thông báo kết
        # quả cho user biết SAU KHI cửa sổ chính đã hiện ra (self.after),
        # tránh popup chặn trước khi app kịp vẽ xong.
        try:
            migrated, event_count, migrate_notes = db.migrate_from_excel_if_needed(
                db_path=self.history_path.get(), history_xlsx=history.HISTORY_FILE_DEFAULT)
        except Exception:
            migrated, event_count, migrate_notes = False, 0, []
        if migrated and event_count:
            note_text = ("\n\n" + "\n".join(migrate_notes[:5])) if migrate_notes else ""
            self.after(600, lambda: messagebox.showinfo(
                "Old data imported",
                f"Found your previous RSVP_History.xlsx and imported {event_count} event(s) "
                f"(plus any matching Recipients/Gift Contribution/Attendance & Payment files) "
                f"into the new database (rsvp_data.db). Your old Excel files were NOT modified "
                f"or deleted — they're just no longer the primary source of data." + note_text
            ))


        if not HAS_TKCALENDAR:
            messagebox.showwarning(
                "tkcalendar not installed",
                "tkcalendar is not installed, so date fields will use plain text entry "
                "(type manually as dd/mm/yyyy).\n\n"
                "Install it to get a clickable calendar:\n"
                "pip install tkcalendar"
            )

        # ── Notebook tab styling ──
        # By default, ttk uses the OS theme's tab colors — on Windows this
        # is usually "vista"/"xpnative", which draws the Notebook tab
        # background NATIVELY and ignores our custom "background" color,
        # while still respecting our custom "foreground" (text) color. That
        # mismatch is exactly what caused the bug report: the selected
        # tab's background stayed the native white/light color while its
        # text was set to white, making the label invisible. Switching to
        # the "clam" theme (a pure-Tk theme, not OS-native) makes ttk fully
        # respect our style customization, so both background AND
        # foreground actually render as configured below. This does change
        # the look of other widgets slightly too (buttons, checkboxes,
        # etc.) to the "clam" style, but keeps the app fully usable and
        # fixes the readability bug for good, instead of just tweaking
        # colors that the old theme might ignore again.
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass  # "clam" should always be available, but fall back safely if not
        style.configure(
            "TNotebook.Tab",
            padding=(14, 8),
            font=("Arial", 10, "bold"),
            background="#D9E2F3",
            foreground="#003366",
        )
        style.map(
            "TNotebook.Tab",
            background=[("selected", "#003366"), ("active", "#B7C9EB")],
            foreground=[("selected", "#FFFFFF"), ("active", "#003366")],
        )

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=8, pady=8)
        self.nb = nb

        self.tab_config = ScrollableFrame(nb)
        self.tab_recipients = ScrollableFrame(nb)
        self.tab_compose = ScrollableFrame(nb)
        self.tab_collect = ScrollableFrame(nb)
        self.tab_calendar = ScrollableFrame(nb)
        self.tab_gift = ScrollableFrame(nb)  # MỚI — theo dõi quyên góp quà tặng
        self.tab_history = ScrollableFrame(nb)

        nb.add(self.tab_config, text="  1. Event Setup  ")
        nb.add(self.tab_recipients, text="  2. Recipients  ")
        nb.add(self.tab_compose, text="  3. Compose & Send  ")
        nb.add(self.tab_collect, text="  4. Collect Responses  ")
        nb.add(self.tab_calendar, text="  5. Attendance & Payment  ")
        nb.add(self.tab_gift, text="  6. Gift Contribution  ")
        nb.add(self.tab_history, text="  7. Event History  ")

        self._build_tab_config()
        self._build_tab_recipients()
        self._build_tab_compose()
        self._build_tab_collect()
        self._build_tab_calendar()
        self._build_tab_gift()
        self._build_tab_history()

        # Event date/Start/End time giờ CHỈ nhập ở Tab 1 — Tab 5 chỉ hiển
        # thị lại (chỉ đọc) để tham khảo khi tạo Calendar Invite, và tab Gift
        # Contribution tự nạp lại danh sách người nhận mới nhất từ Tab 2 mỗi
        # lần mở — cả 2 việc này cần làm mới mỗi khi CHUYỂN SANG đúng tab đó.
        nb.bind("<<NotebookTabChanged>>", self._on_tab_changed)

        # Re-wraps long instructional labels (see _make_wrapping_label())
        # whenever the main app window itself is resized, so guidance text
        # stays fully readable instead of getting clipped / needing a
        # horizontal scrollbar when the window is narrowed.
        self._wrap_labels = []
        self.bind("<Configure>", self._on_root_resize)

    def _on_tab_changed(self, event):
        try:
            selected = event.widget.nametowidget(event.widget.select())
        except Exception:
            return
        if selected is self.tab_calendar:
            self._refresh_calendar_datetime_display()
            self._refresh_attendance_list()
            self._refresh_thankyou_body_display()
        elif selected is self.tab_gift:
            self._refresh_gift_contribution_list()

    # ── generic double-click-to-edit-a-cell helper, shared by the ──
    # ── Tab 5 "Attendance & Payment" table and the Tab 7 History table ──
    def _begin_cell_edit(self, tree, row_id, col_name, on_commit):
        """Opens a small Entry box directly on top of the given Treeview
        cell so the user can type a new value in place. Calls
        on_commit(row_id, col_name, new_value) when the edit is confirmed
        (Enter, or clicking away / losing focus). Escape cancels without
        saving. This does NOT write anything to disk by itself — each
        caller's on_commit decides what to do with the new value (e.g.
        update an in-memory roster dict, or just update the Treeview row)."""
        try:
            x, y, width, height = tree.bbox(row_id, col_name)
        except Exception:
            return
        if not width or not height:
            return  # cell is scrolled out of view — bbox() returns empty
        value = tree.set(row_id, col_name)
        entry = tk.Entry(tree)
        entry.insert(0, value)
        entry.select_range(0, "end")
        entry.focus()
        entry.place(x=x, y=y, width=width, height=height)

        done = {"value": False}

        def commit(event=None):
            if done["value"] or not entry.winfo_exists():
                return
            done["value"] = True
            new_value = entry.get()
            entry.destroy()
            on_commit(row_id, col_name, new_value)

        def cancel(event=None):
            done["value"] = True
            entry.destroy()

        entry.bind("<Return>", commit)
        entry.bind("<FocusOut>", commit)
        entry.bind("<Escape>", cancel)

    def _begin_cell_edit_combobox(self, tree, row_id, col_name, values, on_commit):
        """Same idea as _begin_cell_edit(), but opens a READONLY
        ttk.Combobox restricted to `values` instead of a free-text Entry —
        used for the "Actual Attend" column on Tab 5's Attendance &
        Payment table, so it's a proper Yes/No dropdown rather than free
        text. Commits as soon as a value is picked (<<ComboboxSelected>>),
        or on Enter/losing focus; Escape cancels without saving."""
        try:
            x, y, width, height = tree.bbox(row_id, col_name)
        except Exception:
            return
        if not width or not height:
            return  # cell is scrolled out of view — bbox() returns empty
        value = tree.set(row_id, col_name)
        combo = ttk.Combobox(tree, values=list(values), state="readonly")
        if value in values:
            combo.set(value)
        combo.place(x=x, y=y, width=width, height=height)
        combo.focus()

        done = {"value": False}

        def commit(event=None):
            if done["value"] or not combo.winfo_exists():
                return
            done["value"] = True
            new_value = combo.get()
            combo.destroy()
            on_commit(row_id, col_name, new_value)

        def cancel(event=None):
            done["value"] = True
            combo.destroy()

        combo.bind("<<ComboboxSelected>>", commit)
        combo.bind("<Return>", commit)
        combo.bind("<FocusOut>", commit)
        combo.bind("<Escape>", cancel)

    def _on_editable_tree_double_click(self, tree, event, editable_cols, on_commit):
        """Double-click handler: figures out which cell was double-clicked,
        checks it's one of editable_cols (a set of column names — pass None
        to allow editing ANY column), then opens an inline edit box via
        _begin_cell_edit(). Bind with:
            tree.bind("<Double-1>", lambda e: self._on_editable_tree_double_click(
                tree, e, {"name", "amount"}, self._my_commit_handler))"""
        region = tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        col = tree.identify_column(event.x)  # e.g. '#1', '#2', ...
        row_id = tree.identify_row(event.y)
        if not row_id or not col:
            return
        columns = tree["columns"]
        try:
            col_index = int(col.replace("#", "")) - 1
        except ValueError:
            return
        if col_index < 0 or col_index >= len(columns):
            return
        col_name = columns[col_index]
        if editable_cols is not None and col_name not in editable_cols:
            return
        self._begin_cell_edit(tree, row_id, col_name, on_commit)

    def _make_wrapping_label(self, parent, margin=60, **kwargs):
        """Creates a ttk.Label that automatically re-wraps its text (via
        wraplength) whenever ITS OWN TAB is resized — used for long
        instructional/help text that would otherwise get clipped, or
        trigger the horizontal scrollbar, when the window is narrowed.
        `margin` is roughly how much horizontal space (in pixels) to
        reserve for padding/indentation around the label — bump it up a
        bit for labels nested further to the right. Usage is a drop-in
        replacement for ttk.Label(...) — chain .grid(...)/.pack(...) on
        the result exactly the same way:
            self._make_wrapping_label(f, text="...", font=(...)).grid(row=r, ...)

        BUG ĐÃ SỬA: bản trước tính wraplength dựa theo self.winfo_width()
        (chiều rộng của CỬA SỔ GỐC toàn app) — con số này luôn RỘNG HƠN
        vùng nội dung THẬT SỰ nhìn thấy được bên trong 1 tab cụ thể (mất đi
        vì thanh cuộn dọc, padding của Notebook, indent riêng của từng
        frame con lồng nhau...), nên nhiều label không co lại đúng khi cửa
        sổ bị thu nhỏ. Giờ dò ngược lên cây widget cha để tìm đúng
        tk.Canvas của ScrollableFrame đang bọc tab chứa label này, và đăng
        ký vào DANH SÁCH RIÊNG của canvas đó (canvas.wrap_labels, xem
        ScrollableFrame.__init__) — mỗi khi CHÍNH canvas đó đổi kích thước
        (khớp chính xác vùng hiển thị thật), wraplength được tính lại theo
        đúng con số đó, không qua trung gian nào khác."""
        lbl = ttk.Label(parent, **kwargs)
        w = parent
        canvas = None
        while w is not None:
            if isinstance(w, tk.Canvas):
                canvas = w
                break
            w = w.master
        if canvas is not None:
            if not hasattr(canvas, "wrap_labels"):
                canvas.wrap_labels = []
            canvas.wrap_labels.append((lbl, margin))
            try:
                current_width = canvas.winfo_width()
                if current_width > 1:
                    lbl.configure(wraplength=max(current_width - margin, 250))
            except Exception:
                pass
        else:
            # Trường hợp hiếm: label không nằm trong ScrollableFrame nào
            # (không tìm thấy Canvas tổ tiên) — giữ lại cơ chế CŨ dựa theo
            # cửa sổ gốc làm phương án dự phòng, để label luôn có ít nhất
            # 1 cơ chế wraplength thay vì hoàn toàn không có.
            if not hasattr(self, "_wrap_labels"):
                self._wrap_labels = []
            self._wrap_labels.append((lbl, margin))
            try:
                current_width = self.winfo_width()
                if current_width > 1:
                    lbl.configure(wraplength=max(current_width - margin, 250))
            except Exception:
                pass
        return lbl

    def _on_root_resize(self, event):
        """Bound to the main window's <Configure> — CHỈ còn dùng làm
        phương án dự phòng cho các label hiếm gặp không nằm trong
        ScrollableFrame nào (xem nhánh else của _make_wrapping_label() ở
        trên) — số lượng label trong self._wrap_labels giờ thường RỖNG, vì
        phần lớn đã chuyển sang cơ chế theo canvas riêng của từng tab."""
        if event.widget is not self:
            return  # ignore <Configure> events bubbling up from child widgets — only the root window resizing matters here
        for lbl, margin in list(getattr(self, "_wrap_labels", [])):
            try:
                lbl.configure(wraplength=max(event.width - margin, 250))
            except tk.TclError:
                pass  # widget was destroyed — safe to ignore

    def _enable_treeview_copy_paste(self, tree, on_commit=None, editable_cols=None):
        """Adds Ctrl+C / Ctrl+V clipboard support to a Treeview:
        - Ctrl+C copies the selected rows (or all rows, if none selected)
          as tab-separated text, one row per line, in column order —
          pastes cleanly straight into Excel.
        - Ctrl+V pastes tab-separated (or comma-separated) clipboard text
          — e.g. copied from Excel — back INTO the table, starting at the
          currently selected row and filling downward one Treeview row per
          pasted line. It never creates new rows, only fills existing
          ones; pasted columns beyond the table's column count are
          ignored, and if editable_cols is given, any column not in it is
          skipped so paste can't corrupt read-only columns like "No.".
          Each changed cell goes through on_commit(row_id, col_name,
          new_value) if given — so pasting triggers the exact same side
          effects as manually double-click-editing that cell (e.g.
          auto-filling Amount when "Actual Attend" is pasted as "Yes") —
          otherwise the Treeview cell is just updated directly."""
        def copy_selection(event=None):
            rows = tree.selection() or tree.get_children()
            lines = []
            for row_id in rows:
                values = tree.item(row_id, "values")
                lines.append("\t".join("" if v is None else str(v) for v in values))
            tree.clipboard_clear()
            tree.clipboard_append("\n".join(lines))
            return "break"

        def paste_clipboard(event=None):
            try:
                text = tree.clipboard_get()
            except Exception:
                return "break"
            all_rows = list(tree.get_children())
            if not all_rows:
                return "break"
            selected = tree.selection()
            start_row = selected[0] if selected else all_rows[0]
            start_index = all_rows.index(start_row) if start_row in all_rows else 0
            columns = tree["columns"]
            pasted_lines = [ln for ln in text.replace("\r\n", "\n").replace("\r", "\n").split("\n") if ln != ""]
            for offset, line in enumerate(pasted_lines):
                target_index = start_index + offset
                if target_index >= len(all_rows):
                    break  # only fills existing rows, never creates new ones
                row_id = all_rows[target_index]
                cells = line.split("\t") if "\t" in line else line.split(",")
                for col_index, new_value in enumerate(cells):
                    if col_index >= len(columns):
                        break
                    col_name = columns[col_index]
                    if editable_cols is not None and col_name not in editable_cols:
                        continue
                    if on_commit is not None:
                        on_commit(row_id, col_name, new_value.strip())
                    else:
                        tree.set(row_id, col_name, new_value.strip())
            return "break"

        tree.bind("<Control-c>", copy_selection)
        tree.bind("<Control-C>", copy_selection)
        tree.bind("<Control-v>", paste_clipboard)
        tree.bind("<Control-V>", paste_clipboard)

    # ══════════════════════════════════════════════════════════════════
    # TAB 1 — EVENT SETUP
    # ══════════════════════════════════════════════════════════════════
    def _build_tab_config(self):
        f = self.tab_config.body
        pad = {"padx": 10, "pady": 6}

        header_bar = ttk.Frame(f)
        header_bar.grid(row=0, column=0, columnspan=3, sticky="we", **pad)
        ttk.Label(header_bar, text="Event details (these fields auto-fill into the invite email)",
                  font=("Arial", 11, "bold")).pack(side="left")
        # MỚI: "Event mode" — chọn sự kiện này thuộc dạng nào:
        #  - "Event": chỉ RSVP bình thường (mặc định, hành vi y hệt trước giờ)
        #  - "Gift": CHỈ thông báo kêu gọi đóng góp mua quà tặng (Tab 3 có
        #    thêm mode "Send Gift Contribution Notice"), không cần RSVP
        #  - "Event + Gift": vừa RSVP vừa có kêu gọi đóng góp quà — dùng khi
        #    cùng lúc tổ chức tiệc (cần biết số người tham dự) VÀ muốn quyên
        #    góp mua quà tặng (vd tiệc farewell tặng quà người sắp nghỉ việc)
        # Bản thân "Event mode" chỉ là NHÃN LƯU Ý — không tự động khoá/ẩn
        # tính năng nào cả, bạn vẫn luôn có thể dùng bất kỳ Send mode nào ở
        # Tab 3 bất kể chọn gì ở đây; nó giúp bạn nhớ lại mục đích sự kiện
        # khi xem lại History sau này (cột "EventMode").
        ttk.Label(header_bar, text="   Event mode:", font=("Arial", 10, "bold")).pack(side="left")
        self.var_event_mode = tk.StringVar(value="Event")
        self.combo_event_mode = ttk.Combobox(
            header_bar, width=16, state="readonly", textvariable=self.var_event_mode,
            values=["Event", "Gift", "Event + Gift"],
        )
        self.combo_event_mode.pack(side="left", padx=(4, 0))

        self.var_event_name = tk.StringVar(value="Team Building Q3 2026")
        self.var_event_id = tk.StringVar(value="TB2026-Q3")
        # Mỗi khi Event ID đổi (gõ tay hoặc do Load setup/Register event...),
        # tự cập nhật banner trạng thái ở đầu Tab 4 — báo ngay nếu bảng kết
        # quả đang hiện là của 1 Event ID KHÁC (còn sót từ lần Scan trước).
        self.var_event_id.trace_add(
            "write", lambda *a: self._update_scan_status_banner() if hasattr(self, "lbl_scan_status") else None)
        self.var_location = tk.StringVar(value="")
        self.var_budget = tk.StringVar(value="")
        self.var_gift_budget = tk.StringVar(value="")
        self.var_organizer = tk.StringVar(value="")
        self.var_guest_of_honor = tk.StringVar(value="")

        r = 1
        ttk.Label(f, text="Event Name:").grid(row=r, column=0, sticky="w", **pad)
        ttk.Entry(f, textvariable=self.var_event_name, width=50).grid(row=r, column=1, sticky="w", **pad)
        r += 1
        ttk.Label(f, text="Event ID (no spaces/accents, must be unique):").grid(row=r, column=0, sticky="w", **pad)
        ttk.Entry(f, textvariable=self.var_event_id, width=50).grid(row=r, column=1, sticky="w", **pad)
        r += 1
        ttk.Label(f, text="Location:").grid(row=r, column=0, sticky="w", **pad)
        ttk.Entry(f, textvariable=self.var_location, width=50).grid(row=r, column=1, sticky="w", **pad)
        r += 1
        # Đổi tên "Expected Budget" -> "Expected EVENT Budget" để phân biệt rõ
        # với "Expected GIFT Budget" mới thêm ngay bên dưới (2 khoản chi khác
        # nhau: ngân sách tổ chức sự kiện vs ngân sách mua quà tặng).
        ttk.Label(f, text="Expected event budget (e.g. \"5,000 JPY / person\"):").grid(row=r, column=0, sticky="w", **pad)
        ttk.Entry(f, textvariable=self.var_budget, width=50).grid(row=r, column=1, sticky="w", **pad)
        r += 1
        ttk.Label(f, text="Expected gift budget (e.g. \"3,000 JPY / person\"):").grid(row=r, column=0, sticky="w", **pad)
        ttk.Entry(f, textvariable=self.var_gift_budget, width=50).grid(row=r, column=1, sticky="w", **pad)
        r += 1

        ttk.Label(f, text="Event Date:").grid(row=r, column=0, sticky="w", **pad)
        self.date_event = make_date_picker(f, datetime.now() + timedelta(days=21))
        self.date_event.grid(row=r, column=1, sticky="w", **pad)
        r += 1

        # MỚI: Start time / End time CHUYỂN TỪ TAB 5 LÊN ĐÂY (ngay dưới Event
        # Date) — trước đây 2 ô này chỉ tồn tại ở Tab 5 (Attendance & Payment) và
        # KHÔNG được lưu vào History, nên "Load setup from selected event"
        # không nạp lại được, mỗi lần mở Tab 5 lại phải gõ tay lại giờ. Giờ
        # thuộc Tab 1 nên được lưu/nạp cùng các thông tin sự kiện khác, và
        # Tab 5 dùng lại TRỰC TIẾP 2 biến này (self.var_start_time/
        # self.var_end_time) khi tạo Calendar Invite — không còn ô riêng ở
        # Tab 5 nữa.
        ttk.Label(f, text="Start time (HH:MM):").grid(row=r, column=0, sticky="w", **pad)
        self.var_start_time = tk.StringVar(value="18:00")
        ttk.Entry(f, textvariable=self.var_start_time, width=10).grid(row=r, column=1, sticky="w", **pad)
        r += 1
        ttk.Label(f, text="End time (HH:MM):").grid(row=r, column=0, sticky="w", **pad)
        self.var_end_time = tk.StringVar(value="21:00")
        ttk.Entry(f, textvariable=self.var_end_time, width=10).grid(row=r, column=1, sticky="w", **pad)
        r += 1

        ttk.Label(f, text="Event response deadline:").grid(row=r, column=0, sticky="w", **pad)
        self.date_deadline = make_date_picker(f, datetime.now() + timedelta(days=10))
        self.date_deadline.grid(row=r, column=1, sticky="w", **pad)
        r += 1

        # MỚI: "Gift contribution deadline" — hạn RIÊNG để mọi người đóng
        # góp tiền mua quà, khác với "Event response deadline" (hạn RSVP có
        # tham dự hay không) ở trên — 2 việc này thường không cùng ngày (vd
        # cần chốt RSVP sớm hơn để chuẩn bị tiệc, nhưng có thể gia hạn thời
        # gian đóng góp quà lâu hơn 1 chút). Dùng cho email "Send Gift
        # Contribution Notice" ở Tab 3 (xem build_gift_fixed_block()).
        ttk.Label(f, text="Gift contribution deadline:").grid(row=r, column=0, sticky="w", **pad)
        self.date_gift_deadline = make_date_picker(f, datetime.now() + timedelta(days=10))
        self.date_gift_deadline.grid(row=r, column=1, sticky="w", **pad)
        r += 1

        # MỚI: Organizer / Guest of Honor — dùng cho email "Send Gift
        # Contribution Notice" ở Tab 3 (Organizer = người nhận đóng góp,
        # Guest of Honor = người được tặng quà, vd người sắp nghỉ việc) —
        # nhưng vẫn hiển thị luôn ở Event mode "Event" (không bị ẩn) vì đôi
        # khi hữu ích để ghi chú ai là người tổ chức/nhân vật chính dù
        # không quyên góp quà, không bắt buộc phải điền.
        ttk.Label(f, text="Organizer (contact person for gift contributions):").grid(row=r, column=0, sticky="w", **pad)
        ttk.Entry(f, textvariable=self.var_organizer, width=50).grid(row=r, column=1, sticky="w", **pad)
        r += 1
        ttk.Label(f, text="Guest of Honor (who the gift is for):").grid(row=r, column=0, sticky="w", **pad)
        ttk.Entry(f, textvariable=self.var_guest_of_honor, width=50).grid(row=r, column=1, sticky="w", **pad)
        r += 1

        ttk.Label(f, text="Organizer note (free text — you can write in any language;\n"
                          "translate it per-language on the Compose tab):").grid(row=r, column=0, sticky="nw", **pad)
        self.entry_note = tk.Text(f, width=55, height=3)
        self.entry_note.insert("1.0", "Please respond before the deadline so we can prepare accurate numbers.")
        self.entry_note.grid(row=r, column=1, sticky="w", **pad)
        r += 1

        ttk.Button(f, text="📝 Register event (save to History + reset form for a NEW event)",
                   command=self._register_event).grid(row=r, column=0, columnspan=2, sticky="w", **pad)
        r += 1
        ttk.Label(f, text="→ Saves as a NEW event in History, then clears this form for the next one.",
                  font=("Arial", 8, "italic")).grid(row=r, column=0, columnspan=2, sticky="w", padx=4)
        r += 1

        ttk.Separator(f).grid(row=r, column=0, columnspan=3, sticky="ew", pady=10)
        r += 1

        ttk.Label(f, text="Reload setup from a past event saved in history:",
                  font=("Arial", 10, "italic")).grid(row=r, column=0, columnspan=2, sticky="w", **pad)
        r += 1
        self.combo_load_history = ttk.Combobox(f, width=50, state="readonly")
        self.combo_load_history.grid(row=r, column=1, sticky="w", **pad)
        ttk.Button(f, text="🔄 Refresh past-event list", command=self._refresh_history_combo)\
            .grid(row=r, column=0, sticky="w", **pad)
        r += 1
        ttk.Button(f, text="⬅ Load setup from selected event", command=self._load_from_history)\
            .grid(row=r, column=1, sticky="w", **pad)
        r += 1

        ttk.Separator(f).grid(row=r, column=0, columnspan=3, sticky="ew", pady=10)
        r += 1
        self._make_wrapping_label(f, text="Already sent invites for this Event ID, but the details above changed since "
                          "(location/date/deadline/budget/note)? Push the update into History directly —\n"
                          "no email needs to be sent for this. (To also NOTIFY people by email, use "
                          "'Send update invite' mode on Tab 3 instead — that sends AND updates History.)",
                  font=("Arial", 9, "italic")).grid(row=r, column=0, columnspan=3, sticky="w", **pad)
        r += 1
        ttk.Button(f, text="💾 Update this event in History (no email sent)", command=self._update_history_from_tab1)\
            .grid(row=r, column=0, columnspan=2, sticky="w", **pad)

        self._refresh_history_combo()

    def _register_event(self):
        """Đăng ký sự kiện ĐANG NHẬP (Tab 1) vào RSVP_History.xlsx như 1 sự
        kiện — rồi RESET TOÀN BỘ form (Tab 1-5) về mặc định để bắt đầu nhập
        1 sự kiện MỚI KHÁC, sạch sẽ, không còn dính tên/địa điểm/danh sách
        người nhận/nội dung đã soạn của sự kiện vừa đăng ký. Sự kiện vừa
        đăng ký vẫn còn nguyên trong History — nạp lại bất cứ lúc nào qua
        '⬅ Load setup from selected event'."""
        event_id = self.var_event_id.get().strip()
        if not event_id:
            messagebox.showwarning("Missing Event ID", "Enter an Event ID before registering the event.")
            return

        if not messagebox.askyesno(
            "Register new event",
            f"Register event '{event_id}' to the database?\n\n"
            "After registering, the WHOLE form (Tab 1 → 5: event info, recipient list, "
            "composed content, scanned results...) will be CLEARED so you can start entering "
            "a NEW event. The event you just registered is still safely stored in History, "
            "and can be reloaded any time via '⬅ Load setup from selected event'."
        ):
            return

        if self.recipients:
            self._save_recipients_to_db(silent=True)  # chuẩn hoá tên + lưu file người nhận trước khi đăng ký

        record = {
            "EventID": event_id,
            "EventName": self.var_event_name.get(),
            "EventDate": get_date_str(self.date_event),
            "Deadline": get_date_str(self.date_deadline),
            "Location": self.var_location.get(),
            "Budget": self.var_budget.get(),
            "OrganizerNote": self._source_note_text(),
            "EventMode": self.var_event_mode.get(),
            "Organizer": self.var_organizer.get(),
            "GuestOfHonor": self.var_guest_of_honor.get(),
            "GiftBudget": self.var_gift_budget.get(),
            "GiftDeadline": get_date_str(self.date_gift_deadline),
            "StartTime": self.var_start_time.get(),
            "EndTime": self.var_end_time.get(),
        }
        if self.recipient_file.get():
            record["RecipientFile"] = self.recipient_file.get()

        try:
            db.save_event_record(record, self.history_path.get())
        except Exception as e:
            messagebox.showerror("Error", f"Couldn't register the event:\n{e}")
            return

        self._refresh_history_tree()
        self._refresh_history_combo()
        self._reset_for_new_event()
        messagebox.showinfo(
            "Registered",
            f"Event '{event_id}' registered to History.\n\n"
            "The form has been reset to defaults — ready to enter a new event."
        )

    def _reset_for_new_event(self):
        """Đưa TOÀN BỘ form về trạng thái mặc định — như vừa mở app lần đầu
        — để bắt đầu nhập 1 sự kiện HOÀN TOÀN MỚI. KHÔNG xoá các tuỳ chỉnh
        mang tính CÁ NHÂN/lâu dài (vd self.fixed_overrides — 'câu văn mặc
        định' bạn đã lưu để dùng cho MỌI sự kiện — và
        fixed_wording_overrides.json/prompt_overrides.json nói chung), chỉ
        xoá dữ liệu RIÊNG của sự kiện vừa đăng ký."""
        # Tab 1
        self.var_event_name.set("")
        self.var_event_id.set("")
        self.var_location.set("")
        self.var_budget.set("")
        self.var_gift_budget.set("")
        self.var_organizer.set("")
        self.var_guest_of_honor.set("")
        if hasattr(self, "combo_event_mode"):
            self.combo_event_mode.current(0)  # về lại "Event"
        set_date_str(self.date_event, (datetime.now() + timedelta(days=21)).strftime("%d/%m/%Y"))
        set_date_str(self.date_deadline, (datetime.now() + timedelta(days=10)).strftime("%d/%m/%Y"))
        set_date_str(self.date_gift_deadline, (datetime.now() + timedelta(days=10)).strftime("%d/%m/%Y"))
        # Start/End time nay thuộc Tab 1 (ngay dưới Event Date) — reset về mặc định ở đây.
        self.var_start_time.set("18:00")
        self.var_end_time.set("21:00")
        self.entry_note.delete("1.0", "end")
        self.entry_note.insert("1.0", "Please respond before the deadline so we can prepare accurate numbers.")

        # Tab 2 — danh sách người nhận thuộc về sự kiện CŨ, không mang sang sự kiện mới
        self.recipients = []
        self.recipient_file.set("")
        self._refresh_recipient_tree()
        if hasattr(self, "_group_expansion_cache"):
            self._group_expansion_cache = {}

        # Tab 3 — bản dịch đầy đủ đã lưu (full_translations) là nội dung CỦA
        # sự kiện cũ, không tái sử dụng cho sự kiện mới; fixed_overrides (câu
        # văn mặc định cá nhân) KHÔNG bị đụng vì đó là tuỳ chọn lâu dài.
        self.full_translations = {"en": "", "ja": "", "vi": "", "bilingual": ""}
        self.gift_full_translations = {"en": "", "ja": "", "vi": "", "bilingual": ""}
        self.combo_send_mode.current(0)  # về lại "Send first Invite"
        self._refresh_send_button_label()
        self.txt_translation_paste.delete("1.0", "end")
        self.var_send_to_override.set("")
        self.var_auto_send.set(False)

        # Tab 4
        self.responses = {}
        self._pending_recipients = []
        self._last_scanned_event_id = None
        self._last_scan_time = None
        self.tree_responses.delete(*self.tree_responses.get_children())
        self.tree_pending.delete(*self.tree_pending.get_children())
        self.lbl_summary.config(text="No responses scanned yet.")
        self.lbl_deadline_banner.config(text="")
        self.txt_reminder_body.delete("1.0", "end")

        # Tab 5
        self._yes_emails = []
        self.list_yes.delete(0, "end")
        if hasattr(self, "combo_calendar_lang"):
            self.combo_calendar_lang.current(0)  # về lại English
        self.var_appt_body_default = build_calendar_body("en", *self._calendar_body_args())
        self.txt_appt_body.delete("1.0", "end")
        self.txt_appt_body.insert("1.0", self.var_appt_body_default)

        # Tab 6 — Gift Contribution — thuộc về sự kiện CŨ, không mang sang sự kiện mới
        self._gift_roster = {}
        if hasattr(self, "var_gift_search"):
            self.var_gift_search.set("")
        if hasattr(self, "tree_gift"):
            self.tree_gift.delete(*self.tree_gift.get_children())
            self._update_gift_contributed_count()

        self._refresh_compose_preview()
        self._update_scan_status_banner()

    def _update_history_from_tab1(self):
        """Ghi các trường Tab 1 hiện tại (Location/EventDate/Deadline/Budget/
        Note) — và cả RecipientFile nếu Tab 2 đang có danh sách — vào dòng
        History khớp Event ID, mà KHÔNG cần gửi email gì cả. Dùng khi bạn chỉ
        cần sửa lại thông tin đã lưu (vd sửa nhầm địa điểm) mà không cần báo
        cho người nhận. Nếu Event ID chưa từng có trong History, sẽ tạo dòng
        mới (giống hành vi bình thường của History)."""
        event_id = self.var_event_id.get().strip()
        if not event_id:
            messagebox.showwarning(
                "Missing Event ID",
                "Enter an Event ID first (used to identify which row in History to update).")
            return

        # Chuẩn hoá + lưu lại file người nhận (nếu Tab 2 đang có danh sách)
        # TRƯỚC khi ghi vào History — cùng cơ chế với Tab 3 Send / Tab 4 Save,
        # đảm bảo cột RecipientFile trong History luôn trỏ đúng file mới nhất
        # thay vì tên file gốc bạn Browse vào (vd file TEMPLATE dùng chung).
        if self.recipients:
            self._save_recipients_to_db(silent=True)

        record = {
            "EventID": event_id,
            "EventName": self.var_event_name.get(),
            "EventDate": get_date_str(self.date_event),
            "Deadline": get_date_str(self.date_deadline),
            "Location": self.var_location.get(),
            "Budget": self.var_budget.get(),
            "OrganizerNote": self._source_note_text(),
            "EventMode": self.var_event_mode.get(),
            "Organizer": self.var_organizer.get(),
            "GuestOfHonor": self.var_guest_of_honor.get(),
            "GiftBudget": self.var_gift_budget.get(),
            "GiftDeadline": get_date_str(self.date_gift_deadline),
            "StartTime": self.var_start_time.get(),
            "EndTime": self.var_end_time.get(),
        }
        if self.recipient_file.get():
            record["RecipientFile"] = self.recipient_file.get()

        try:
            db.save_event_record(record, self.history_path.get())
        except Exception as e:
            messagebox.showerror("Error", f"Couldn't update History:\n{e}")
            return

        self._refresh_history_tree()
        self._refresh_history_combo()
        extra = "\n(including the latest RecipientFile)" if self.recipient_file.get() else ""
        messagebox.showinfo(
            "History updated",
            f"Updated Location / Event Date / Deadline / Budget / Note for Event ID "
            f"'{event_id}' in History{extra}.\n\n"
            "Other columns (SentDate, Yes/No/Maybe, CalendarSent, etc.) were KEPT AS-IS — "
            "nothing was cleared or overwritten.\n\n"
            "⚠️ This does NOT send any email to recipients — if you want to notify them of the "
            "change, use 'Send update invite' mode on Tab 3 instead."
        )

    def _refresh_history_combo(self):
        try:
            records = db.load_history(self.history_path.get())
        except Exception:
            records = []
        self._history_records = records
        labels = [f'{r["EventID"]} — {r["EventName"]}' for r in records]
        self.combo_load_history["values"] = labels
        if labels:
            self.combo_load_history.current(len(labels) - 1)

    def _load_from_history(self):
        idx = self.combo_load_history.current()
        if idx < 0 or idx >= len(self._history_records):
            messagebox.showinfo("Nothing selected", "Please select a past event from the list first.")
            return

        # BUG ĐÃ SỬA: trước đây dùng thẳng self._history_records[idx] — 1 bản
        # CACHE chỉ được nạp lúc mở app / lúc bấm '🔄 Refresh past-event
        # list'. Nếu History đã được cập nhật ở nơi khác trong CÙNG phiên
        # làm việc (vd Tab 2 '📥 Load list' hoặc '🗂 Update RecipientFile in
        # History' vừa ghi RecipientFile mới) mà bạn chưa bấm Refresh lại ở
        # Tab 1, cache này KHÔNG tự biết — nên Load setup vẫn dùng dữ liệu
        # CŨ (vd RecipientFile trỏ về file TEMPLATE gốc), dù bản thân file
        # RSVP_History.xlsx trên đĩa đã đúng từ lâu — đúng như bạn gặp phải.
        # Giờ đọc lại TRỰC TIẾP từ đĩa theo đúng Event ID vừa chọn, luôn đảm
        # bảo dùng dữ liệu MỚI NHẤT bất kể có bấm Refresh trước đó hay không.
        stale_rec = self._history_records[idx]
        event_id_to_load = stale_rec.get("EventID")
        try:
            fresh_records = db.load_history(self.history_path.get())
        except Exception:
            fresh_records = self._history_records
        matches = [r for r in fresh_records if r.get("EventID") == event_id_to_load]
        rec = matches[0] if matches else stale_rec
        self._history_records = fresh_records  # đồng bộ luôn cache, để lần Refresh sau nhất quán

        # BUG ĐÃ SỬA: trước đây hàm này KHÔNG nạp EventID / EventDate / Deadline
        # — 3 ô đó vẫn giữ nguyên giá trị cũ còn sót lại trên form (vd: giá trị
        # mặc định "TB2026-Q3" lúc khởi động app), không hề khớp với sự kiện vừa
        # chọn. Hậu quả: Event ID trên form bị lệch với Event ID thật đã dùng khi
        # gửi mail, nên Tab 4 (Collect Responses) tìm theo Event ID sai và không
        # ra kết quả nào. Giờ nạp đủ cả 3 ô này.
        self.var_event_id.set(rec.get("EventID") or "")
        self.var_event_name.set(rec.get("EventName") or "")
        self.var_location.set(rec.get("Location") or "")
        self.var_budget.set(rec.get("Budget") or "")
        set_date_str(self.date_event, rec.get("EventDate"))
        set_date_str(self.date_deadline, rec.get("Deadline"))
        self.entry_note.delete("1.0", "end")
        self.entry_note.insert("1.0", rec.get("OrganizerNote") or "")

        # MỚI: nạp thêm 7 trường của tính năng "Event mode" (Event/Gift/
        # Event + Gift) — EventMode/Organizer/GuestOfHonor/GiftBudget/
        # GiftDeadline đã thêm ở Tab 1, và StartTime/EndTime (chuyển từ
        # Tab 5 lên Tab 1). Dùng .get(key) or "" / or default để an toàn
        # với các dòng History CŨ được tạo TRƯỚC khi các cột này tồn tại
        # (giá trị sẽ là None sau khi tự động migrate — xem
        # history._migrate_history_columns()). GiftDeadline không có sẵn ở
        # các dòng cũ -> mặc định về CÙNG giá trị với Deadline (RSVP), thay
        # vì để trống khó hiểu.
        event_mode = rec.get("EventMode") or "Event"
        if hasattr(self, "combo_event_mode") and event_mode in self.combo_event_mode["values"]:
            self.var_event_mode.set(event_mode)
        self.var_organizer.set(rec.get("Organizer") or "")
        self.var_guest_of_honor.set(rec.get("GuestOfHonor") or "")
        self.var_gift_budget.set(rec.get("GiftBudget") or "")
        self.var_start_time.set(rec.get("StartTime") or "18:00")
        self.var_end_time.set(rec.get("EndTime") or "21:00")
        set_date_str(self.date_gift_deadline, rec.get("GiftDeadline") or rec.get("Deadline"))

        # MỚI: nạp lại "Amount paid" (Tab 5) đã lưu cho sự kiện này — cột
        # "AmountPaid" có thể trống với các sự kiện CŨ tạo trước khi tính
        # năng này tồn tại (giá trị None sau ALTER TABLE, xem db.py), mặc
        # định về "0" cho an toàn thay vì hiện chữ "None".
        if hasattr(self, "var_amount_paid"):
            self.var_amount_paid.set(rec.get("AmountPaid") or "0")
            self._refresh_remaining_amount()

        # MỚI: đổi từ đọc file Excel (Participant_List_{EventID}.xlsx) sang
        # đọc TRỰC TIẾP từ database — Tab 2 giờ lưu trữ trong DB, file Excel
        # chỉ còn được TẠO RA khi bấm "Export to Excel" trên Tab 2 (không
        # còn là nguồn dữ liệu chính, xem db.load_recipients()).
        recipient_note = ""
        loaded_people = db.load_recipients(event_id_to_load, self.history_path.get())
        if loaded_people:
            self.recipients = loaded_people
            self._refresh_recipient_tree()
            recipient_note = f"\n\n📋 Auto-loaded {len(self.recipients)} people into Tab 2 from the database."
        else:
            recipient_note = "\n\n(This event has no recipient list saved in the database yet.)"

        # MỚI: khôi phục kết quả Responded (Tab 4) và bảng Attendance &
        # Payment (Tab 5) đã lưu trước đó cho sự kiện này, từ sheet
        # "Responded result" / "Attendance & Payment" trong
        # Attendance_Payment_{EventID}.xlsx — KHÔNG tự động quét lại
        # Outlook. Chỉ khi bấm '📨 Scan Inbox for Vote results' ở Tab 4 mới
        # thực sự quét lại và lấy kết quả MỚI NHẤT (xem _collect_responses()).
        responded_loaded = self._load_responded_result_from_file(event_id_to_load)
        attendance_loaded = self._load_attendance_sheet_from_file(event_id_to_load)
        gift_loaded = self._load_gift_roster_from_db(event_id_to_load)
        restore_note = ""
        if responded_loaded:
            restore_note += ("\n\n📨 Restored the last-scanned Responded results into Tab 4 "
                              "(no live Outlook scan was performed — click "
                              "'📨 Scan Inbox for Vote results' there to refresh).")
        if attendance_loaded:
            restore_note += "\n\n📊 Restored the Attendance & Payment tracking table into Tab 5."
        if gift_loaded:
            restore_note += "\n\n🎁 Restored the Gift Contribution tracking table into Tab 6."

        messagebox.showinfo(
            "Loaded",
            "Event ID / Event Name / Location / Budget / Event Date / Deadline / Note "
            "have all been loaded from the past event."
            + recipient_note + restore_note +
            "\n\n⚠️ If you're creating a NEW event (using this one as a template): remember to "
            "CHANGE the Event ID (and dates) to new values before sending the invite, to avoid "
            "clashing with the old event — and on Tab 2, click '💾 Save to Excel' again so the "
            "recipient list is saved separately under the new Event ID (avoiding overwriting the "
            "old event's file).\n\n"
            "If you're reloading THIS SAME event just to Collect Responses / Send Calendar "
            "Invite, you can leave the Event ID as loaded."
        )

    # ══════════════════════════════════════════════════════════════════
    # TAB 2 — RECIPIENTS
    # ══════════════════════════════════════════════════════════════════
    def _standardized_recipient_path(self):
        """Trả về đường dẫn CHUẨN HOÁ cho file danh sách người nhận của sự
        kiện hiện tại: 'Participant_List_{EventID}.xlsx', nằm CÙNG THƯ MỤC
        với file đang trỏ tới ở ô 'Recipient Excel file' (hoặc thư mục hiện
        tại nếu ô đó đang trống). Trả về None nếu Event ID (Tab 1) đang
        trống — không đủ thông tin để đặt tên chuẩn.

        Việc chuẩn hoá tên file này (thay vì giữ nguyên tên bạn Browse vào,
        vd 'Danh_sach_nguoi_tham_gia_TEMPLATE.xlsx') giúp Tab 1 '⬅ Load setup
        from selected event' sau này tự tìm và nạp lại ĐÚNG file người nhận
        cho sự kiện đó, vì tên file luôn suy ra được trực tiếp từ Event ID
        (không phụ thuộc bạn đặt tên gốc là gì)."""
        event_id = self.var_event_id.get().strip()
        if not event_id:
            return None
        current = self.recipient_file.get().strip()
        folder = os.path.dirname(os.path.abspath(current)) if current else os.getcwd()
        return os.path.join(folder, f"Participant_List_{event_id}.xlsx")

    def _build_tab_recipients(self):
        f = self.tab_recipients.body
        top = ttk.Frame(f)
        top.pack(fill="x", padx=10, pady=8)

        # MỚI: kiến trúc lưu trữ đã đổi sang DB — danh sách người nhận giờ
        # tự động lưu vào database mỗi khi thay đổi (thêm/xoá dòng, Load
        # list, Expand group...), KHÔNG còn tự ghi ra Excel liên tục nữa.
        # "Recipient Excel file" ở đây giờ CHỈ dùng làm nguồn NHẬP (import)
        # ban đầu — vd nạp từ file mẫu Danh_sach_nguoi_tham_gia_TEMPLATE —
        # không phải nơi lưu trữ chính. Muốn có file Excel để gửi/báo cáo,
        # dùng nút "📊 Export to Excel".
        ttk.Label(top, text="Import from Excel file:").pack(side="left")
        ttk.Entry(top, textvariable=self.recipient_file, width=45).pack(side="left", padx=6)
        ttk.Button(top, text="📂 Browse...", command=self._browse_recipient_file).pack(side="left", padx=4)
        ttk.Button(top, text="📥 Import list", command=self._load_recipients).pack(side="left", padx=4)
        ttk.Button(top, text="📊 Export to Excel", command=self._export_recipients_to_excel).pack(side="left", padx=4)

        expand_bar = ttk.Frame(f)
        expand_bar.pack(fill="x", padx=10, pady=(0, 6))
        ttk.Button(expand_bar, text="🔎 Expand group emails in list (incl. sub-groups)",
                   command=self._expand_group_recipients).pack(side="left")
        self._make_wrapping_label(expand_bar, text="  → Any row that is a company Distribution List (group email, "
                                   "e.g. 'EET Employees All') gets replaced with its REAL individual "
                                   "members, so Tab 4 can track exactly who hasn't responded yet.",
                  font=("Arial", 8, "italic")).pack(side="left")

        history_push_bar = ttk.Frame(f)
        history_push_bar.pack(fill="x", padx=10, pady=(0, 6))
        ttk.Button(history_push_bar, text="💾 Save recipients now",
                   command=lambda: self._save_recipients_to_db(silent=False)).pack(side="left")
        self._make_wrapping_label(history_push_bar, text="  → Manually save this list to the database for the current "
                                          "Event ID (Tab 1) right away. This already happens automatically "
                                          "after every change (Import/Add/Delete row/Expand group) — use "
                                          "this button only if you want to be sure right now.",
                  font=("Arial", 8, "italic")).pack(side="left")

        # MỚI: ô tìm kiếm nhanh theo Name/Email — lọc TRỰC TIẾP bảng bên dưới
        # khi gõ (không cần bấm nút), khớp theo dạng "chứa chuỗi" (không phân
        # biệt hoa/thường), tìm trên CẢ Name lẫn Email cùng lúc.
        search_bar = ttk.Frame(f)
        search_bar.pack(fill="x", padx=10, pady=(0, 6))
        ttk.Label(search_bar, text="🔎 Search (name or email):").pack(side="left")
        self.var_recipient_search = tk.StringVar(value="")
        search_entry = ttk.Entry(search_bar, textvariable=self.var_recipient_search, width=40)
        search_entry.pack(side="left", padx=6)
        self.var_recipient_search.trace_add("write", lambda *a: self._apply_recipient_filter())
        ttk.Button(search_bar, text="✕ Clear", command=lambda: self.var_recipient_search.set(""))\
            .pack(side="left", padx=4)

        cols = ("name", "email")
        tree_container, self.tree_recipients = make_scrollable_treeview(f, columns=cols, height=18)
        self.tree_recipients.heading("name", text="Name")
        self.tree_recipients.heading("email", text="Email")
        self.tree_recipients.column("name", width=280)
        self.tree_recipients.column("email", width=320)
        tree_container.pack(fill="both", expand=True, padx=10, pady=6)

        edit_bar = ttk.Frame(f)
        edit_bar.pack(fill="x", padx=10, pady=6)
        self.var_new_name = tk.StringVar()
        self.var_new_email = tk.StringVar()
        ttk.Entry(edit_bar, textvariable=self.var_new_name, width=28).pack(side="left", padx=3)
        ttk.Entry(edit_bar, textvariable=self.var_new_email, width=28).pack(side="left", padx=3)
        ttk.Button(edit_bar, text="➕ Add person", command=self._add_recipient_row).pack(side="left", padx=6)
        ttk.Button(edit_bar, text="🗑 Delete selected row", command=self._delete_recipient_row).pack(side="left", padx=6)

        self.lbl_recipient_count = ttk.Label(f, text="No list loaded yet.")
        self.lbl_recipient_count.pack(anchor="w", padx=10, pady=4)

    def _browse_recipient_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.recipient_file.set(path)

    def _load_recipients(self, silent=False):
        path = self.recipient_file.get()
        if not os.path.exists(path):
            if not silent:
                messagebox.showerror("Error", f"File not found: {path}")
            return False
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb["DanhSach"] if "DanhSach" in wb.sheetnames else wb.active
        self.recipients = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            name, email = (row + (None, None))[:2]
            if not email or "@" not in str(email):
                continue
            if "(ví dụ)" in str(name or "") or "(example)" in str(name or "").lower():
                continue
            self.recipients.append((str(name or "").strip(), str(email).strip()))
        self._refresh_recipient_tree()

        # BUG ĐÃ SỬA (kiến trúc CŨ): trước đây bước này chỉ CHUẨN HOÁ TÊN FILE
        # trên đĩa rồi ghi đường dẫn đó vào RSVP_History.xlsx. GIỜ ĐÃ ĐỔI
        # SANG DB: import xong là lưu THẲNG self.recipients vào database
        # (bảng recipients, khớp theo Event ID) — không còn phụ thuộc vào
        # việc ghi/đọc lại 1 file Excel trung gian nữa.
        self._save_recipients_to_db(silent=True)
        return True

    def _save_recipients_to_db(self, silent=False):
        """Lưu self.recipients (danh sách người nhận đang có trên Tab 2)
        vào database cho Event ID hiện tại — đây là hàm PERSIST THẬT SỰ
        trong kiến trúc mới (thay cho việc ghi ra Excel trước đây). Gọi
        NGAY sau mọi thay đổi (Import/Add/Delete row/Expand group), và
        trước khi đăng ký/gửi sự kiện, để đảm bảo dữ liệu trong DB luôn
        khớp với những gì đang hiển thị trên Tab 2."""
        event_id = self.var_event_id.get().strip()
        if not event_id:
            if not silent:
                messagebox.showwarning(
                    "Missing Event ID",
                    "Enter an Event ID on Tab 1 first — it's needed to know which event "
                    "this recipient list belongs to in the database."
                )
            return False
        if not self.recipients:
            if not silent:
                messagebox.showwarning("No list yet", "There are no recipients to save yet.")
            return False
        try:
            db.save_recipients(event_id, self.recipients, self.history_path.get())
        except Exception as e:
            if not silent:
                messagebox.showerror("Error", f"Couldn't save recipients to the database:\n{e}")
            return False
        if not silent:
            messagebox.showinfo(
                "Saved",
                f"Saved {len(self.recipients)} people to the database for Event ID '{event_id}'."
            )
        return True

    def _refresh_recipient_tree(self):
        # Giờ chỉ là 1 lớp mỏng gọi _apply_recipient_filter() — hàm đó mới
        # thực sự đọc self.recipients + lọc theo ô search (nếu có), để mọi
        # nơi gọi _refresh_recipient_tree() (Load list, Add/Delete row,
        # Expand group, reset form...) đều tự động tôn trọng từ khoá đang
        # tìm thay vì hiện lại TOÀN BỘ danh sách và làm mất kết quả search.
        self._apply_recipient_filter()

    def _apply_recipient_filter(self, *args):
        query = self.var_recipient_search.get().strip().lower() if hasattr(self, "var_recipient_search") else ""
        self.tree_recipients.delete(*self.tree_recipients.get_children())
        matched = 0
        for name, email in self.recipients:
            if query and query not in (name or "").lower() and query not in (email or "").lower():
                continue
            self.tree_recipients.insert("", "end", values=(name, email))
            matched += 1
        total = len(self.recipients)
        if query:
            self.lbl_recipient_count.config(text=f"Showing {matched} / {total} people (search: '{self.var_recipient_search.get()}')")
        else:
            self.lbl_recipient_count.config(text=f"Total: {total} people")

    def _add_recipient_row(self):
        name, email = self.var_new_name.get().strip(), self.var_new_email.get().strip()
        if not email or "@" not in email:
            messagebox.showwarning("Missing email", "Enter a valid email before adding.")
            return
        self.recipients.append((name, email))
        self.var_new_name.set("")
        self.var_new_email.set("")
        self._refresh_recipient_tree()
        self._save_recipients_to_db(silent=True)

    def _delete_recipient_row(self):
        sel = self.tree_recipients.selection()
        if not sel:
            return
        for item_id in sel:
            values = self.tree_recipients.item(item_id, "values")
            self.recipients = [r for r in self.recipients if not (r[0] == values[0] and r[1] == values[1])]
        self._refresh_recipient_tree()
        self._save_recipients_to_db(silent=True)

    def _export_recipients_to_excel(self, silent=False):
        """Xuất self.recipients (danh sách người nhận đang có, đang lưu
        trong DB) ra 1 file Excel — CHỈ khi bấm nút "📊 Export to Excel",
        KHÔNG còn tự động chạy mỗi khi đổi dữ liệu như kiến trúc cũ. Dùng
        để gửi cho người khác hoặc lưu trữ ngoài app."""
        if not self.recipients:
            if not silent:
                messagebox.showwarning("No data", "There is no recipient list to export yet.")
            return
        std_path = self._standardized_recipient_path()
        old_path = self.recipient_file.get().strip()
        path = std_path or old_path
        if not path:
            path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx")])
            if not path:
                return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DanhSach"
        ws.cell(row=3, column=1, value="Họ tên").font = Font(bold=True)
        ws.cell(row=3, column=2, value="Email").font = Font(bold=True)
        for i, (name, email) in enumerate(self.recipients, start=4):
            ws.cell(row=i, column=1, value=name)
            ws.cell(row=i, column=2, value=email)
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 32
        wb.save(path)
        if not silent:
            messagebox.showinfo("Exported", f"Exported {len(self.recipients)} people to:\n{path}")

    def _expand_group_recipients(self):
        """Với mỗi dòng trong danh sách hiện tại, kiểm tra xem đó có phải 1
        group email (Exchange Distribution List) không — nếu phải, thay dòng
        đó bằng TẤT CẢ thành viên thật của group (kể cả sub-group lồng bên
        trong, xem outlook_com.expand_group_members()). Dòng nào KHÔNG phải
        group thì giữ nguyên. Chạy trong background thread vì có thể mất vài
        giây/group khi query Exchange GAL."""
        if not self.recipients:
            messagebox.showwarning("No list", "Load or add recipients first.")
            return

        original = list(self.recipients)

        def worker():
            new_list = []
            seen_emails = set()
            groups_expanded = []
            errors = []
            for name, email in original:
                try:
                    members = outlook_com.expand_group_members(email)
                except Exception as e:
                    members = None
                    errors.append(f"{email}: {e}")
                if members is None:
                    # Không phải group (hoặc không resolve được) -> giữ nguyên dòng gốc
                    key = email.lower()
                    if key not in seen_emails:
                        seen_emails.add(key)
                        new_list.append((name, email))
                    continue
                groups_expanded.append((name or email, len(members)))
                for m_name, m_email in members:
                    key = m_email.lower()
                    if key in seen_emails:
                        continue
                    seen_emails.add(key)
                    new_list.append((m_name, m_email))

            def apply_result():
                self.recipients = new_list
                self._refresh_recipient_tree()
                self._save_recipients_to_db(silent=True)
                lines = [f"Total after expanding: {len(new_list)} people (was {len(original)} rows)."]
                if groups_expanded:
                    lines.append("\nGroups expanded:")
                    lines.extend(f"  • {n} → {c} members" for n, c in groups_expanded)
                else:
                    lines.append("\nNo group email detected in the list — everyone was already an "
                                  "individual address (or Outlook couldn't resolve them as a group; "
                                  "see notes below if that's unexpected).")
                if errors:
                    lines.append("\n⚠️ Some rows could not be checked (Outlook/COM error):")
                    lines.extend(f"  • {e}" for e in errors)
                messagebox.showinfo("Expand groups — done", "\n".join(lines))

            self.after(0, apply_result)

        threading.Thread(target=worker, daemon=True).start()

    # ══════════════════════════════════════════════════════════════════
    # TAB 3 — COMPOSE & SEND (VOTING BUTTONS + LANGUAGE + TRANSLATION)
    # ══════════════════════════════════════════════════════════════════
    def _build_tab_compose(self):
        f = self.tab_compose.body
        pad = {"padx": 10, "pady": 5}

        # ── send mode selector (TOP-MOST — important choice, made first) ──
        send_mode_bar = ttk.Frame(f, relief="ridge", borderwidth=1)
        send_mode_bar.grid(row=0, column=0, columnspan=2, sticky="we", padx=10, pady=(8, 4))
        ttk.Label(send_mode_bar, text="📤 Send mode:", font=("Arial", 11, "bold"))\
            .pack(side="left", padx=(8, 6), pady=6)
        self.combo_send_mode = ttk.Combobox(
            send_mode_bar, width=28, state="readonly",
            values=["Send first Invite", "Send update invite", "Send Gift Contribution Notice"],
        )
        self.combo_send_mode.current(0)  # default: first invite
        self.combo_send_mode.pack(side="left", padx=4, pady=6)
        self.combo_send_mode.bind("<<ComboboxSelected>>", lambda e: (self._refresh_send_button_label(), self._refresh_compose_preview()))
        self._make_wrapping_label(send_mode_bar,
                  text="  ↳ 'Update invite' adds a change-notice banner + different subject prefix, but\n"
                       "still lets people re-vote. 'Gift Contribution Notice' sends a Guest-of-Honor/\n"
                       "Organizer/Budget notice from Tab 1 — NO voting buttons (see 'Event mode' on Tab 1).",
                  font=("Arial", 8, "italic")).pack(side="left", padx=4, pady=6)

        # ── language selector ──
        ttk.Label(f, text="Email language:", font=("Arial", 10, "bold"))\
            .grid(row=1, column=0, sticky="w", **pad)
        self.combo_email_lang = ttk.Combobox(
            f, width=32, state="readonly",
            values=[LANG_LABELS["en"], LANG_LABELS["ja"], LANG_LABELS["vi"], LANG_LABELS["bilingual"]],
        )
        self.combo_email_lang.current(0)  # default: English
        self.combo_email_lang.grid(row=1, column=1, sticky="w", **pad)
        self.combo_email_lang.bind("<<ComboboxSelected>>", lambda e: self._refresh_compose_preview())

        self.var_subject_preview = tk.StringVar(value="")
        ttk.Label(f, textvariable=self.var_subject_preview, font=("Arial", 9, "italic"))\
            .grid(row=2, column=0, columnspan=2, sticky="w", padx=10)

        self.var_greeting_preview = tk.StringVar(value="")
        ttk.Label(f, textvariable=self.var_greeting_preview, font=("Arial", 9, "italic"))\
            .grid(row=3, column=0, columnspan=2, sticky="w", padx=10, pady=(2, 8))

        # Email is assembled in THIS order: Greeting (above) → EDITABLE → FIXED.
        # ── editable block preview (organizer's free-text background note, shown FIRST) ──
        ttk.Label(f, text="🟩 EDITABLE part — YOUR background notes for this specific event, shown\n"
                          "right after the greeting (write in any language; translate below if needed):",
                  font=("Arial", 10, "bold"), foreground="#00703C").grid(row=4, column=0, columnspan=2, sticky="w", **pad)
        editable_container, self.txt_editable_preview = make_scrollable_text(f, width=100, height=5, bg="#e8f5e9")
        editable_container.grid(row=5, column=0, columnspan=2, sticky="w", padx=10)

        # ── fixed block preview (event details from Tab 1 + voting instructions, shown AFTER editable) ──
        ttk.Label(f, text="🟧 FIXED part — auto-filled from Tab 1 (event name/date/location/deadline/budget)\n"
                          "plus the standard voting wording, shown AFTER your note. You CAN edit this box —\n"
                          "click 'Save as default' below to reuse your edited wording for future events too:",
                  font=("Arial", 10, "bold"), foreground="#8a4b00").grid(row=6, column=0, columnspan=2, sticky="w", **pad)
        fixed_container, self.txt_fixed_preview = make_scrollable_text(f, width=100, height=9, bg="#fff3e0")
        fixed_container.grid(row=7, column=0, columnspan=2, sticky="w", padx=10)

        fixed_btns = ttk.Frame(f)
        fixed_btns.grid(row=8, column=0, columnspan=2, sticky="w", padx=10, pady=4)
        ttk.Button(fixed_btns, text="🔄 Refresh preview from Tab 1 / Tab 2", command=self._refresh_compose_preview)\
            .pack(side="left", padx=(0, 6))
        ttk.Button(fixed_btns, text="💾 Save FIXED wording as default for this language", command=self._save_fixed_default)\
            .pack(side="left", padx=6)
        ttk.Button(fixed_btns, text="↺ Reset FIXED wording to system default", command=self._reset_fixed_default)\
            .pack(side="left", padx=6)

        # ── prompt customization (NEW) ──
        ttk.Separator(f).grid(row=9, column=0, columnspan=2, sticky="ew", pady=8)
        ttk.Label(f, text="🎨 Customize Copilot prompt (for single-language targets — icons already built in):",
                  font=("Arial", 10, "bold")).grid(row=10, column=0, columnspan=2, sticky="w", **pad)
        self._make_wrapping_label(f, text="This box shows the exact prompt that will be sent to Copilot (system default,\n"
                          "already includes emoji-icon instructions ⏰📍💰📋👥). Edit it freely and click\n"
                          "'Save' to keep your version — it persists until you Reset. The bilingual target\n"
                          "uses its own built-in default (see 'Show system default prompt' below).",
                  font=("Arial", 9, "italic")).grid(row=11, column=0, columnspan=2, sticky="w", padx=10, pady=(2, 4))
        
        prompt_container, self.txt_custom_prompt = make_scrollable_text(f, width=100, height=11, bg="#fffce0")
        # Always show SOMETHING — the saved override if present, otherwise the
        # system default (with icon instructions) — so the user can see exactly
        # what will be sent, and edit it directly instead of starting from blank.
        current_prompt = self.prompt_overrides.get("single", "").strip() or DEFAULT_PROMPT_SINGLE
        self.txt_custom_prompt.insert("1.0", current_prompt)
        prompt_container.grid(row=12, column=0, columnspan=2, sticky="w", padx=10, pady=4)
        
        prompt_btns = ttk.Frame(f)
        prompt_btns.grid(row=13, column=0, columnspan=2, sticky="w", padx=10, pady=4)
        ttk.Button(prompt_btns, text="💾 Save custom prompt as default", command=self._save_custom_prompt)\
            .pack(side="left", padx=(0, 6))
        ttk.Button(prompt_btns, text="↺ Reset prompt to system default", command=self._reset_custom_prompt)\
            .pack(side="left", padx=6)
        ttk.Button(prompt_btns, text="📘 Show system default prompt (single + bilingual)", command=self._show_system_prompt)\
            .pack(side="left", padx=6)
        
        # ── translation helper (Copilot bridge) ──
        ttk.Separator(f).grid(row=14, column=0, columnspan=2, sticky="ew", pady=8)
        ttk.Label(f, text="Translate the FULL email (greeting + note + event details) via Copilot copy/paste bridge:",
                  font=("Arial", 10, "bold")).grid(row=15, column=0, columnspan=2, sticky="w", **pad)
        ttk.Label(f, text="(Source is the language currently shown above — switch off Bilingual to pick a source)",
                  font=("Arial", 9, "italic")).grid(row=16, column=0, columnspan=2, sticky="w", padx=10)

        helper = ttk.Frame(f)
        helper.grid(row=17, column=0, columnspan=2, sticky="w", padx=10)

        ttk.Label(helper, text="Translate into:").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.combo_translate_target = ttk.Combobox(helper, width=26, state="readonly", values=TRANSLATE_TARGETS)
        self.combo_translate_target.current(1)  # default Japanese
        self.combo_translate_target.grid(row=0, column=1, sticky="w", padx=4, pady=4)
        ttk.Button(helper, text="📋 Copy full email + prompt to clipboard (for Copilot)",
                   command=self._copy_email_for_translation).grid(row=0, column=2, sticky="w", padx=8, pady=4)

        ttk.Label(helper, text="Paste the translated result from Copilot here — this will be used\n"
                              "as the complete, ready-to-send email (no section labels).\n"
                              "⚠️ Pasting a NEW result? Click '🗑 Clear' FIRST — pasting into a\n"
                              "non-empty box appends instead of replacing, combining old + new text.")\
            .grid(row=1, column=0, columnspan=3, sticky="w", padx=4, pady=(8, 2))
        paste_container, self.txt_translation_paste = make_scrollable_text(helper, width=90, height=6)
        paste_container.grid(row=2, column=0, columnspan=3, sticky="w", padx=4)
        self._make_wrapping_label(helper, text="⚠️ If words look glued together with no spaces (a known Copilot-copy quirk,\n"
                              "outside this tool's control), click 'Clean up' below, then click 'Save' again\n"
                              "to apply the cleaned-up version — Clean up alone does not change what is sent.",
                  font=("Arial", 8, "italic"), foreground="#8a4b00")\
            .grid(row=3, column=0, columnspan=3, sticky="w", padx=4, pady=(0, 2))
        ttk.Button(helper, text="💾 Save as translated version for this language", command=self._save_translated_email)\
            .grid(row=4, column=0, sticky="w", padx=4, pady=6)
        ttk.Button(helper, text="🗑 Clear saved translation for this language", command=self._clear_translated_email)\
            .grid(row=4, column=1, sticky="w", padx=4, pady=6)
        ttk.Button(helper, text="🧹 Clean up (fix lost spaces/line breaks)", command=self._cleanup_pasted_text)\
            .grid(row=4, column=2, sticky="w", padx=4, pady=6)

        self.lbl_translation_status = ttk.Label(
            helper, text="Full translated email ready: EN ❌  |  JA ❌  |  VI ❌  |  Bilingual ❌")
        self.lbl_translation_status.grid(row=5, column=0, columnspan=3, sticky="w", padx=4, pady=2)

        # ── send controls ──
        ttk.Separator(f).grid(row=18, column=0, columnspan=2, sticky="ew", pady=8)

        self._make_wrapping_label(f, text='Send to (optional override — e.g. a department group email).\n'
                          "Leave blank to send individually to everyone on Tab 2.\n"
                          "Tab 2's roster is always used for vote tracking either way.",
                  font=("Arial", 9, "italic")).grid(row=19, column=0, columnspan=2, sticky="w", **pad)
        self.var_send_to_override = tk.StringVar(value="")
        ttk.Entry(f, textvariable=self.var_send_to_override, width=50).grid(row=20, column=0, columnspan=2, sticky="w", padx=10)

        ttk.Label(f, text="Voting options (leave as-is unless needed):").grid(row=21, column=0, sticky="w", **pad)
        self.var_voting_options = tk.StringVar(value="Yes;No;Maybe")
        ttk.Entry(f, textvariable=self.var_voting_options, width=30).grid(row=21, column=1, sticky="w", **pad)

        self.var_auto_send = tk.BooleanVar(value=False)
        ttk.Checkbutton(f, text="Send immediately without review (unchecked = open Outlook for you to click Send)",
                         variable=self.var_auto_send).grid(row=22, column=0, columnspan=2, sticky="w", **pad)

        self.var_send_btn_label = tk.StringVar(value="✉ Send Invite via Outlook (Voting Buttons)")
        ttk.Button(f, textvariable=self.var_send_btn_label, command=self._send_invite)\
            .grid(row=23, column=0, sticky="w", **pad)

        self._refresh_send_button_label()
        self._refresh_compose_preview()

    def _current_lang_code(self):
        label = self.combo_email_lang.get()
        return LANG_LABEL_TO_CODE.get(label, "en")

    def _is_update_mode(self):
        return self.combo_send_mode.get() == "Send update invite"

    def _is_gift_mode(self):
        return self.combo_send_mode.get() == "Send Gift Contribution Notice"

    def _refresh_send_button_label(self):
        if not hasattr(self, "var_send_btn_label"):
            return
        if self._is_gift_mode():
            self.var_send_btn_label.set("🎁 Send Gift Contribution Notice via Outlook")
        else:
            self.var_send_btn_label.set("✉ Send Invite via Outlook (Voting Buttons)")

    def _active_full_translations(self):
        """Dict lưu bản dịch Copilot đầy đủ ĐANG DÙNG — tách riêng cho mode
        Gift (self.gift_full_translations) và mode Invite/Update invite
        (self.full_translations), để 2 loại nội dung hoàn toàn khác nhau
        không ghi đè lẫn nhau khi đổi qua lại Send mode trên cùng 1 ngôn
        ngữ."""
        return self.gift_full_translations if self._is_gift_mode() else self.full_translations

    def _refresh_subject_preview_only(self):
        """Cập nhật CHỈ phần Subject preview (không đụng tới 2 ô nội dung
        EDITABLE/FIXED) — dùng khi Gửi mail hoặc Copy-để-dịch, để KHÔNG xoá
        mất nội dung bạn vừa gõ TAY TRỰC TIẾP vào ô EDITABLE trên Tab 3.

        BUG ĐÃ SỬA: trước đây cả '📨 Send' lẫn '📋 Copy full email' đều gọi
        _refresh_compose_preview() (bản ĐẦY ĐỦ) trước khi đọc nội dung —
        hàm đó NẠP LẠI ô Editable từ ghi chú gốc ở Tab 1 (self.entry_note),
        nên nếu bạn gõ thêm/sửa trực tiếp vào ô Editable trên Tab 3 (thay vì
        quay lại Tab 1) mà CHƯA đồng bộ ngược, phần vừa gõ đó bị XOÁ MẤT
        ngay trước khi gửi/copy — y hệt hiện tượng bạn gặp."""
        lang_code = self._current_lang_code()
        event_name = self.var_event_name.get()
        event_id = self.var_event_id.get()
        if self._is_gift_mode():
            subject = build_gift_subject(lang_code, event_id, self.var_guest_of_honor.get())
        else:
            subject = build_subject(lang_code, event_id, event_name, is_update=self._is_update_mode())
        self._compose_subject = subject
        self.var_subject_preview.set(f"Subject preview: {subject}")

    def _editable_box_text_for_translation(self):
        """Đọc nội dung THẬT ĐANG CÓ trên ô EDITABLE của Tab 3 — đúng như
        bạn đang thấy/đã gõ trên màn hình (kể cả gõ tay trực tiếp vào đó) —
        bỏ đi hậu tố '[not yet translated...]' nếu có (hậu tố đó chỉ để
        hiển thị trong app, không phải nội dung thật cần dịch)."""
        text = self.txt_editable_preview.get("1.0", "end").strip()
        for flag in NOT_TRANSLATED_FLAG.values():
            flag = flag.strip()
            if flag and text.endswith(flag):
                text = text[: -len(flag)].rstrip()
                break
        return text

    def _source_note_text(self):
        return self.entry_note.get("1.0", "end").strip()

    def _lang_content(self, lang_code, event_name, event_date, location, deadline, budget):
        """Returns (fixed_text_or_None, editable_text, using_override) for a SINGLE
        language (en/ja/vi — not bilingual). fixed_text is None when a full Copilot-
        translated override is active for this language — in that case editable_text
        already holds the COMPLETE translated email (greeting + note + event details)."""
        override = self.full_translations.get(lang_code, "").strip()
        if override:
            return None, override, True
        fixed = self.fixed_overrides.get(lang_code, "").strip() or \
            build_fixed_block(lang_code, event_name, event_date, location, deadline, budget)
        editable = build_editable_block(lang_code, self._source_note_text(), False)
        return fixed, editable, False

    def _single_lang_full_body(self, lang_code, event_name, event_date, location, deadline, budget):
        """Full assembled body (greeting + note + fixed) for ONE language — used both for
        sending and as a building block for the bilingual (JA+EN) combination."""
        override = self.full_translations.get(lang_code, "").strip()
        if override:
            return override
        greeting = build_greeting(lang_code)
        fixed = self.fixed_overrides.get(lang_code, "").strip() or \
            build_fixed_block(lang_code, event_name, event_date, location, deadline, budget)
        editable = build_editable_block(lang_code, self._source_note_text(), False)
        return f"{greeting}\n\n{editable}\n\n{fixed}".strip()

    # ── MỚI: các hàm build nội dung riêng cho mode "Send Gift Contribution
    # Notice" — song song với _lang_content()/_single_lang_full_body() ở
    # trên nhưng dùng build_gift_fixed_block() thay vì build_fixed_block(),
    # và đọc/ghi self.gift_full_translations (KHÔNG dùng self.full_translations,
    # để tránh 2 loại nội dung khác nhau ghi đè lẫn nhau khi đổi Send mode). ──
    def _gift_lang_content(self, lang_code, guest_of_honor, start_time, event_date, location,
                            organizer, deadline, gift_budget):
        override = self.gift_full_translations.get(lang_code, "").strip()
        if override:
            return None, override, True
        fixed = build_gift_fixed_block(lang_code, guest_of_honor, start_time, event_date, location,
                                        organizer, deadline, gift_budget)
        editable = build_editable_block(lang_code, self._source_note_text(), False)
        return fixed, editable, False

    def _gift_single_lang_full_body(self, lang_code, guest_of_honor, start_time, event_date, location,
                                     organizer, deadline, gift_budget):
        override = self.gift_full_translations.get(lang_code, "").strip()
        if override:
            return override
        greeting = build_greeting(lang_code)
        fixed = build_gift_fixed_block(lang_code, guest_of_honor, start_time, event_date, location,
                                        organizer, deadline, gift_budget)
        editable = build_editable_block(lang_code, self._source_note_text(), False)
        return f"{greeting}\n\n{editable}\n\n{fixed}".strip()

    def _refresh_compose_preview(self):
        lang_code = self._current_lang_code()
        event_name = self.var_event_name.get()
        event_id = self.var_event_id.get()
        location = self.var_location.get()
        budget = self.var_budget.get()
        event_date = get_date_str(self.date_event)
        deadline = get_date_str(self.date_deadline)
        gift_deadline = get_date_str(self.date_gift_deadline)

        # MỚI: nhánh RIÊNG hoàn toàn cho mode "Send Gift Contribution Notice"
        # — cùng cấu trúc UI (Subject/Greeting/Editable/Fixed/Bilingual) như
        # nhánh Invite bên dưới, nhưng dùng nội dung + subject + dict lưu bản
        # dịch Copilot RIÊNG cho Gift (xem _gift_lang_content()/
        # _gift_single_lang_full_body() ở trên) — return sớm, không chạy tiếp
        # xuống logic Invite/Update invite bên dưới. Dùng "Gift contribution
        # deadline" (RIÊNG, khác "Event response deadline" của RSVP) làm hạn
        # đóng góp trong nội dung email.
        if self._is_gift_mode():
            guest_of_honor = self.var_guest_of_honor.get()
            organizer = self.var_organizer.get()
            gift_budget = self.var_gift_budget.get()
            start_time = self.var_start_time.get()

            subject = build_gift_subject(lang_code, event_id, guest_of_honor)
            self._compose_subject = subject
            self.var_subject_preview.set(f"Subject preview: {subject}")

            if lang_code == "bilingual":
                self.var_greeting_preview.set("Greeting (auto): shown inside the combined bilingual draft below (Japanese first, English second).")
                override = self.gift_full_translations.get("bilingual", "").strip()
                if override:
                    combined = override
                else:
                    ja_full = self._gift_single_lang_full_body(
                        "ja", guest_of_honor, start_time, event_date, location, organizer, gift_deadline, gift_budget)
                    en_full = self._gift_single_lang_full_body(
                        "en", guest_of_honor, start_time, event_date, location, organizer, gift_deadline, gift_budget)
                    combined = "[English below]\n\n" + ja_full + BILINGUAL_SEPARATOR + en_full
                fixed_display = ("→ Bilingual mode: the FIXED box is not used here. The complete bilingual "
                                  "draft (Japanese first, English second) is shown in the EDITABLE box below — "
                                  "you can hand-edit it there before sending.")
                editable_display = combined
                self.txt_fixed_preview.config(state="normal")
                self.txt_fixed_preview.delete("1.0", "end")
                self.txt_fixed_preview.insert("1.0", fixed_display)
                self.txt_fixed_preview.config(state="disabled")
            else:
                self.var_greeting_preview.set(f"Greeting (auto, appears first): {build_greeting(lang_code)}")
                fixed_text, editable_text, override_used = self._gift_lang_content(
                    lang_code, guest_of_honor, start_time, event_date, location, organizer, gift_deadline, gift_budget)
                fixed_display = fixed_text if fixed_text is not None else \
                    "→ Using a full Copilot-translated email (see EDITABLE box below — this FIXED box is unused for this language)."
                editable_display = editable_text
                self.txt_fixed_preview.config(state="normal")
                self.txt_fixed_preview.delete("1.0", "end")
                self.txt_fixed_preview.insert("1.0", fixed_display)
                if override_used:
                    self.txt_fixed_preview.config(state="disabled")

            self.txt_editable_preview.delete("1.0", "end")
            self.txt_editable_preview.insert("1.0", editable_display)
            self._refresh_translation_status()
            return

        subject = build_subject(lang_code, event_id, event_name, is_update=self._is_update_mode())
        self._compose_subject = subject
        self.var_subject_preview.set(f"Subject preview: {subject}")

        if lang_code == "bilingual":
            self.var_greeting_preview.set("Greeting (auto): shown inside the combined bilingual draft below (Japanese first, English second).")
            override = self.full_translations.get("bilingual", "").strip()
            if override:
                combined = override
            else:
                ja_full = self._single_lang_full_body("ja", event_name, event_date, location, deadline, budget)
                en_full = self._single_lang_full_body("en", event_name, event_date, location, deadline, budget)
                combined = "[English below]\n\n" + ja_full + BILINGUAL_SEPARATOR + en_full
            if self._is_update_mode():
                combined = build_update_notice("bilingual") + "\n\n" + combined
            fixed_display = ("→ Bilingual mode: the FIXED box is not used here. The complete bilingual "
                              "draft (Japanese first, English second) is shown in the EDITABLE box below — "
                              "you can hand-edit it there before sending.")
            editable_display = combined
            self.txt_fixed_preview.config(state="normal")
            self.txt_fixed_preview.delete("1.0", "end")
            self.txt_fixed_preview.insert("1.0", fixed_display)
            self.txt_fixed_preview.config(state="disabled")
        else:
            self.var_greeting_preview.set(f"Greeting (auto, appears first): {build_greeting(lang_code)}")
            fixed_text, editable_text, override_used = self._lang_content(
                lang_code, event_name, event_date, location, deadline, budget)
            fixed_display = fixed_text if fixed_text is not None else \
                "→ Using a full Copilot-translated email (see EDITABLE box below — this FIXED box is unused for this language)."
            editable_display = editable_text
            if self._is_update_mode():
                # Chèn banner "thông tin đã thay đổi" NGAY ĐẦU ô Editable — vẫn
                # là văn bản có thể sửa/xoá tay như phần note bình thường, chỉ
                # là được tự động thêm sẵn khi chọn 'Send update invite'.
                editable_display = build_update_notice(lang_code) + "\n" + editable_display

            self.txt_fixed_preview.config(state="normal")
            self.txt_fixed_preview.delete("1.0", "end")
            self.txt_fixed_preview.insert("1.0", fixed_display)
            if override_used:
                self.txt_fixed_preview.config(state="disabled")
            # else: leave editable, per point 2 — user can hand-edit + save as new default

        self.txt_editable_preview.delete("1.0", "end")
        self.txt_editable_preview.insert("1.0", editable_display)

        self._refresh_translation_status()

    def _refresh_translation_status(self):
        translations = self._active_full_translations()
        def mark(code):
            return "✅" if translations.get(code, "").strip() else "❌"
        self.lbl_translation_status.config(
            text=f"Full translated email ready: EN {mark('en')}  |  JA {mark('ja')}  |  "
                 f"VI {mark('vi')}  |  Bilingual {mark('bilingual')}"
        )

    def _save_fixed_default(self):
        if self._is_gift_mode():
            messagebox.showinfo(
                "Not available for Gift mode",
                "The FIXED wording for 'Send Gift Contribution Notice' is always freshly built "
                "from Tab 1 (Guest of Honor/Organizer/Location/Deadline/Gift budget), so there's "
                "no cross-event default to save here.\n\n"
                "You can still hand-edit the FIXED box above before sending this specific email — "
                "that edit just won't be remembered for future events."
            )
            return
        lang_code = self._current_lang_code()
        if lang_code not in ("en", "ja", "vi"):
            messagebox.showinfo(
                "Switch language first",
                "Custom default wording is saved per single language.\n\n"
                "Switch 'Email language' to English, Japanese, or Vietnamese first."
            )
            return
        if self.full_translations.get(lang_code, "").strip():
            messagebox.showwarning(
                "Clear the Copilot translation first",
                "This language is currently using a full Copilot-translated email override, "
                "so the FIXED box isn't active. Clear that translation first (button below) "
                "if you want to edit and save the FIXED wording instead."
            )
            return
        text = self.txt_fixed_preview.get("1.0", "end").strip()
        if not text:
            messagebox.showwarning("Empty", "The FIXED box is empty.")
            return
        self.fixed_overrides[lang_code] = text
        history.save_fixed_overrides(self.fixed_overrides)
        messagebox.showinfo(
            "Saved",
            f"Saved your edited wording as the new default FIXED text for "
            f"{self.combo_email_lang.get()}.\n\n"
            "It will be used automatically for this and future events (until you Reset it)."
        )

    def _reset_fixed_default(self):
        if self._is_gift_mode():
            messagebox.showinfo(
                "Not available for Gift mode",
                "There's no saved default to reset for Gift mode — the FIXED box is always freshly "
                "built from Tab 1. Click '🔄 Refresh preview from Tab 1 / Tab 2' instead to reload it."
            )
            return
        lang_code = self._current_lang_code()
        if lang_code not in ("en", "ja", "vi"):
            messagebox.showinfo(
                "Switch language first",
                "Switch 'Email language' to English, Japanese, or Vietnamese first."
            )
            return
        self.fixed_overrides[lang_code] = ""
        history.save_fixed_overrides(self.fixed_overrides)
        self._refresh_compose_preview()
        messagebox.showinfo("Reset", f"{self.combo_email_lang.get()} FIXED wording reset to the system default.")

    def _save_custom_prompt(self):
        """Save the custom prompt template to disk for reuse."""
        custom_prompt = self.txt_custom_prompt.get("1.0", "end").strip()
        if not custom_prompt:
            messagebox.showwarning("Empty", "The prompt box is empty. Type a prompt, or click "
                                             "'Reset prompt to system default' to restore the built-in one.")
            return
        self.prompt_overrides["single"] = custom_prompt
        history.save_prompt_overrides(self.prompt_overrides)
        messagebox.showinfo(
            "Saved",
            "Custom prompt template saved.\n\n"
            "It will be used for single-language translations (English/Japanese/Vietnamese "
            "as the TRANSLATE-INTO target) from now on — including this exact wording, icons, "
            "and instructions — until you Reset it."
        )

    def _reset_custom_prompt(self):
        """Reset prompt back to system default (with icon instructions) — NOT blank."""
        self.prompt_overrides["single"] = ""
        history.save_prompt_overrides(self.prompt_overrides)
        self.txt_custom_prompt.config(state="normal")
        self.txt_custom_prompt.delete("1.0", "end")
        self.txt_custom_prompt.insert("1.0", DEFAULT_PROMPT_SINGLE)
        messagebox.showinfo("Reset", "Prompt reset to the system default (shown in the box above — "
                                      "this is exactly what gets sent to Copilot, icon instructions included).")

    def _show_system_prompt(self):
        """Display the full built-in default prompts (both single-language and bilingual),
        so the user can see exactly what's used when no custom override is saved."""
        messagebox.showinfo(
            "System Default Prompt — Single language (EN/JA/VI)",
            "Used when 'Translate into' = English / Japanese / Vietnamese, and no custom "
            "prompt is saved (this is also what's pre-filled in the editable box above):\n\n"
            + DEFAULT_PROMPT_SINGLE
        )
        messagebox.showinfo(
            "System Default Prompt — Bilingual (Japanese + English)",
            "Used when 'Translate into' = Bilingual (Japanese + English). This one is "
            "fixed/built-in (no separate customization box for it yet):\n\n"
            + DEFAULT_PROMPT_BILINGUAL
        )

    def _copy_email_for_translation(self):
        lang_code = self._current_lang_code()
        if lang_code == "bilingual":
            messagebox.showinfo(
                "Switch language first",
                "Pick the SOURCE language you're drafting in first (English/Japanese/Vietnamese).\n\n"
                "You can still choose 'Bilingual (Japanese + English)' as the TRANSLATE-INTO target "
                "below — Copilot will produce both languages from that single source."
            )
            return
        self._refresh_subject_preview_only()
        greeting = build_greeting(lang_code)
        # Rebuild FIXED content directly (bypass the preview box on purpose):
        # when a full-translation override is already saved for this language
        # (self.full_translations[lang_code]), _refresh_compose_preview() fills
        # txt_fixed_preview with a PLACEHOLDER string — "→ Using a full
        # Copilot-translated email (see EDITABLE box below...)" — not real
        # event details. Reading that placeholder here and sending it to
        # Copilot as "content to translate" produced garbage/stale-looking
        # results whenever you tried to Copy-for-translation again after
        # already having saved a translation for the current source language.
        # Always use the LIVE Tab-1 event details instead, regardless of
        # whatever override happens to be saved.
        event_name = self.var_event_name.get()
        event_date = get_date_str(self.date_event)
        location = self.var_location.get()
        deadline = get_date_str(self.date_deadline)
        budget = self.var_budget.get()
        if self._is_gift_mode():
            fixed_text = build_gift_fixed_block(
                lang_code, self.var_guest_of_honor.get(), self.var_start_time.get(), event_date,
                location, self.var_organizer.get(), get_date_str(self.date_gift_deadline), self.var_gift_budget.get())
        else:
            fixed_text = self.fixed_overrides.get(lang_code, "").strip() or \
                build_fixed_block(lang_code, event_name, event_date, location, deadline, budget)
        # BUG ĐÃ SỬA: trước đây đọc note từ Tab 1 (self._source_note_text()),
        # nghĩa là nếu bạn gõ thêm/sửa trực tiếp vào ô EDITABLE trên Tab 3
        # (thay vì quay lại Tab 1) thì phần đó bị BỎ QUA hoàn toàn khi Copy —
        # y hệt hiện tượng "nội dung đã gõ biến mất". Giờ đọc TRỰC TIẾP từ ô
        # Editable đang hiển thị trên Tab 3 — đúng nguyên văn những gì bạn
        # thấy trên màn hình lúc bấm Copy.
        editable_text = self._editable_box_text_for_translation()
        if not fixed_text and not editable_text:
            messagebox.showwarning("Nothing to translate", "The email preview is empty.")
            return

        target_label = self.combo_translate_target.get()
        target_code = self._target_lang_code()

        # Email order is: Greeting -> organizer's note -> event details/voting instructions.
        note_part = editable_text if editable_text else "(no note was written for this event)"
        combined = f"{greeting}\n\n{note_part}\n\n{fixed_text}"

        # Build prompt.
        # IMPORTANT: for the single-language box, read directly from what's shown/edited
        # in txt_custom_prompt — that box is ALWAYS pre-filled with either the saved
        # override or the system default (which already includes icon instructions), so
        # "what you see in the box is exactly what gets sent" — no more silently falling
        # back to an old icon-less hardcoded prompt just because nothing was Saved yet.
        if target_code == "bilingual":
            custom_prompt_bilingual = self.prompt_overrides.get("bilingual", "").strip()
            prompt_template = custom_prompt_bilingual or DEFAULT_PROMPT_BILINGUAL
            prompt = prompt_template + "\n\n" + combined
        else:
            prompt_template = self.txt_custom_prompt.get("1.0", "end").strip() or DEFAULT_PROMPT_SINGLE
            # Replace [TARGET_LANGUAGE] placeholder with the actual language name
            prompt_template = prompt_template.replace("[TARGET_LANGUAGE]", target_label)
            prompt = prompt_template + "\n\n" + combined

        self.clipboard_clear()
        self.clipboard_append(prompt)

        stale_override_note = ""
        if self._active_full_translations().get(lang_code, "").strip():
            stale_override_note = (
                f"\n\n⚠️ Note: a translation is STILL SAVED for {self.combo_email_lang.get()} "
                "from earlier — the preview above will keep showing that OLD saved version "
                "until you paste and Save the NEW result below (or click '🗑 Clear saved "
                "translation' first if you don't want it replaced)."
            )

        messagebox.showinfo(
            "Copied",
            "The full email (greeting + note + event details) plus translation instructions "
            "were copied to the clipboard — freshly rebuilt from the CURRENT Tab 1/Tab 2 "
            "settings and note, regardless of any previously saved translation." + stale_override_note +
            "\n\nNext steps:\n"
            "1. Open Copilot (or any AI assistant)\n"
            "2. Paste (Ctrl+V) and send\n"
            "3. Copy the translated reply\n"
            "4. Come back here, paste it in the box below, and click "
            "'Save as translated version for this language'"
        )

    def _target_lang_code(self):
        target = self.combo_translate_target.get()
        return {
            "English": "en", "Japanese": "ja", "Vietnamese": "vi",
            "Bilingual (Japanese + English)": "bilingual",
        }.get(target, "en")

    def _cleanup_pasted_text(self):
        """Apply best-effort repair to the pasted box: first remove a
        duplicated whole-email copy if detected (see dedupe_pasted_translation
        docstring — this can come from Copilot's own reply, not just from
        re-pasting), then fix lost spaces/line breaks on what remains.
        Deduplication runs AUTOMATICALLY (no confirmation prompt) — duplicate
        content is never something you'd want to keep in the final email, and
        a Yes/No dialog here was too easy to accidentally dismiss/misread."""
        text = self.txt_translation_paste.get("1.0", "end").strip()
        if not text:
            messagebox.showwarning("Empty", "Nothing to clean up — the paste box is empty.")
            return

        working_text = text
        dedupe_note = ""
        if detect_possible_duplicate_paste(text):
            deduped = dedupe_pasted_translation(text)
            removed = len(text) - len(deduped)
            if removed > 20:
                working_text = deduped
                dedupe_note = (
                    f"\n\n🔁 Also detected and removed ~{removed} characters of "
                    "DUPLICATED content at the start (the whole email appeared "
                    "twice — this can happen in Copilot's own reply, e.g. a "
                    "malformed draft immediately followed by a self-corrected "
                    "version — not just from pasting twice). Kept the later copy."
                )

        cleaned = cleanup_pasted_translation(working_text)
        self.txt_translation_paste.delete("1.0", "end")
        self.txt_translation_paste.insert("1.0", cleaned)
        messagebox.showinfo(
            "Cleaned up — remember to Save",
            "Applied best-effort fixes: stray '**'/'***' markers removed, spacing "
            "added around icons, line breaks restored before '•' bullets and "
            "after Japanese '。' sentence endings." + dedupe_note +
            "\n\nThis can't perfectly reconstruct every lost space inside a plain "
            "sentence — please read through the result once before saving.\n\n"
            "⚠️ IMPORTANT: this only updates the box above. It does NOT change "
            "what actually gets sent until you click '💾 Save as translated "
            "version for this language' again."
        )

    def _save_translated_email(self):
        text = self.txt_translation_paste.get("1.0", "end").strip()
        if not text:
            messagebox.showwarning("Empty", "Paste the translated email first.")
            return

        dedupe_note = ""
        if detect_possible_duplicate_paste(text):
            deduped = dedupe_pasted_translation(text)
            removed = len(text) - len(deduped)
            if removed > 20:
                text = cleanup_pasted_translation(deduped)
                self.txt_translation_paste.delete("1.0", "end")
                self.txt_translation_paste.insert("1.0", text)
                dedupe_note = (
                    f"\n\n🔁 Also detected and automatically removed ~{removed} "
                    "characters of DUPLICATED content at the start before saving "
                    "(the whole email appeared twice — can happen in Copilot's own "
                    "reply, not just from pasting twice). Kept the later copy."
                )

        code = self._target_lang_code()
        self._active_full_translations()[code] = text

        # Auto-switch "Email language" to match what was just saved — this used
        # to be a manual step ("switch Email language to see it") that was easy
        # to forget, causing the tool to silently send the OLD default-template
        # email instead of the translation you just pasted. Now it's applied
        # immediately, so what you see in the preview is guaranteed to be what
        # gets sent.
        target_label = LANG_LABELS.get(code)
        if target_label:
            self.combo_email_lang.set(target_label)

        self._refresh_compose_preview()
        messagebox.showinfo(
            "Saved & Applied",
            f"Saved the {self.combo_translate_target.get()} translation, and switched "
            f"'Email language' to {target_label} so it's shown in the preview below "
            "and will be used when you click Send — no extra step needed." + dedupe_note
        )

    def _clear_translated_email(self):
        code = self._target_lang_code()
        self._active_full_translations()[code] = ""
        self.txt_translation_paste.delete("1.0", "end")
        self._refresh_compose_preview()

    def _compose_full_body(self):
        """Assemble the FINAL email body for sending.
        ALWAYS prioritize full translation override (from Copilot) if present.
        If no override exists, build manually from components."""
        lang_code = self._current_lang_code()

        # ✅ ALWAYS check for full translation override FIRST (from Copilot)
        # MỚI: dùng đúng dict theo mode hiện tại (Gift dùng self.gift_full_translations
        # riêng, không lẫn với self.full_translations của Invite/Update invite).
        full_override = self._active_full_translations().get(lang_code, "").strip()
        if full_override:
            # Use the COMPLETE translated email (greeting + note + details already in it)
            return full_override
        
        # If no override, build manually from components
        if lang_code == "bilingual":
            # BUG ĐÃ SỬA: UI ghi rõ ở ô FIXED (bilingual mode) "...shown in the
            # EDITABLE box below — you can hand-edit it there before sending",
            # nhưng trước đây hàm này KHÔNG hề đọc ô Editable — luôn build lại
            # từ đầu bằng self._source_note_text() (ghi chú gốc Tab 1), nên
            # MỌI chỉnh sửa tay trực tiếp trên Tab 3 đều bị bỏ qua lúc gửi
            # thật. Giờ ưu tiên đọc TRỰC TIẾP từ ô Editable đang hiển thị —
            # đúng bản nháp song ngữ bạn đã xem/sửa trước khi gửi.
            box_text = self.txt_editable_preview.get("1.0", "end").strip()
            if box_text:
                return box_text
            # Ô đang trống (vd chưa từng bấm sang Bilingual/chưa Refresh
            # preview lần nào) -> build mới như cũ, làm fallback an toàn.
            # MỚI: fallback cũng phải phân biệt Gift mode (dùng thông tin
            # Guest of Honor/Organizer/Gift budget) vs Invite thường.
            if self._is_gift_mode():
                ja_full = self._gift_single_lang_full_body(
                    "ja", self.var_guest_of_honor.get(), self.var_start_time.get(),
                    get_date_str(self.date_event), self.var_location.get(),
                    self.var_organizer.get(), get_date_str(self.date_gift_deadline), self.var_gift_budget.get())
                en_full = self._gift_single_lang_full_body(
                    "en", self.var_guest_of_honor.get(), self.var_start_time.get(),
                    get_date_str(self.date_event), self.var_location.get(),
                    self.var_organizer.get(), get_date_str(self.date_gift_deadline), self.var_gift_budget.get())
            else:
                ja_full = self._single_lang_full_body("ja", self.var_event_name.get(),
                                                       get_date_str(self.date_event),
                                                       self.var_location.get(),
                                                       get_date_str(self.date_deadline),
                                                       self.var_budget.get())
                en_full = self._single_lang_full_body("en", self.var_event_name.get(),
                                                       get_date_str(self.date_event),
                                                       self.var_location.get(),
                                                       get_date_str(self.date_deadline),
                                                       self.var_budget.get())
            return "[English below]\n\n" + ja_full + BILINGUAL_SEPARATOR + en_full
        else:
            # Build single language: greeting + editable + fixed
            # (đọc TRỰC TIẾP từ 2 ô đang hiển thị trên Tab 3 — đã đúng nội
            # dung Gift hay Invite tuỳ mode, vì _refresh_compose_preview() đã
            # điền đúng nội dung cho từng mode; không cần branch thêm ở đây)
            greeting = build_greeting(lang_code)
            fixed_text = self.txt_fixed_preview.get("1.0", "end").strip()
            editable_text = self.txt_editable_preview.get("1.0", "end").strip()
            return f"{greeting}\n\n{editable_text}\n\n{fixed_text}".strip()

    def _send_invite(self):
        if not self.recipients:
            messagebox.showwarning("No recipients", "Go to Tab 2 and load the recipient list first.")
            return
        # BUG ĐÃ SỬA: trước đây bước này bị THIẾU ở đây (dù đã có ở
        # _save_to_history() và _update_history_from_tab1()) — nên nếu bạn
        # gửi invite mà CHƯA từng tự bấm '💾 Save to Excel' ở Tab 2 sau khi
        # đặt Event ID, cột RecipientFile ghi vào History lúc gửi vẫn là tên
        # file GỐC bạn Browse vào (vd file TEMPLATE dùng chung nhiều sự kiện)
        # thay vì file 'Participant_List_{EventID}.xlsx' đã chuẩn hoá — dù
        # file chuẩn hoá đó CÓ THỂ đã tồn tại sẵn trên đĩa từ 1 bước khác.
        self._save_recipients_to_db(silent=True)
        self._refresh_subject_preview_only()
        subject = self._compose_subject
        body = self._compose_full_body()
        voting_options = self.var_voting_options.get().strip()
        auto_send = self.var_auto_send.get()
        send_to_override = self.var_send_to_override.get().strip() or None
        is_update = self._is_update_mode()
        is_gift = self._is_gift_mode()

        def worker():
            try:
                outlook_com.send_voting_invite(
                    self.recipients, subject, body,
                    voting_options=voting_options, auto_send=auto_send,
                    send_to_override=send_to_override,
                    # MỚI: email "Send Gift Contribution Notice" KHÔNG có
                    # Voting Buttons/ảnh hướng dẫn vote — đây chỉ là thông
                    # báo kêu gọi đóng góp, không phải 1 cuộc bỏ phiếu.
                    use_voting_buttons=not is_gift,
                )
                # Record the actual send time NOW (not whenever "Save event to
                # History" happens to be clicked later on Tab 4, which could be
                # minutes/hours/days after the real send) — Tab 4's Save uses
                # this if present. Also used by Tab 5 to find the matching sent
                # confirmation email to attach to the Calendar Invite.
                send_time = datetime.now()
                if is_update:
                    # Theo dõi RIÊNG lần gửi update-invite (khác _invite_sent_date
                    # = lần gửi invite ĐẦU TIÊN) — Tab 5 dùng CẢ HAI để tìm và
                    # đính kèm đủ 2 email (gốc + update) vào Calendar Invite.
                    self._update_invite_sent_date = send_time
                elif not is_gift:
                    self._invite_sent_date = send_time

                # ALSO write/update the History row right now — previously the
                # History file only got a row when "Save event to History" was
                # manually clicked on Tab 4, which could be much later (or
                # never), leaving SentDate/EventID missing from History even
                # though the email had already gone out. This writes the
                # fields known at send-time immediately; Tab 4's later Save
                # will UPDATE this same row (matched by EventID) with vote
                # counts, cost tracking, etc. — no duplicate rows are created.
                try:
                    record = {
                        "EventID": self.var_event_id.get(),
                        "EventName": self.var_event_name.get(),
                        "EventDate": get_date_str(self.date_event),
                        "Deadline": get_date_str(self.date_deadline),
                        "Location": self.var_location.get(),
                        "Budget": self.var_budget.get(),
                        "EmailLanguage": self.combo_email_lang.get(),
                        "OrganizerNote": self._source_note_text(),
                        "RecipientFile": self.recipient_file.get(),
                        "TotalInvited": len(self.recipients),
                        "EventMode": self.var_event_mode.get(),
                        "Organizer": self.var_organizer.get(),
                        "GuestOfHonor": self.var_guest_of_honor.get(),
                        "GiftBudget": self.var_gift_budget.get(),
                        "GiftDeadline": get_date_str(self.date_gift_deadline),
                        "StartTime": self.var_start_time.get(),
                        "EndTime": self.var_end_time.get(),
                    }
                    # Chỉ ghi ĐÚNG 1 trong 2 cột theo chế độ gửi — cột còn lại
                    # KHÔNG được đưa vào record, nên db.save_event_record()
                    # (đã sửa để giữ nguyên giá trị cũ khi UPDATE 1 dòng có sẵn)
                    # sẽ không đụng tới nó. Vd: gửi update invite lần 2 sẽ chỉ
                    # cập nhật UpdateInviteDate, SentDate gốc vẫn giữ nguyên.
                    # MỚI: gửi "Gift Contribution Notice" KHÔNG đụng tới CẢ 2
                    # cột này — đây không phải 1 lần gửi RSVP invite, nên
                    # không nên ghi đè/thay đổi trạng thái theo dõi invite.
                    if is_gift:
                        pass
                    elif is_update:
                        record["UpdateInviteDate"] = send_time.strftime("%Y-%m-%d %H:%M")
                    else:
                        record["SentDate"] = send_time.strftime("%Y-%m-%d %H:%M")
                    db.save_event_record(record, self.history_path.get())
                    self.after(0, self._refresh_history_tree)
                except Exception:
                    pass  # best-effort — a History-write failure shouldn't block the send confirmation

                mode_label = "Gift Contribution Notice" if is_gift else ("UPDATE invite" if is_update else "invite")
                if is_gift:
                    history_note = "🗂 History updated with the latest Tab 1 details for this EventID."
                elif is_update:
                    history_note = "🗂 History updated with the update-invite time for this EventID."
                else:
                    history_note = "🗂 History updated with the send time for this EventID."
                self.after(0, lambda: messagebox.showinfo(
                    "Done",
                    f"{mode_label} email {'SENT' if auto_send else 'OPENED for review'} "
                    + (f"to group address: {send_to_override}" if send_to_override
                       else f"individually for {len(self.recipients)} people")
                    + ".\n\n"
                    + ("" if auto_send else "Check it in Outlook, then click Send.")
                    + f"\n\n{history_note}"
                ))
            except Exception as e:
                err_msg = str(e)
                self.after(0, lambda: messagebox.showerror(
                    "Error",
                    f"Could not send via Outlook:\n{err_msg}\n\n"
                    "Check: is Outlook desktop open & signed in? Is pywin32 installed?"
                ))

        threading.Thread(target=worker, daemon=True).start()

    # ══════════════════════════════════════════════════════════════════
    # TAB 4 — COLLECT RESPONSES
    # ══════════════════════════════════════════════════════════════════
    def _build_tab_collect(self):
        f = self.tab_collect.body

        info_frame = ttk.LabelFrame(f, text="🔍 How 'Scan Inbox' works — read before using")
        info_frame.pack(fill="x", padx=10, pady=(8, 4))
        self._make_wrapping_label(
            info_frame, justify="left",
            text=(
                "• Scans the ENTIRE Inbox (every folder, including subfolders) for emails whose "
                "Subject contains the EXACT current Event ID from Tab 1 — it doesn't matter whether "
                "the email is the original invite, an update invite, or a reminder, as long as the "
                "Subject contains the Event ID it counts.\n"
                "• For each sender, only the MOST RECENT vote is kept if they clicked more than once "
                "(votes are not added up).\n"
                "• Does NOT re-run automatically when you change the Event ID on Tab 1 — the table "
                "below only updates when you click '📨 Scan Inbox for Vote results'. If you switch to "
                "a different event and forget to click it again, the table will still show the results "
                "of the OLD Event ID (check the status line right below to be sure)."
            ),
            font=("Arial", 8)).pack(anchor="w", padx=8, pady=(4, 2))
        self.lbl_scan_status = ttk.Label(info_frame, text="", font=("Arial", 9, "bold"))
        self.lbl_scan_status.pack(anchor="w", padx=8, pady=(2, 6))

        top = ttk.Frame(f)
        top.pack(fill="x", padx=10, pady=8)
        ttk.Button(top, text="📨 Scan Inbox for Vote results", command=self._collect_responses)\
            .pack(side="left", padx=4)
        ttk.Button(top, text="🗂 Save event", command=self._save_to_history)\
            .pack(side="left", padx=4)

        # ── folder scanning note ──
        # Automatic: tool scans ALL folders (AdvancedSearch) to find vote results
        # No manual folder selection needed — kept simple for better UX
        self.var_scan_all_folders = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            f,
            text="✅ Tool automatically scans ALL folders for vote results (powered by AdvancedSearch)",
            variable=self.var_scan_all_folders,
        ).pack(anchor="w", padx=10, pady=6)

        ttk.Label(f, text="✅ Responded / not yet responded (based on Tab 2 + scanned votes):",
                  font=("Arial", 9, "bold")).pack(anchor="w", padx=10, pady=(4, 0))
        # MỚI: double-click ô "Vote" để sửa tay — dùng cho những người chỉ
        # đổi ý/báo miệng thay vì bấm lại nút Vote trong email (xem
        # _on_response_tree_double_click()/_commit_response_vote_edit()).
        # MỚI: thêm hẳn 1 cột checkbox "Manual edit" (✅/⬜) bên trái cột
        # "Name" — tick vào đó để mở dropdown Yes/No/Maybe sửa ngay (giống
        # double-click ô Vote); bỏ tick lại (với dòng ĐANG sửa tay) để trả
        # phiếu vote đó về cho lần Scan Inbox sau tự do cập nhật (xem
        # _on_response_manual_check_click()).
        self._make_wrapping_label(
            f, text="💡 Tick the \"Manual edit\" checkbox (or double-click the \"Vote\" cell) to correct "
                    "someone's vote by hand (Yes/No/Maybe) — useful when they just told you verbally or "
                    "changed their mind instead of clicking the button in the email. Manually-edited rows "
                    "are shown with a light blue background and stay checked. Un-ticking the checkbox for "
                    "an already manually-edited row lets the NEXT \"📨 Scan Inbox for Vote results\" freely "
                    "update it again from real emails; leaving it ticked KEEPS your manual value even after "
                    "re-scanning, unless a NEWER email vote is found for that same person.",
            font=("Arial", 8, "italic")).pack(anchor="w", padx=10)
        cols = ("manual_edit", "name", "email", "vote", "received")
        tree_container, self.tree_responses = make_scrollable_treeview(f, columns=cols, height=10)
        headers = ["✏️", "Name", "Email", "Vote", "Received At"]
        widths = [70, 220, 260, 120, 160]
        for c, label, w in zip(cols, headers, widths):
            self.tree_responses.heading(c, text=label)
            anchor = "center" if c == "manual_edit" else "w"
            self.tree_responses.column(c, width=w, anchor=anchor)
        self.tree_responses.bind("<Double-1>", self._on_response_tree_double_click)
        self.tree_responses.bind("<Button-1>", self._on_response_manual_check_click)
        tree_container.pack(fill="both", expand=True, padx=10, pady=6)

        self.lbl_summary = ttk.Label(f, text="No responses scanned yet.", font=("Arial", 10, "bold"))
        self.lbl_summary.pack(anchor="w", padx=10, pady=4)

        # ══════════════════════════════════════════════════════════════
        # KHUNG PHÍA DƯỚI — CHƯA PHẢN HỒI (pending), + soạn/gửi email nhắc nhở
        # ══════════════════════════════════════════════════════════════
        ttk.Separator(f).pack(fill="x", padx=10, pady=8)
        self.lbl_deadline_banner = ttk.Label(f, text="", font=("Arial", 10, "bold"))
        self.lbl_deadline_banner.pack(anchor="w", padx=10, pady=(0, 4))

        ttk.Label(f, text="🕓 NOT yet responded (no vote) — auto-updates every time you Scan Inbox:",
                  font=("Arial", 9, "bold"), foreground="#8a4b00").pack(anchor="w", padx=10, pady=(2, 0))
        pending_cols = ("name", "email")
        pending_container, self.tree_pending = make_scrollable_treeview(f, columns=pending_cols, height=6)
        for c, label, w in zip(pending_cols, ["Name", "Email"], [280, 320]):
            self.tree_pending.heading(c, text=label)
            self.tree_pending.column(c, width=w)
        pending_container.pack(fill="both", expand=False, padx=10, pady=6)

        reminder_frame = ttk.LabelFrame(f, text="📨 Compose & send a reminder email to people who haven't responded")
        reminder_frame.pack(fill="x", padx=10, pady=6)

        rbtn = ttk.Frame(reminder_frame)
        rbtn.pack(fill="x", padx=6, pady=4)
        ttk.Label(rbtn, text="Language:").pack(side="left", padx=(0, 4))
        self.combo_reminder_lang = ttk.Combobox(
            rbtn, width=26, state="readonly",
            values=[LANG_LABELS["en"], LANG_LABELS["ja"], LANG_LABELS["vi"], LANG_LABELS["bilingual"]],
        )
        self.combo_reminder_lang.current(0)  # default: English
        self.combo_reminder_lang.pack(side="left", padx=(0, 6))
        # Changing the language auto-regenerates the draft right away (same
        # behavior as combo_email_lang on Tab 3), so the user sees the new
        # language's content immediately instead of needing an extra click
        # on "Regenerate".
        self.combo_reminder_lang.bind("<<ComboboxSelected>>", lambda e: self._generate_reminder_draft())
        ttk.Button(rbtn, text="🔄 Regenerate reminder text (from Tab 1)",
                   command=self._generate_reminder_draft).pack(side="left", padx=(0, 6))
        self.var_reminder_attach = tk.BooleanVar(value=True)
        ttk.Checkbutton(rbtn, text="📎 Attach original invite email (looked up from RSVP_History)",
                        variable=self.var_reminder_attach).pack(side="left", padx=6)
        # The "Send now without review" checkbox has been REMOVED — per the
        # requirement, EVERY reminder email (both RSVP and Gift, see Tab 6)
        # now ALWAYS only opens Outlook so the user can review it and click
        # Send themselves; there is no longer an auto-send option from
        # within the app (see _send_reminder() below — now always called
        # with auto_send=False). The BACKGROUND script run via Task
        # Scheduler (send_scheduled_reminders.py) is a SEPARATE case — it
        # still sends directly as before, since that scenario has no one
        # sitting at the machine.

        ttk.Label(reminder_frame, text="Reminder email content (review/edit before sending):",
                  font=("Arial", 9, "italic")).pack(anchor="w", padx=6, pady=(4, 0))
        reminder_text_container, self.txt_reminder_body = make_scrollable_text(
            reminder_frame, width=100, height=7)
        reminder_text_container.pack(fill="x", padx=6, pady=4)

        self.lbl_reminder_send = ttk.Button(
            reminder_frame, text="📨 Send reminder email to 0 people who haven't responded",
            command=self._send_reminder)
        self.lbl_reminder_send.pack(anchor="w", padx=6, pady=(2, 8))

        # MỚI: đã BỎ khối "Actual cost tracking (Actual attendees/Cost per
        # person/Total income/Total expense/Calculate balance)" khỏi UI
        # theo yêu cầu — tính năng theo dõi tổng quát này đã được thay thế
        # hoàn toàn bởi Tab 5 "Attendance & Payment" (theo dõi CHI TIẾT
        # theo TỪNG NGƯỜI — actual attend/free/amount — thay vì chỉ vài con
        # số tổng gộp gõ tay). Các cột "ActualAttendees"/"CostPerPerson"/
        # "TotalIncome"/"TotalExpense"/"Balance" vẫn còn trong database
        # (không xoá schema, để KHÔNG mất dữ liệu các sự kiện CŨ đã từng
        # điền — xem Tab 7 Event History nếu cần xem lại) — chỉ không còn
        # ai ghi/đọc chúng nữa từ giờ (xem _save_to_history(), đã bỏ 5 key
        # này khỏi record để merge-preserve giữ nguyên giá trị cũ, không ghi
        # đè thành rỗng).
        self._update_scan_status_banner()

    def _load_folder_list(self):
        def worker():
            try:
                paths = outlook_com.list_folder_paths()
            except Exception as e:
                err_msg = str(e)
                self.after(0, lambda: messagebox.showerror(
                    "Error", f"Could not load folder list from Outlook:\n{err_msg}"))
                return
            self.after(0, lambda: self._fill_folder_listbox(paths))

        threading.Thread(target=worker, daemon=True).start()

    def _fill_folder_listbox(self, paths):
        self.list_folders.delete(0, "end")
        for p in paths:
            self.list_folders.insert("end", p)
        # Pre-select "Inbox" if present
        for i in range(self.list_folders.size()):
            if self.list_folders.get(i).strip().lower() == "inbox":
                self.list_folders.selection_set(i)

    def _selected_scan_folders(self):
        sel = self.list_folders.curselection()
        paths = [self.list_folders.get(i) for i in sel]
        if not paths:
            paths = ["Inbox"]  # default if nothing loaded/selected yet
        elif "inbox" not in [p.lower() for p in paths]:
            paths = ["Inbox"] + paths  # Inbox is always included
        return paths

    def _collect_responses(self):
        if not self.recipients:
            messagebox.showwarning("No list", "Go to Tab 2 and load the recipient list first.")
            return
        event_id = self.var_event_id.get()
        scan_all = self.var_scan_all_folders.get()
        folder_paths = None if scan_all else self._selected_scan_folders()

        def worker():
            try:
                responses, skipped = outlook_com.scan_voting_responses(
                    event_id, folder_paths=folder_paths, scan_all=scan_all)
            except Exception as e:
                err_msg = str(e)
                self.after(0, lambda: messagebox.showerror(
                    "Error", f"Could not scan folders:\n{err_msg}\n\n"
                             "Check that Outlook is open & signed in, and pywin32 is installed."
                ))
                return
            # MỚI: các phiếu vote đã SỬA TAY trước đó (self.responses[...]
            # ["manual"] == True, xem _commit_response_vote_edit()) KHÔNG
            # được để lần Scan Inbox này ghi đè mất — GIỮ LẠI phiếu sửa tay,
            # TRỪ KHI vừa quét được 1 email vote MỚI HƠN đúng người đó (so
            # theo "received" — nghĩa là họ đã thực sự bấm Vote lại sau thời
            # điểm bạn sửa tay, nên coi email mới đó là quyết định mới nhất,
            # đáng tin hơn). Không có mốc thời gian nào để so (thiếu
            # "received" ở 1 trong 2 phía) thì ưu tiên GIỮ bản sửa tay, an
            # toàn hơn là để mất 1 sửa đổi thủ công một cách âm thầm.
            merged = dict(responses)
            for email_key, prev in self.responses.items():
                if not prev.get("manual"):
                    continue
                new_scan = merged.get(email_key)
                prev_time = prev.get("received")
                new_time = new_scan.get("received") if new_scan else None
                keep_manual = (
                    new_scan is None or prev_time is None or new_time is None or new_time <= prev_time
                )
                if keep_manual:
                    merged[email_key] = prev
            self.responses = merged
            # Ghi lại ĐÚNG Event ID vừa quét + thời điểm quét, để banner trạng
            # thái ở đầu Tab 4 báo được chính xác dữ liệu bên dưới thuộc về sự
            # kiện nào — tránh hiểu nhầm "dữ liệu event mới" trong khi thực ra
            # đó là kết quả CŨ còn sót lại từ lần Scan của 1 Event ID khác.
            self._last_scanned_event_id = event_id
            self._last_scan_time = datetime.now()
            # Tự động mở rộng bất kỳ dòng nào trong Tab 2 là group email (Exchange
            # Distribution List, kể cả sub-group lồng bên trong) thành từng thành
            # viên thật, để khung 'Đã/Chưa phản hồi' liệt kê đúng từng người —
            # NGAY CẢ với sự kiện đã gửi mời từ trước, không cần quay lại Tab 2 bấm
            # '🔎 Expand group emails' rồi gửi lại (xem _build_effective_roster()).
            roster = self._build_effective_roster()
            self.after(0, lambda: self._refresh_response_tree(skipped, roster))
            self.after(0, self._update_scan_status_banner)
            # MỚI: tự động lưu kết quả vừa quét được vào sheet "Responded
            # result" của Attendance_Payment_{EventID}.xlsx ngay sau khi
            # Scan Inbox xong (best-effort, không thông báo lỗi nếu có) —
            # để "Load setup from selected event" ở Tab 1 sau này khôi phục
            # lại được kết quả quét gần nhất mà KHÔNG cần quét lại Outlook.
            self.after(0, lambda: self._save_responded_result_to_file(silent=True))

        threading.Thread(target=worker, daemon=True).start()

    def _update_scan_status_banner(self):
        """Cập nhật dòng trạng thái ở đầu Tab 4 — cho biết bảng bên dưới
        đang hiển thị kết quả quét của Event ID nào, quét lúc nào, và CẢNH
        BÁO RÕ nếu Event ID ở Tab 1 đã đổi khác so với lần quét gần nhất
        (tức dữ liệu đang hiển thị là CŨ, chưa phản ánh sự kiện hiện tại)."""
        current_id = self.var_event_id.get().strip()
        last_id = getattr(self, "_last_scanned_event_id", None)
        last_time = getattr(self, "_last_scan_time", None)
        last_time_str = last_time.strftime("%Y-%m-%d %H:%M") if last_time else "?"

        if not current_id:
            text = "🔍 Current Event ID: (not entered on Tab 1 yet)"
            color = "#8a4b00"
        elif last_id is None:
            text = (f"🔍 Current Event ID: '{current_id}' — NOT scanned yet this session. "
                    f"Click '📨 Scan Inbox for Vote results' to fetch results.")
            color = "#8a4b00"
        elif last_id != current_id:
            text = (f"⚠️ The table below is showing results for the OLD Event ID ('{last_id}', scanned at "
                    f"{last_time_str}) — NOT the current Event ID ('{current_id}'). Click Scan Inbox "
                    f"again to refresh!")
            color = "#c42b1c"
        else:
            text = f"✅ Showing results for Event ID '{current_id}' — scanned at {last_time_str}."
            color = "#0e700e"
        self.lbl_scan_status.config(text=text, foreground=color)

    def _build_effective_roster(self):
        """Trả về roster THỰC TẾ để đối chiếu vote ở Tab 4 — mỗi dòng trong
        Tab 2 là group email (Exchange Distribution List) được TỰ ĐỘNG thay
        bằng các thành viên thật (đệ quy qua sub-group, xem
        outlook_com.expand_group_members()), để khung 'Đã/Chưa phản hồi'
        liệt kê đúng TỪNG NGƯỜI thay vì 1 dòng group email mơ hồ.

        KHÔNG sửa self.recipients (Tab 2 vẫn giữ nguyên như đã lưu) — chỉ áp
        dụng cho việc THEO DÕI/HIỂN THỊ ở Tab 4. Muốn áp dụng vĩnh viễn vào
        chính Tab 2 (vd để lần gửi mời SAU tự đúng luôn từ đầu), dùng nút
        '🔎 Expand group emails in list' + 'Save to Excel' ở Tab 2.

        Kết quả mỗi group được CACHE lại (self._group_expansion_cache) để
        không phải hỏi lại Exchange GAL mỗi lần bấm Scan Inbox trong cùng
        phiên làm việc — chỉ query lần đầu tiên gặp mỗi group email."""
        if not hasattr(self, "_group_expansion_cache"):
            self._group_expansion_cache = {}  # email.lower() -> list[(name,email)] hoặc None

        roster = []
        seen = set()
        for name, email in self.recipients:
            key = email.lower()
            if key not in self._group_expansion_cache:
                try:
                    self._group_expansion_cache[key] = outlook_com.expand_group_members(email)
                except Exception:
                    self._group_expansion_cache[key] = None  # lỗi COM/không resolve được -> coi như người thường

            members = self._group_expansion_cache[key]
            if members is None:
                # Không phải group (hoặc không resolve được) -> giữ nguyên dòng gốc
                if key not in seen:
                    seen.add(key)
                    roster.append((name, email))
            else:
                for m_name, m_email in members:
                    mk = m_email.lower()
                    if mk not in seen:
                        seen.add(mk)
                        roster.append((m_name, m_email))
        return roster

    def _refresh_response_tree(self, skipped=0, roster=None):
        roster = roster if roster is not None else self.recipients
        # MỚI: nhớ lại roster vừa dùng để vẽ bảng — dùng khi sửa tay 1 ô Vote
        # (_commit_response_vote_edit()) cần vẽ lại toàn bộ bảng mà KHÔNG phải
        # tính lại _build_effective_roster() (có thể gọi COM để mở rộng group
        # email, không nên chạy lại chỉ vì sửa 1 ô).
        self._last_responses_roster = roster
        self.tree_responses.delete(*self.tree_responses.get_children())
        self.tree_responses.tag_configure("extra", background="#fff3cd")
        # MỚI: tô nền xanh nhạt cho các dòng có phiếu vote đã được SỬA TAY
        # (double-click ô Vote — xem _commit_response_vote_edit()), để phân
        # biệt trực quan với phiếu quét được thật sự từ email.
        self.tree_responses.tag_configure("manual", background="#dceeff")
        counts = {"Yes": 0, "No": 0, "Maybe": 0, "No response": 0}
        matched_emails = set()
        pending = []  # list[(name, email)] — người trong roster mà CHƯA vote gì
        for name, email in roster:
            email_key = email.lower()
            matched_emails.add(email_key)
            r = self.responses.get(email_key)
            if r:
                vote = r["vote"]
                received = r["received"].strftime("%Y-%m-%d %H:%M") if r.get("received") else ""
                display_name = r.get("name") or name
                counts[vote] = counts.get(vote, 0) + 1
                tags = ("manual",) if r.get("manual") else ()
            else:
                vote, received, display_name = "No response", "", name
                counts["No response"] += 1
                pending.append((name, email))
                tags = ()
            manual_display = "✅" if (r and r.get("manual")) else "⬜"
            self.tree_responses.insert("", "end", values=(manual_display, display_name, email, vote, received),
                                        tags=tags)

        # BUG ĐÃ SỬA: trước đây, bất kỳ ai bấm vote mà KHÔNG có mặt trong danh
        # sách Recipients (Tab 2) — vd: thành viên của 1 group email đã gửi
        # tới qua "Send to override" ở Tab 3, mỗi người trả lời bằng chính hộp
        # thư cá nhân của họ chứ không phải địa chỉ group — thì phiếu vote đó
        # ĐÃ được tool tìm thấy (nằm sẵn trong self.responses) nhưng bị ÂM
        # THẦM BỎ QUA lúc hiển thị, vì vòng lặp trên chỉ duyệt qua roster.
        # Giờ `roster` đã tự MỞ RỘNG group email thành từng thành viên thật
        # (xem _build_effective_roster), nên phần lớn trường hợp này KHÔNG
        # còn xảy ra nữa — nhưng vẫn giữ lại khối bên dưới làm lưới an toàn
        # cho những phiếu vote thật sự "lạ" (vd trả lời từ 1 địa chỉ hoàn
        # toàn không liên quan tới roster/group đã biết). Hiển thị thêm các
        # phiếu "ngoài danh sách" này (tô màu vàng nhạt để phân biệt) và
        # CỘNG luôn vào Yes/No/Maybe — vì đó vẫn là phiếu vote thật, hợp lệ,
        # khớp đúng Event ID — để tổng số phản ánh đúng thực tế.
        extra_rows = []
        for email, r in self.responses.items():
            if email in matched_emails:
                continue
            vote = r["vote"]
            received = r["received"].strftime("%Y-%m-%d %H:%M") if r.get("received") else ""
            display_name = (r.get("name") or email) + "  ⚠ (outside known list/group)"
            counts[vote] = counts.get(vote, 0) + 1
            # Dòng "extra" (ngoài danh sách) ưu tiên tag "extra" (vàng nhạt) —
            # nếu vừa "extra" vừa "manual", vẫn hiện "extra" vì đó là tín hiệu
            # quan trọng hơn cần chú ý (ai đó không có trong roster).
            tags = ("manual",) if r.get("manual") else ("extra",)
            manual_display = "✅" if r.get("manual") else "⬜"
            extra_rows.append((email, manual_display, display_name, vote, received, tags))
        for email, manual_display, display_name, vote, received, tags in extra_rows:
            self.tree_responses.insert("", "end", values=(manual_display, display_name, email, vote, received),
                                        tags=tags)

        extra_note = (f"  |  ➕ {len(extra_rows)} people outside the list/group also responded "
                      f"(added to the counts below)") if extra_rows else ""
        group_note = "" if len(roster) == len(self.recipients) else \
            f"  (auto-expanded {len(self.recipients)} Tab 2 row(s) → {len(roster)} actual people)"
        self.lbl_summary.config(
            text=(f"Total tracked: {len(roster)}{group_note}  |  ✅ Yes: {counts.get('Yes',0)}  |  "
                  f"❌ No: {counts.get('No',0)}  |  ❔ Maybe: {counts.get('Maybe',0)}  |  "
                  f"⬜ No response: {counts.get('No response',0)}"
                  + extra_note
                  + (f"  |  (⚠ {skipped} emails matched the Event ID but were not valid votes)" if skipped else ""))
        )
        self._refresh_calendar_yes_list()
        self._refresh_pending_panel(pending)

    def _on_response_tree_double_click(self, event):
        """MỚI: double-click ô "Vote" trên bảng Tab 4 mở 1 dropdown Yes/No/
        Maybe để sửa tay — dùng cho những người chỉ đổi ý hoặc báo miệng
        thay vì bấm lại nút Vote trong email. Các cột khác (Name/Email/
        Received At) KHÔNG cho sửa trực tiếp ở đây (Name/Email vốn lấy từ
        Tab 2, sửa nhầm dễ làm lệch với danh sách gốc)."""
        tree = self.tree_responses
        region = tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        col = tree.identify_column(event.x)
        row_id = tree.identify_row(event.y)
        if not row_id or not col:
            return
        columns = tree["columns"]
        try:
            col_index = int(col.replace("#", "")) - 1
        except ValueError:
            return
        if col_index < 0 or col_index >= len(columns):
            return
        col_name = columns[col_index]
        if col_name != "vote":
            return
        self._begin_cell_edit_combobox(tree, row_id, col_name, ["Yes", "No", "Maybe"],
                                        self._commit_response_vote_edit)

    def _commit_response_vote_edit(self, row_id, col_name, new_value):
        """Ghi lại 1 phiếu vote SỬA TAY vào self.responses (dict nguồn dữ
        liệu chung, xem đầu class) rồi vẽ lại TOÀN BỘ Tab 4 — vì mọi nơi
        khác đọc vote (Tab 5 Yes/Maybe list qua _refresh_calendar_yes_list(),
        Tab 5 Attendance & Payment qua _refresh_attendance_list(), khung
        "Chưa phản hồi"/banner deadline qua _refresh_pending_panel()) đều
        LẤY DỮ LIỆU TỪ CHÍNH self.tree_responses/self.responses, nên chỉ cần
        vẽ lại đúng 1 chỗ này là toàn bộ các tab liên quan tự động khớp
        theo — KHÔNG cần sửa gì thêm ở nơi khác. Sửa tay cũng auto-save
        xuống database ngay (giống mọi auto-save khác trong app), và đánh
        dấu "manual": True để lần "📨 Scan Inbox for Vote results" tiếp
        theo KHÔNG vô tình ghi đè mất phiếu sửa tay này (xem
        _collect_responses() — chỉ ghi đè nếu tìm thấy email vote MỚI HƠN
        thời điểm sửa tay)."""
        new_value = (new_value or "").strip()
        if new_value not in ("Yes", "No", "Maybe"):
            return
        # Đọc lại email/name TRỰC TIẾP từ chính ô đang hiển thị trên dòng đó
        # (không dựa vào row_id/iid) — đơn giản và luôn đúng bất kể iid được
        # Tkinter tự sinh ra là gì.
        email = self.tree_responses.set(row_id, "email").strip()
        if not email:
            return
        email_key = email.lower()
        display_name = self.tree_responses.set(row_id, "name")
        # Bỏ hậu tố cảnh báo "⚠ (outside known list/group)" (nếu có, xem
        # _refresh_response_tree()) trước khi lưu — hậu tố đó chỉ để HIỂN
        # THỊ, không phải 1 phần của tên thật.
        display_name = display_name.split("  ⚠")[0].strip()
        existing = self.responses.get(email_key, {})
        self.responses[email_key] = {
            "name": existing.get("name") or display_name or email,
            "vote": new_value,
            "received": datetime.now(),  # thời điểm SỬA TAY — dùng để so sánh "mới hơn hay cũ hơn" ở lần Scan sau
            "manual": True,
        }
        self._refresh_response_tree(roster=getattr(self, "_last_responses_roster", None))
        self._save_responded_result_to_file(silent=True)

    def _on_response_manual_check_click(self, event):
        """MỚI: single-click ô "Manual edit" (cột checkbox ✅/⬜, đầu tiên
        bên trái "Name") trên Tab 4:
          - Dòng CHƯA sửa tay (⬜) -> mở NGAY dropdown Yes/No/Maybe ở ô
            "Vote" cùng dòng đó để chọn giá trị mới (tick + sửa cùng lúc,
            đúng yêu cầu "tick chọn rồi manual cho phần vote") — commit
            qua chính _commit_response_vote_edit() ở trên, hàm đó tự đặt
            "manual": True nên ô checkbox sẽ tự chuyển thành ✅ sau khi
            bảng được vẽ lại.
          - Dòng ĐÃ sửa tay (✅) -> bấm lại để BỎ tick, chỉ xoá cờ "manual"
            (giữ nguyên giá trị Vote hiện tại) — để lần Scan Inbox tiếp
            theo được tự do cập nhật lại phiếu này từ email thật, không
            còn bị khoá bởi bản sửa tay cũ nữa."""
        tree = self.tree_responses
        region = tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        col = tree.identify_column(event.x)
        row_id = tree.identify_row(event.y)
        if not row_id or col != "#1":  # "#1" = cột "manual_edit" (cột đầu tiên)
            return
        email = tree.set(row_id, "email").strip()
        if not email:
            return
        email_key = email.lower()
        info = self.responses.get(email_key)
        if info and info.get("manual"):
            # Đang ✅ -> bấm để bỏ tick: chỉ xoá cờ manual, GIỮ NGUYÊN vote.
            info["manual"] = False
            self._refresh_response_tree(roster=getattr(self, "_last_responses_roster", None))
            self._save_responded_result_to_file(silent=True)
        else:
            # Đang ⬜ -> bấm để tick: mở dropdown Vote ngay để chọn giá trị
            # sửa tay (commit sẽ tự đặt manual=True, xem _commit_response_vote_edit()).
            self._begin_cell_edit_combobox(tree, row_id, "vote", ["Yes", "No", "Maybe"],
                                            self._commit_response_vote_edit)

    def _refresh_pending_panel(self, pending):
        """Cập nhật khung 'Chưa phản hồi' phía dưới + banner hạn phản hồi +
        nút gửi nhắc nhở, dựa trên `pending` (list[(name,email)]) vừa tính
        được từ _refresh_response_tree(). Gọi mỗi lần Scan Inbox xong, đúng
        như yêu cầu 'khi user bấm Scan inbox, cập nhật cả 2 khung'."""
        self._pending_recipients = pending

        self.tree_pending.delete(*self.tree_pending.get_children())
        for name, email in pending:
            self.tree_pending.insert("", "end", values=(name, email))

        # ── banner hạn phản hồi (Response Deadline ở Tab 1) ──
        deadline_str = get_date_str(self.date_deadline)
        overdue = False
        try:
            d, m, y = deadline_str.split("/")
            deadline_date = datetime(int(y), int(m), int(d)).date()
            overdue = datetime.now().date() >= deadline_date
        except Exception:
            deadline_date = None

        if not pending:
            banner = f"✅ Response deadline: {deadline_str}  —  everyone in Tab 2 has responded."
        elif deadline_date is None:
            banner = f"🕓 {len(pending)} people still haven't responded. (Couldn't read deadline date: '{deadline_str}')"
        elif overdue:
            banner = (f"🔴 Response deadline REACHED/PASSED ({deadline_str}) — {len(pending)} people still haven't "
                       f"replied. Review the reminder content below, then click send.")
        else:
            days_left = (deadline_date - datetime.now().date()).days
            banner = (f"🕓 Response deadline: {deadline_str} ({days_left} days left) — "
                      f"{len(pending)} people still haven't replied.")
        self.lbl_deadline_banner.config(text=banner)

        self.lbl_reminder_send.config(
            text=f"📨 Send reminder email to {len(pending)} people who haven't responded")

        # Tự động tạo sẵn nội dung nhắc nhở lần đầu (nếu ô đang trống), để
        # người dùng có ngay bản nháp mà không cần bấm "Tạo lại" trước.
        if not self.txt_reminder_body.get("1.0", "end").strip():
            self._generate_reminder_draft()

    def _lookup_sent_date_hint(self, event_id):
        """Tra cột SentDate trong RSVP_History.xlsx theo EventID — dùng làm
        hint để chọn ĐÚNG email trong Sent Items khi có nhiều email cùng
        chứa event_id trong Subject (vd: gửi nhắc nhiều lần). Dùng chung cho
        cả '📅 Send Calendar Invite' (Tab 5) và '📨 Gửi nhắc nhở' (Tab 4).
        Trả về datetime hoặc None nếu không tìm thấy/không đọc được."""
        if not event_id:
            return None
        try:
            for rec in db.load_history(self.history_path.get()):
                if rec.get("EventID") == event_id:
                    raw = rec.get("SentDate")
                    if raw:
                        raw_str = str(raw).strip()
                        for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S"):
                            try:
                                return datetime.strptime(raw_str, fmt)
                            except ValueError:
                                continue
                    break
        except Exception:
            pass  # best-effort — History không đọc được thì bỏ qua hint
        return None

    def _generate_reminder_draft(self):
        """Tự động soạn sẵn nội dung email nhắc nhở dựa trên thông tin sự
        kiện đang có ở Tab 1 (Event Name/Date/Location/Deadline/Budget) +
        ngôn ngữ đang chọn ở combo_reminder_lang (English/Japanese/
        Vietnamese/Bilingual — giống 4 lựa chọn ở Tab 3). Người dùng xem/sửa
        tay trong ô bên dưới trước khi gửi — hàm này chỉ tạo BẢN NHÁP, không
        tự gửi gì cả."""
        event_name = self.var_event_name.get()
        event_date = get_date_str(self.date_event)
        location = self.var_location.get()
        deadline = get_date_str(self.date_deadline)
        budget = self.var_budget.get()

        lang_label = self.combo_reminder_lang.get() or LANG_LABELS["en"]
        lang_code = LANG_LABEL_TO_CODE.get(lang_label, "en")

        draft = build_reminder_body(lang_code, event_name, event_date, location, deadline, budget)
        self.txt_reminder_body.delete("1.0", "end")
        self.txt_reminder_body.insert("1.0", draft)

    def _send_reminder(self):
        pending = getattr(self, "_pending_recipients", []) or []
        if not pending:
            messagebox.showinfo(
                "No one left to remind",
                "Everyone in Tab 2 has already responded (Yes/No/Maybe) — no need to send a reminder.\n\n"
                "If you think there are still people who haven't replied, click "
                "'📨 Scan Inbox for Vote results' again first to refresh."
            )
            return

        body = self.txt_reminder_body.get("1.0", "end").strip()
        if not body:
            messagebox.showwarning(
                "Empty content",
                "Click '🔄 Regenerate reminder text' or type the content by hand before sending.")
            return

        event_id = self.var_event_id.get().strip()
        event_name = self.var_event_name.get()
        lang_label = self.combo_reminder_lang.get() or LANG_LABELS["en"]
        lang_code = LANG_LABEL_TO_CODE.get(lang_label, "en")
        # Subject PHẢI vẫn chứa đúng event_id (giống mail gốc) — để lần Scan
        # Inbox sau còn quét/khớp được các phiếu vote mới trả lời trên chính
        # email nhắc nhở này (cơ chế quét chỉ dựa vào event_id nằm trong
        # Subject, không phân biệt đó là mail mời gốc hay mail nhắc, và
        # không phân biệt ngôn ngữ — build_reminder_subject() luôn giữ
        # nguyên "[Reminder-{event_id}]"/"【リマインド-{event_id}】" ở đầu).
        subject = build_reminder_subject(lang_code, event_id, event_name)

        attach = self.var_reminder_attach.get()
        # MỚI: LUÔN False — mọi email nhắc nhở giờ chỉ mở Outlook để review,
        # không còn tuỳ chọn tự động gửi thẳng từ trong app (xem ghi chú ở
        # nơi checkbox cũ đã bị bỏ, phía trên trong _build_tab_collect()).
        auto_send = False
        sent_date_hint = self._lookup_sent_date_hint(event_id) if attach else None
        pending_count = len(pending)

        def worker():
            try:
                mail, attached = outlook_com.send_reminder_email(
                    pending, subject, body,
                    auto_send=auto_send,
                    attach_event_id=(event_id if attach else None),
                    attach_hint_datetime=sent_date_hint,
                )
                if attach and attached:
                    attach_note = "\n\n📎 Found and attached the original invite email."
                elif attach:
                    attach_note = (
                        "\n\n⚠️ Couldn't find the original invite email in Sent Items to attach "
                        "(it may have been deleted/moved, or sent from a different account) — the "
                        "reminder email was still created, just without an attachment."
                    )
                else:
                    attach_note = ""
                action = "opened"
                review_note = " Review it, then click Send in Outlook."

                # MỚI: luôn ghi lại thời điểm BẤM gửi/mở email nhắc nhở (bất
                # kể auto_send hay review mode) vào cột "LastReminderSentDate"
                # — giống cách _send_invite() ghi SentDate ngay lúc gửi/mở,
                # để Tab 6 (Event History) luôn có dữ liệu để xem lại, thay
                # vì bị trống như trước đây. Cột này CHỈ để tra cứu lịch sử,
                # KHÔNG liên quan tới cơ chế chống gửi trùng của cột
                # "ReminderSent" bên dưới (cột đó vẫn CHỈ set khi auto_send=
                # True chắc chắn đã gửi, giữ nguyên logic cũ cho script nền).
                history_log_note = ""
                try:
                    db.save_event_record(
                        {"EventID": event_id,
                         "LastReminderSentDate": datetime.now().strftime("%Y-%m-%d %H:%M")},
                        self.history_path.get()
                    )
                    self.after(0, self._refresh_history_tree)
                except Exception:
                    history_log_note = ("\n\n⚠️ Couldn't write the reminder-sent time to "
                                         "the database — "
                                         "the email was still " + action + " normally.")

                # Chỉ đánh dấu "ReminderSent" trong History khi email THỰC SỰ
                # đã được gửi (auto_send=True -> mail.Send() đã chạy) — nếu
                # chỉ Display() để review thì chưa chắc bạn sẽ bấm Send, nên
                # KHÔNG đánh dấu (để tránh script chạy nền hiểu nhầm là đã
                # nhắc rồi trong khi thực ra chưa ai nhận được gì). Việc này
                # giúp send_scheduled_reminders.py không gửi TRÙNG LẶP nếu
                # bạn đã tự nhắc tay trước khi tới ngày deadline.
                mark_note = ""
                if auto_send:
                    try:
                        for rec in db.load_history(self.history_path.get()):
                            if rec.get("EventID") == event_id:
                                rec["ReminderSent"] = datetime.now().strftime("%Y-%m-%d %H:%M")
                                db.save_event_record(rec, self.history_path.get())
                                mark_note = ("\n\nMarked in the database that this event HAS "
                                             "been reminded — the background script (if Task "
                                             "Scheduler is set up) will not send it again.")
                                break
                    except Exception:
                        pass  # best-effort — không có History record thì thôi, không chặn việc gửi

                self.after(0, lambda: messagebox.showinfo(
                    "Done",
                    f"{action.capitalize()} a reminder email for {pending_count} people who haven't responded."
                    + review_note + attach_note + mark_note + history_log_note))
            except Exception as e:
                err_msg = str(e)
                self.after(0, lambda: messagebox.showerror(
                    "Error", f"Couldn't send the reminder email:\n{err_msg}"))

        threading.Thread(target=worker, daemon=True).start()

    def _export_report(self):
        if not self.recipients:
            messagebox.showwarning("No data", "There is no list/response data to export yet.")
            return
        event_id = self.var_event_id.get()
        out_path = f"{event_id}_Report.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Details"
        headers = ["Name", "Email", "Vote", "Received At"]
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="003366")
            c.alignment = Alignment(horizontal="center")
        colors = {"Yes": "C6EFCE", "No": "FFC7CE", "Maybe": "FFEB9C", "No response": "F0F0F0"}
        row = 2
        for iid in self.tree_responses.get_children():
            _manual, name, email, vote, received = self.tree_responses.item(iid, "values")
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=email)
            vc = ws.cell(row=row, column=3, value=vote)
            vc.fill = PatternFill("solid", fgColor=colors.get(vote, "FFFFFF"))
            ws.cell(row=row, column=4, value=received)
            row += 1
        for col, w in zip("ABCD", [26, 32, 14, 18]):
            ws.column_dimensions[col].width = w
        wb.save(out_path)
        abs_path = os.path.abspath(out_path)
        self._last_report_path = out_path
        # BUG ĐÃ SỬA: trước đây hộp thoại chỉ hiện tên file ngắn (vd
        # "Farewell_Tu_Jul26_Report.xlsx"), không cho biết nó nằm ở THƯ MỤC
        # nào — vì out_path là đường dẫn TƯƠNG ĐỐI, Python lưu vào đúng thư
        # mục mà app đang chạy (working directory) — thường là thư mục chứa
        # rsvp_app.py + RSVP_History.xlsx, nhưng không phải lúc nào cũng rõ
        # ràng với người dùng. Giờ hiện luôn đường dẫn ĐẦY ĐỦ + có nút mở
        # thẳng thư mục đó trong File Explorer.
        self._show_export_done_dialog(abs_path)

    def _show_export_done_dialog(self, abs_path):
        win = tk.Toplevel(self)
        win.title("Exported")
        win.resizable(False, False)
        frame = ttk.Frame(win, padding=16)
        frame.pack(fill="both", expand=True)
        ttk.Label(frame, text="✅ Report saved to:", font=("Arial", 10, "bold"))\
            .pack(anchor="w")
        path_entry = ttk.Entry(frame, width=70)
        path_entry.insert(0, abs_path)
        path_entry.configure(state="readonly")
        path_entry.pack(anchor="w", pady=(4, 10), fill="x")

        btns = ttk.Frame(frame)
        btns.pack(anchor="e", fill="x")

        def open_folder():
            folder = os.path.dirname(abs_path)
            try:
                os.startfile(folder)  # chỉ có trên Windows — đúng môi trường chạy app này
            except Exception as e:
                messagebox.showerror("Couldn't open the folder", str(e))

        ttk.Button(btns, text="📂 Open containing folder", command=open_folder).pack(side="left", padx=(0, 6))
        ttk.Button(btns, text="OK", command=win.destroy).pack(side="left")
        win.transient(self)
        win.grab_set()

    def _save_to_history(self):
        # Đảm bảo danh sách người nhận đã được lưu ra Excel với TÊN CHUẨN HOÁ
        # (Participant_List_{EventID}.xlsx) TRƯỚC khi ghi đường dẫn đó vào
        # History — kể cả khi bạn quên bấm '💾 Save to Excel' ở Tab 2 trước.
        # Nhờ vậy cột RecipientFile trong History luôn trỏ đúng tới file mới
        # nhất, và Tab 1 'Load setup from selected event' sau này nạp lại
        # được chính xác.
        if self.recipients:
            self._save_recipients_to_db(silent=True)

        counts = {"Yes": 0, "No": 0, "Maybe": 0, "No response": 0}
        for iid in self.tree_responses.get_children():
            _, _, _, vote, _ = self.tree_responses.item(iid, "values")
            counts[vote] = counts.get(vote, 0) + 1

        record = {
            "EventID": self.var_event_id.get(),
            "EventName": self.var_event_name.get(),
            "EventDate": get_date_str(self.date_event),
            "Deadline": get_date_str(self.date_deadline),
            "Location": self.var_location.get(),
            "Budget": self.var_budget.get(),
            "EmailLanguage": self.combo_email_lang.get(),
            "OrganizerNote": self._source_note_text(),
            "RecipientFile": self.recipient_file.get(),
            "TotalInvited": len(self.recipients),
            "Yes": counts.get("Yes", 0),
            "No": counts.get("No", 0),
            "Maybe": counts.get("Maybe", 0),
            "NoResponse": counts.get("No response", 0),
            "ReportFile": getattr(self, "_last_report_path", ""),
            "CalendarSent": getattr(self, "_calendar_sent_flag", "Not sent"),
            # MỚI: đã BỎ 5 key "ActualAttendees"/"CostPerPerson"/
            # "TotalIncome"/"TotalExpense"/"Balance" khỏi record — khối UI
            # nhập tay các số liệu này ở Tab 4 đã bị xoá (thay bằng Tab 5
            # "Attendance & Payment" chi tiết hơn). KHÔNG ghi đè các cột
            # này thành rỗng — save_event_record() dùng merge-preserve
            # (chỉ cập nhật cột CÓ MẶT trong record), nên bỏ hẳn key ra
            # khỏi đây giữ nguyên giá trị CŨ (nếu có, từ trước khi tính
            # năng này bị gỡ) thay vì xoá mất.
            "EventMode": self.var_event_mode.get(),
            "Organizer": self.var_organizer.get(),
            "GuestOfHonor": self.var_guest_of_honor.get(),
            "GiftBudget": self.var_gift_budget.get(),
            "GiftDeadline": get_date_str(self.date_gift_deadline),
            "StartTime": self.var_start_time.get(),
            "EndTime": self.var_end_time.get(),
        }
        # Use the ACTUAL time the confirmation email was sent (Tab 3), not
        # whenever this "Save event to History" button happens to be clicked —
        # those can differ by minutes, hours, or even days. Only set the key
        # if we actually tracked a send this session; otherwise leave it out
        # so db.save_event_record()'s existing "now" fallback still
        # applies (e.g. re-saving an older event without re-sending).
        if getattr(self, "_invite_sent_date", None):
            record["SentDate"] = self._invite_sent_date.strftime("%Y-%m-%d %H:%M")
        db.save_event_record(record, self.history_path.get())
        self._refresh_history_combo()
        self._refresh_history_tree()
        # MỚI: "Save event" giờ không chỉ lưu vào History mà còn lưu snapshot
        # kết quả Responded (Tab 4) vào sheet "Responded result" của
        # Attendance_Payment_{EventID}.xlsx — để "Load setup from selected
        # event" ở Tab 1 sau này khôi phục lại được mà KHÔNG cần quét lại
        # Outlook (xem _load_from_history()). Cũng lưu lại luôn bảng
        # Attendance & Payment (Tab 5) hiện tại, phòng khi có thay đổi chưa
        # kịp auto-save (best-effort, không chặn nếu lỗi).
        self._save_responded_result_to_file(silent=True)
        self._save_attendance_sheet_to_file(silent=True)
        messagebox.showinfo("Saved", f"Event '{record['EventID']}' saved to history.")

    # ══════════════════════════════════════════════════════════════════
    # TAB 5 — SEND MEETING (CALENDAR INVITE) TO EVERYONE WHO VOTED "YES"
    # ══════════════════════════════════════════════════════════════════
    def _build_tab_gift(self):
        f = self.tab_gift.body
        pad = {"padx": 10, "pady": 6}

        ttk.Label(f, text="Gift Contribution tracking", font=("Arial", 11, "bold"))\
            .grid(row=0, column=0, columnspan=3, sticky="w", **pad)
        self._make_wrapping_label(f, text="Recipient list pulled from Tab 2 (Recipients). Check off everyone who has\n"
                          "ALREADY contributed money for the gift — the checked state is saved to the\n"
                          "database automatically as soon as you click, and reloaded automatically if\n"
                          "there's saved data (nothing is lost when you close and reopen the app).",
                  font=("Arial", 9, "italic")).grid(row=1, column=0, columnspan=3, sticky="w", padx=10)

        btns = ttk.Frame(f)
        btns.grid(row=2, column=0, columnspan=3, sticky="w", **pad)
        ttk.Button(btns, text="🔄 Reload list from Tab 2", command=self._refresh_gift_contribution_list)\
            .pack(side="left", padx=(0, 6))
        # Import an EXISTING Gift_Contribution_List_*.xlsx file — used when:
        # (1) you want to re-sync checked state from a file that was edited
        # by hand outside the app (e.g. a colleague added notes directly in
        # Excel), or (2) the file isn't named with the current Event ID, so
        # it wasn't picked up automatically by "🔄 Reload list from Tab 2".
        # Unlike the auto-load on tab open (only runs once, and only fills
        # in people who aren't already in the current list), this button
        # ACTIVELY overwrites the checked state for EVERYONE found in the
        # chosen file.
        ttk.Button(btns, text="📂 Load contribution list from file", command=self._load_gift_list_from_file)\
            .pack(side="left", padx=6)
        # MỚI: đã BỎ 2 nút "☑ Check all"/"☐ Uncheck all" — thay bằng 1 dấu
        # ✅/⬜ NGAY TRÊN ĐẦU CỘT "Send email" và cột "Contributed" (bấm
        # trực tiếp vào tiêu đề cột, xem tree.heading(..., command=...) bên
        # dưới và _toggle_all_gift_column()) — mỗi cột tự chọn/bỏ chọn tất
        # cả ĐỘC LẬP với nhau, và vẫn chỉ áp dụng cho các dòng ĐANG HIỂN THỊ
        # (tôn trọng ô tìm kiếm, giống hành vi 2 nút cũ).

        # Quick Name search box — filters the table below directly as you
        # type, case-insensitive. The checked state stays stored in
        # self._gift_roster (the source of truth), so filtering/searching
        # never loses anyone's checked state, even while they're hidden.
        search_bar = ttk.Frame(f)
        search_bar.grid(row=3, column=0, columnspan=3, sticky="w", **pad)
        ttk.Label(search_bar, text="🔎 Search (name):").pack(side="left")
        self.var_gift_search = tk.StringVar(value="")
        ttk.Entry(search_bar, textvariable=self.var_gift_search, width=40).pack(side="left", padx=6)
        self.var_gift_search.trace_add("write", lambda *a: self._apply_gift_filter())
        ttk.Button(search_bar, text="✕ Clear", command=lambda: self.var_gift_search.set(""))\
            .pack(side="left", padx=4)

        # Column order: send_email, check, No. (sequence number), Name, Email, Amount.
        # MỚI: "send_email" — cột checkbox THÊM MỚI, đặt bên trái "check"
        # (Contributed) — người dùng tick chọn AI sẽ nhận email báo cáo
        # quyên góp (xem khối "📧 Send contribution report" ở cuối tab, và
        # _send_gift_report_email()) — HOÀN TOÀN ĐỘC LẬP với "Contributed"
        # (1 người có thể được chọn nhận báo cáo dù họ CHƯA đóng góp, hoặc
        # ngược lại).
        cols = ("send_email", "check", "no", "name", "email", "amount")
        tree_container, self.tree_gift = make_scrollable_treeview(f, columns=cols, height=16)
        # Tiêu đề 2 cột checkbox ("send_email"/"check") tự hiển thị ✅/⬜
        # phản ánh "tất cả dòng ĐANG HIỂN THỊ có đang được tick hết không"
        # — bấm vào TIÊU ĐỀ (không phải 1 ô cụ thể) để chọn/bỏ chọn tất cả
        # cùng lúc cho đúng cột đó (xem _toggle_all_gift_column()).
        self.tree_gift.heading("send_email", text="⬜ Send email",
                                command=lambda: self._toggle_all_gift_column("send_email"))
        self.tree_gift.heading("check", text="⬜ Contributed",
                                command=lambda: self._toggle_all_gift_column("check"))
        self.tree_gift.heading("no", text="No.")
        self.tree_gift.heading("name", text="Name")
        self.tree_gift.heading("email", text="Email")
        self.tree_gift.heading("amount", text="Amount")
        self.tree_gift.column("send_email", width=100, anchor="center")
        self.tree_gift.column("check", width=100, anchor="center")
        self.tree_gift.column("no", width=40, anchor="center")
        self.tree_gift.column("name", width=240)
        self.tree_gift.column("email", width=280)
        self.tree_gift.column("amount", width=100, anchor="e")
        tree_container.grid(row=4, column=0, columnspan=3, sticky="w", padx=10)
        # Clicking the "send_email" or "check" (Contributed) column toggles
        # that row's checked state (tkinter's Treeview has no real checkbox
        # widget, so ✅/⬜ characters simulate one — a common, simple
        # approach that avoids pulling in an extra third-party library).
        self.tree_gift.bind("<Button-1>", self._on_gift_tree_click)

        ttk.Label(f, text="Total contributors:").grid(row=5, column=0, sticky="w", **pad)
        self.var_gift_contributed_count = tk.StringVar(value="0 / 0")
        ttk.Label(f, textvariable=self.var_gift_contributed_count, font=("Arial", 10, "bold"))\
            .grid(row=5, column=1, sticky="w", **pad)

        # Total amount collected so far — sum of "Amount" for everyone
        # currently checked as "Contributed" (see _update_gift_contributed_count()).
        ttk.Label(f, text="Total amount collected:").grid(row=6, column=0, sticky="w", **pad)
        self.var_gift_total_amount = tk.StringVar(value="0")
        ttk.Label(f, textvariable=self.var_gift_total_amount, font=("Arial", 10, "bold"))\
            .grid(row=6, column=1, sticky="w", **pad)

        ttk.Button(f, text="📊 Export to Excel", command=self._export_gift_contribution_list)\
            .grid(row=7, column=0, columnspan=2, sticky="w", **pad)
        self._make_wrapping_label(f, text="→ Contribution data is saved automatically to the database every time you "
                          "check/uncheck someone — this button only creates an Excel file when you need "
                          "one to share or archive. Use '📂 Load contribution list from file' above to "
                          "import/merge from a specific Excel file instead.",
                  font=("Arial", 8, "italic")).grid(row=8, column=0, columnspan=3, sticky="w", padx=14)

        # ── Compose & send a reminder email to people who haven't contributed yet ──
        # Always ONLY opens Outlook for review (mail.Display()) — there is no
        # "Send immediately without review" option here (unlike Tab 4's RSVP
        # reminder), per the requirement that every reminder email (both RSVP
        # and Gift) must be checked by the user and sent by hand from
        # Outlook, never sent automatically from within the app. See
        # _send_gift_reminder().
        gift_reminder_frame = ttk.LabelFrame(
            f, text="📨 Compose & send a reminder email to people who haven't contributed yet")
        gift_reminder_frame.grid(row=9, column=0, columnspan=3, sticky="we", padx=10, pady=6)

        grbtn = ttk.Frame(gift_reminder_frame)
        grbtn.pack(fill="x", padx=6, pady=4)
        ttk.Label(grbtn, text="Language:").pack(side="left", padx=(0, 4))
        self.combo_gift_reminder_lang = ttk.Combobox(
            grbtn, width=26, state="readonly",
            values=[LANG_LABELS["en"], LANG_LABELS["ja"], LANG_LABELS["vi"], LANG_LABELS["bilingual"]],
        )
        self.combo_gift_reminder_lang.current(0)  # default: English
        self.combo_gift_reminder_lang.pack(side="left", padx=(0, 6))
        self.combo_gift_reminder_lang.bind(
            "<<ComboboxSelected>>", lambda e: self._generate_gift_reminder_draft())
        ttk.Button(grbtn, text="🔄 Regenerate reminder text (from Tab 1)",
                   command=self._generate_gift_reminder_draft).pack(side="left", padx=(0, 6))
        self.var_gift_reminder_attach = tk.BooleanVar(value=True)
        ttk.Checkbutton(grbtn, text="📎 Attach original Gift email (looked up from Sent Items)",
                        variable=self.var_gift_reminder_attach).pack(side="left", padx=6)

        ttk.Label(gift_reminder_frame, text="Reminder email content (review/edit before sending):",
                  font=("Arial", 9, "italic")).pack(anchor="w", padx=6, pady=(4, 0))
        gift_reminder_text_container, self.txt_gift_reminder_body = make_scrollable_text(
            gift_reminder_frame, width=100, height=7)
        gift_reminder_text_container.pack(fill="x", padx=6, pady=4)

        self.lbl_gift_reminder_send = ttk.Button(
            gift_reminder_frame, text="📨 Open reminder email for 0 people who haven't contributed",
            command=self._send_gift_reminder)
        self.lbl_gift_reminder_send.pack(anchor="w", padx=6, pady=(2, 8))
        ttk.Label(gift_reminder_frame,
                  text="→ Always opens the email compose window in Outlook so you can review it and "
                       "click Send YOURSELF — there is no auto-send option here.",
                  font=("Arial", 8, "italic")).pack(anchor="w", padx=6, pady=(0, 6))

        # ── MỚI: gửi email báo cáo số tiền đã thu được, tới NHỮNG NGƯỜI ĐÃ
        # TICK CHỌN ở cột "Send email" (hoàn toàn tách biệt với người CHƯA
        # đóng góp ở khối nhắc nhở phía trên) — nội dung email liệt kê danh
        # sách những ai ĐÃ ĐÓNG GÓP (cột "Contributed" = Yes), đánh số lại
        # từ 1, KHÔNG đưa 2 cột checkbox ("Send email"/"Contributed") vào
        # danh sách đó — xem _build_gift_report_workbook()/build_gift_report_body().
        # Nội dung mail CHỈ là 1 thông báo TỔNG QUAN (đã thu bao nhiêu
        # người/bao nhiêu tiền), KHÔNG liệt kê từng người trong nội dung —
        # danh sách chi tiết nằm trong file Excel TỰ ĐỘNG ĐÍNH KÈM.
        report_frame = ttk.LabelFrame(f, text="📧 Send contribution report to selected people")
        report_frame.grid(row=10, column=0, columnspan=3, sticky="we", padx=10, pady=6)
        self._make_wrapping_label(
            report_frame,
            text="Sent to everyone ticked in the \"Send email\" column above. The message body is a "
                 "short summary (how many people contributed, total amount) — the detailed list "
                 "(everyone who has ALREADY contributed, renumbered, without the \"Send email\"/"
                 "\"Contributed\" checkbox columns) is automatically attached as an Excel file.",
            font=("Arial", 8, "italic")).pack(anchor="w", padx=6, pady=(4, 0))

        rbtn = ttk.Frame(report_frame)
        rbtn.pack(fill="x", padx=6, pady=4)
        ttk.Label(rbtn, text="Language:").pack(side="left", padx=(0, 4))
        self.combo_gift_report_lang = ttk.Combobox(
            rbtn, width=26, state="readonly",
            values=[LANG_LABELS["en"], LANG_LABELS["ja"], LANG_LABELS["vi"], LANG_LABELS["bilingual"]],
        )
        self.combo_gift_report_lang.current(0)  # default: English
        self.combo_gift_report_lang.pack(side="left", padx=(0, 6))
        self.combo_gift_report_lang.bind("<<ComboboxSelected>>", lambda e: self._generate_gift_report_draft())
        ttk.Button(rbtn, text="🔄 Regenerate report text", command=self._generate_gift_report_draft)\
            .pack(side="left", padx=(0, 6))

        ttk.Label(report_frame, text="Report email content (review/edit before sending):",
                  font=("Arial", 9, "italic")).pack(anchor="w", padx=6, pady=(4, 0))
        report_text_container, self.txt_gift_report_body = make_scrollable_text(
            report_frame, width=100, height=10)
        report_text_container.pack(fill="x", padx=6, pady=4)

        self.lbl_gift_report_send = ttk.Button(
            report_frame, text="📧 Send report to 0 selected people", command=self._send_gift_report_email)
        self.lbl_gift_report_send.pack(anchor="w", padx=6, pady=(2, 8))
        ttk.Label(report_frame,
                  text="→ Always opens the email compose window in Outlook so you can review it and "
                       "click Send YOURSELF — there is no auto-send option here.",
                  font=("Arial", 8, "italic")).pack(anchor="w", padx=6, pady=(0, 6))

    def _pending_gift_contributors(self):
        """Returns list[(name, email)] of everyone NOT yet checked as
        "Contributed" in self._gift_roster (the source of truth for Tab 6 —
        independent of whatever the search box is currently filtering)."""
        roster = getattr(self, "_gift_roster", None) or {}
        return [(info["name"], email) for email, info in roster.items() if not info["checked"]]

    def _update_gift_reminder_button_label(self):
        if not hasattr(self, "lbl_gift_reminder_send"):
            return
        count = len(self._pending_gift_contributors())
        self.lbl_gift_reminder_send.config(text=f"📨 Open reminder email for {count} people who haven't contributed")

    def _generate_gift_reminder_draft(self):
        """Auto-composes a draft gift-contribution reminder email using the
        event info currently on Tab 1 (Guest of Honor/Location/Start time/
        Event date/Organizer/Gift deadline/Gift budget) plus the language
        selected in combo_gift_reminder_lang. Only produces a DRAFT — nothing
        is sent — the user reviews/edits it in the box below before sending."""
        guest_of_honor = self.var_guest_of_honor.get()
        start_time = self.var_start_time.get()
        event_date = get_date_str(self.date_event)
        location = self.var_location.get()
        organizer = self.var_organizer.get()
        deadline = get_date_str(self.date_gift_deadline)
        gift_budget = self.var_gift_budget.get()

        lang_label = self.combo_gift_reminder_lang.get() or LANG_LABELS["en"]
        lang_code = LANG_LABEL_TO_CODE.get(lang_label, "en")

        draft = build_gift_reminder_body(lang_code, guest_of_honor, start_time, event_date,
                                          location, organizer, deadline, gift_budget)
        self.txt_gift_reminder_body.delete("1.0", "end")
        self.txt_gift_reminder_body.insert("1.0", draft)

    def _send_gift_reminder(self):
        pending = self._pending_gift_contributors()
        if not pending:
            messagebox.showinfo(
                "No one left to remind",
                "Everyone in the Tab 6 list is already checked as 'Contributed' — no need to send a reminder."
            )
            return

        body = self.txt_gift_reminder_body.get("1.0", "end").strip()
        if not body:
            messagebox.showwarning(
                "Empty content",
                "Click '🔄 Regenerate reminder text' or type the content by hand before sending.")
            return

        event_id = self.var_event_id.get().strip()
        guest_of_honor = self.var_guest_of_honor.get()
        lang_label = self.combo_gift_reminder_lang.get() or LANG_LABELS["en"]
        lang_code = LANG_LABEL_TO_CODE.get(lang_label, "en")
        # Uses its OWN subject prefix "[Reminder-Gift-{event_id}]" — doesn't
        # need to match any Scan Inbox logic (Gift mode has no Voting
        # Buttons / isn't vote-scanned), just needs to be clearly distinct
        # from the original Gift email.
        subject = build_gift_reminder_subject(lang_code, event_id, guest_of_honor)

        attach = self.var_gift_reminder_attach.get()
        pending_count = len(pending)

        def worker():
            try:
                # ALWAYS auto_send=False — only opens the Outlook window for
                # review, there is no direct-send option (unlike Tab 4's
                # RSVP reminder, which used to have a "Send now" checkbox;
                # here it's deliberately absent, per the requirement that
                # every reminder — RSVP or Gift — must be sent by the user
                # clicking Send themselves).
                mail, attached = outlook_com.send_gift_reminder_email(
                    pending, subject, body,
                    auto_send=False,
                    attach_event_id=(event_id if attach else None),
                )
                if attach and attached:
                    attach_note = "\n\n📎 Found and attached the original Gift email."
                elif attach:
                    attach_note = (
                        "\n\n⚠️ Couldn't find the original Gift email in Sent Items to attach "
                        "(it may have been deleted/moved, or sent from a different account) — the "
                        "reminder email was still created, just without an attachment."
                    )
                else:
                    attach_note = ""

                self.after(0, lambda: messagebox.showinfo(
                    "Reminder email opened",
                    f"Opened a reminder email compose window for {pending_count} people who haven't contributed.\n"
                    "Review it, then click Send in Outlook." + attach_note))
            except Exception as e:
                err_msg = str(e)
                self.after(0, lambda: messagebox.showerror(
                    "Error", f"Couldn't open the reminder email:\n{err_msg}"))

        threading.Thread(target=worker, daemon=True).start()

    def _load_gift_list_from_file(self):
        """Lets you pick ANY Excel file by hand (typically an older
        'Gift_Contribution_List_*.xlsx', e.g. exported before switching to
        the database, or a copy edited by a colleague outside the app),
        then MERGES/UPDATES the contribution state into self._gift_roster
        and immediately saves it to the database:
        - Accepts ANY .xlsx file (doesn't have to match the current Event
          ID's standard export filename).
        - OVERWRITES the checked state (and amount) for EVERYONE found in
          the chosen file (not just people who aren't already in the
          current list, like the normal Tab-open reload does).
        - Automatically adds NEW people to the list if the file contains
          someone not currently in Tab 2 (e.g. the Tab 2 list changed since
          the file was exported)."""
        path = filedialog.askopenfilename(
            title="Select a Gift Contribution List file",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not path:
            return
        if not hasattr(self, "_gift_roster"):
            self._gift_roster = {}
        try:
            rows = read_gift_contribution_rows(path)
        except Exception as e:
            messagebox.showerror("Error", f"Couldn't read the file:\n{e}")
            return

        for row in rows:
            existing = self._gift_roster.get(row["email"], {})
            self._gift_roster[row["email"]] = {
                "name": row["name"] or existing.get("name", row["email"]),
                "checked": row["checked"],
                "amount": row["amount"],
                # "send_email" không có trong file Excel (chỉ là lựa chọn
                # riêng của app, xem cột "Send email") — GIỮ NGUYÊN giá trị
                # đã có (nếu người này đã từng được tick chọn nhận báo cáo
                # trước đó), không để việc import file vô tình xoá mất.
                "send_email": existing.get("send_email", False),
            }

        self._apply_gift_filter()
        # Lưu ngay vào database (thay vì ghi lại Excel như kiến trúc cũ) —
        # để lần mở lại Tab 6 sau này tự nạp đúng những gì vừa import.
        self._save_gift_roster_to_db(silent=True)
        messagebox.showinfo(
            "Loaded",
            f"Loaded/updated {len(rows)} people from file:\n{path}\n\n"
            f"Total in list now: {len(self._gift_roster)} people."
        )

    def _build_gift_report_workbook(self):
        """Dựng 1 openpyxl Workbook liệt kê CHỈ những người ĐÃ ĐÓNG GÓP
        (self._gift_roster[...]["checked"] == True), đánh số lại từ 1 —
        dùng làm file đính kèm của email báo cáo (xem
        _send_gift_report_email()). KHÔNG gồm 2 cột checkbox "Send email"/
        "Contributed" (2 cột đó chỉ có ý nghĩa thao tác trên UI, không phải
        dữ liệu cần báo cáo cho người nhận) — chỉ còn No./Name/Email/
        Amount."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contributors"
        headers = ["No.", "Name", "Email", "Amount"]
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="003366")
            c.alignment = Alignment(horizontal="center")
        row_idx = 2
        seq = 0
        total_amount = 0.0
        for email, info in getattr(self, "_gift_roster", {}).items():
            if not info.get("checked"):
                continue
            seq += 1
            amount = info.get("amount", 0.0)
            total_amount += amount
            ws.cell(row=row_idx, column=1, value=seq)
            ws.cell(row=row_idx, column=2, value=info.get("name") or email)
            ws.cell(row=row_idx, column=3, value=email)
            ws.cell(row=row_idx, column=4, value=amount or None)
            row_idx += 1
        total_row = row_idx + 1
        ws.cell(row=total_row, column=2, value="TOTAL COLLECTED:").font = Font(bold=True)
        c_amt = ws.cell(row=total_row, column=4, value=total_amount)
        c_amt.font = Font(bold=True)
        c_amt.number_format = "#,##0"
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 32
        ws.column_dimensions["D"].width = 14
        return wb

    def _gift_report_body_args(self):
        """Lấy đúng (guest_of_honor, event_name, contributor_count,
        total_amount) hiện tại từ Tab 1 + bảng Gift Contribution (Tab 6) —
        dùng để build nội dung email báo cáo mặc định
        (build_gift_report_body()) — nội dung chỉ là 1 THÔNG BÁO TỔNG QUAN,
        danh sách chi tiết nằm trong file Excel đính kèm (xem
        _build_gift_report_workbook())."""
        contributor_count = sum(
            1 for info in getattr(self, "_gift_roster", {}).values() if info.get("checked"))
        return (
            self.var_guest_of_honor.get(),
            self.var_event_name.get(),
            str(contributor_count),
            getattr(self, "var_gift_total_amount", tk.StringVar(value="0")).get(),
        )

    def _generate_gift_report_draft(self):
        """Auto-composes a draft contribution-report email using the event
        info on Tab 1 plus the CURRENT Gift Contribution table (Tab 6) and
        the language selected in combo_gift_report_lang. Only produces a
        DRAFT — nothing is sent — the user reviews/edits it in the box
        below before sending."""
        lang_label = self.combo_gift_report_lang.get() or LANG_LABELS["en"]
        lang_code = LANG_LABEL_TO_CODE.get(lang_label, "en")
        draft = build_gift_report_body(lang_code, *self._gift_report_body_args())
        self.txt_gift_report_body.delete("1.0", "end")
        self.txt_gift_report_body.insert("1.0", draft)

    def _send_gift_report_email(self):
        """MỚI: gửi email báo cáo số tiền đã quyên góp được tới những ai đã
        tick chọn ở cột "Send email" (self._gift_roster[...]["send_email"])
        — HOÀN TOÀN ĐỘC LẬP với cột "Contributed" (người nhận báo cáo
        không nhất thiết phải là người đã đóng góp). Nội dung mail chỉ là
        1 thông báo TỔNG QUAN (đã thu được bao nhiêu người/bao nhiêu tiền)
        — danh sách chi tiết từng người ĐÃ đóng góp (đánh số lại từ 1,
        KHÔNG gồm 2 cột checkbox) nằm trong file Excel TỰ ĐỘNG ĐÍNH KÈM
        (xem _build_gift_report_workbook()). Không có Voting Buttons (dùng
        outlook_com.send_gift_report_email(), hàm gửi thông báo thuần tuý
        + đính kèm file — KHÁC với send_voting_invite() vốn không hỗ trợ
        đính kèm)."""
        recipients = [
            (info.get("name") or email, email)
            for email, info in getattr(self, "_gift_roster", {}).items()
            if info.get("send_email")
        ]
        if not recipients:
            messagebox.showwarning(
                "No recipients selected",
                "Tick the \"Send email\" checkbox (top-left column) for at least one person first — "
                "click the ⬜/✅ mark in the column header to select everyone currently shown at once."
            )
            return

        body = self.txt_gift_report_body.get("1.0", "end").strip()
        if not body:
            messagebox.showwarning(
                "Empty content",
                "Click '🔄 Regenerate report text' or type the content by hand before sending.")
            return

        event_id = self.var_event_id.get().strip()
        guest_of_honor = self.var_guest_of_honor.get()
        lang_label = self.combo_gift_report_lang.get() or LANG_LABELS["en"]
        lang_code = LANG_LABEL_TO_CODE.get(lang_label, "en")
        subject = build_gift_report_subject(lang_code, event_id, guest_of_honor)

        # Chuẩn bị file Excel đính kèm TRƯỚC (trong main thread, vì cần đọc
        # self._gift_roster/các StringVar UI — an toàn hơn đọc chúng từ
        # worker thread), lưu vào 1 file tạm — không phải nơi lưu trữ
        # chính (dữ liệu thật vẫn ở database), chỉ để đính kèm email.
        try:
            wb = self._build_gift_report_workbook()
            excel_path = os.path.join(
                tempfile.gettempdir(), f"Gift_Contribution_Report_{event_id or 'event'}.xlsx")
            wb.save(excel_path)
        except Exception as e:
            messagebox.showerror("Error", f"Couldn't prepare the contribution report attachment:\n{e}")
            return

        recipient_count = len(recipients)

        def worker():
            try:
                mail, attached = outlook_com.send_gift_report_email(
                    recipients, subject, body, excel_path=excel_path)
                attach_note = ("\n\n📎 Attached the contribution report." if attached else
                                "\n\n⚠️ Couldn't attach the contribution report file — the email was "
                                "still created without it.")
                self.after(0, lambda: messagebox.showinfo(
                    "Done",
                    f"Contribution report email opened for {recipient_count} selected people. "
                    "Review it, then click Send in Outlook." + attach_note))
            except Exception as e:
                err_msg = str(e)
                self.after(0, lambda: messagebox.showerror(
                    "Error", f"Could not create the report email:\n{err_msg}"))

        threading.Thread(target=worker, daemon=True).start()

    def _refresh_gift_contribution_list(self):
        """Reloads the Name/Email list from Tab 2 (self.recipients) into
        self._gift_roster — the SOURCE OF TRUTH (not the Treeview itself —
        the Treeview is only a filterable VIEW, see _apply_gift_filter()).
        Priority order for each person's checked state and amount:
        1. Keep whatever is already in self._gift_roster (if it was already
           loaded this session — e.g. a new person just appeared on Tab 2),
           otherwise
        2. If that person has never been in self._gift_roster yet, try
           reading from the database (bảng gift_contributions) — to RESTORE
           the checked state (and amount) from last time, even after
           closing/reopening the app."""
        if not hasattr(self, "tree_gift"):
            return
        if not hasattr(self, "_gift_roster"):
            self._gift_roster = {}  # email -> {"name":..., "checked": bool, "amount": float, "send_email": bool}

        # Step 1: read the saved state from the database as a fallback —
        # ONLY used for people not already in self._gift_roster.
        event_id = self.var_event_id.get().strip()
        saved_state = {}
        if event_id:
            try:
                saved_state = db.load_gift_roster(event_id, self.history_path.get())
            except Exception:
                pass  # best-effort

        new_roster = {}
        for name, email in self.recipients:
            if email in self._gift_roster:
                checked = self._gift_roster[email]["checked"]
                amount = self._gift_roster[email].get("amount", 0.0)
                send_email = self._gift_roster[email].get("send_email", False)
            else:
                prev = saved_state.get(email, {})
                checked = prev.get("checked", False)
                amount = prev.get("amount", 0.0)
                send_email = prev.get("send_email", False)
            new_roster[email] = {"name": name, "checked": checked, "amount": amount, "send_email": send_email}
        self._gift_roster = new_roster
        self._apply_gift_filter()

    def _apply_gift_filter(self, *args):
        """Redraws the Treeview from self._gift_roster, only showing rows
        that match the search box (if any) — this does NOT touch the actual
        checked state (which lives in self._gift_roster), so searching/
        filtering never loses anyone's checked state. The "No." column shows
        each visible row's position (1, 2, 3, ...) in the CURRENTLY SHOWN
        list, so it stays contiguous even while filtering."""
        if not hasattr(self, "tree_gift") or not hasattr(self, "_gift_roster"):
            return
        query = self.var_gift_search.get().strip().lower() if hasattr(self, "var_gift_search") else ""
        self.tree_gift.delete(*self.tree_gift.get_children())
        seq = 0
        for email, info in self._gift_roster.items():
            name = info["name"]
            if query and query not in (name or "").lower():
                continue
            seq += 1
            send_email_display = "✅" if info.get("send_email") else "⬜"
            check = "✅" if info["checked"] else "⬜"
            amount_display = f"{info.get('amount', 0.0):,.0f}" if info.get("amount") else ""
            # Use EMAIL as the Treeview row's iid -> look up/toggle the
            # right person in self._gift_roster directly, without matching
            # displayed Name/Email strings (which could collide).
            self.tree_gift.insert("", "end", iid=email,
                                   values=(send_email_display, check, seq, name, email, amount_display))
        self._update_gift_contributed_count()
        self._update_gift_header_checkmarks()

    def _update_gift_header_checkmarks(self):
        """Cập nhật dấu ✅/⬜ hiển thị NGAY TRÊN TIÊU ĐỀ cột "Send email"/
        "Contributed" — ✅ khi TẤT CẢ dòng ĐANG HIỂN THỊ (tôn trọng ô tìm
        kiếm) đều đã tick ở cột đó, ⬜ nếu không (kể cả khi danh sách đang
        hiển thị rỗng). Gọi lại sau mỗi lần vẽ bảng (_apply_gift_filter())
        để tiêu đề luôn phản ánh đúng trạng thái hiện tại."""
        if not hasattr(self, "tree_gift"):
            return
        shown = self.tree_gift.get_children()
        all_send = bool(shown) and all(
            self._gift_roster.get(iid, {}).get("send_email") for iid in shown)
        all_checked = bool(shown) and all(
            self._gift_roster.get(iid, {}).get("checked") for iid in shown)
        self.tree_gift.heading("send_email", text=("✅" if all_send else "⬜") + " Send email")
        self.tree_gift.heading("check", text=("✅" if all_checked else "⬜") + " Contributed")

    def _on_gift_tree_click(self, event):
        region = self.tree_gift.identify("region", event.x, event.y)
        if region != "cell":
            return
        col = self.tree_gift.identify_column(event.x)
        iid = self.tree_gift.identify_row(event.y)
        if not iid:
            return
        email = iid  # iid IS the email — see _apply_gift_filter()
        if email not in self._gift_roster:
            return
        if col == "#1":  # "send_email" — cột checkbox chọn người nhận báo cáo, ĐỘC LẬP với "checked"
            self._gift_roster[email]["send_email"] = not self._gift_roster[email].get("send_email", False)
        elif col == "#2":  # "check" — cột "Contributed"
            new_checked = not self._gift_roster[email]["checked"]
            self._gift_roster[email]["checked"] = new_checked
            # Auto-fill the Amount from Tab 1's "Expected gift budget" the
            # moment someone is checked as contributed; clear it back to 0 if
            # unchecked, so the total only ever counts people currently marked
            # as having contributed. (The amount can still be corrected later by
            # editing the exported Excel file and reloading it via "📂 Load
            # contribution list from file", if the actual amount differs from
            # the expected budget.)
            if new_checked:
                self._gift_roster[email]["amount"] = parse_amount_from_text(self.var_gift_budget.get())
            else:
                self._gift_roster[email]["amount"] = 0.0
        else:
            return  # cột khác (No./Name/Email/Amount) không tick được bằng click
        self._apply_gift_filter()
        self.tree_gift.see(iid)
        # Auto-save to the database right after every check/uncheck — no
        # separate button needed, so progress isn't lost if you forget to
        # save manually before closing the app. Saves silently (no
        # messagebox) so it doesn't interrupt every single click.
        self._save_gift_roster_to_db(silent=True)

    def _toggle_all_gift_column(self, col_key):
        """MỚI: thay thế 2 nút "☑ Check all"/"☐ Uncheck all" cũ — bấm vào
        TIÊU ĐỀ cột "send_email" hoặc "check" (xem tree.heading(...,
        command=...) trong _build_tab_gift()) để chọn/bỏ chọn TẤT CẢ các
        dòng ĐANG HIỂN THỊ (tôn trọng ô tìm kiếm, giống hệt hành vi 2 nút
        cũ) cho ĐÚNG cột đó — 2 cột hoạt động HOÀN TOÀN ĐỘC LẬP với nhau.
        Trạng thái MỚI luôn là ĐẢO NGƯỢC của trạng thái hiện tại: nếu tất
        cả dòng đang hiển thị đã tick hết -> bỏ tick hết; ngược lại (kể cả
        khi chỉ tick 1 phần) -> tick hết."""
        if col_key not in ("send_email", "check"):
            return
        shown = self.tree_gift.get_children()
        if not shown:
            return
        currently_all = all(self._gift_roster.get(iid, {}).get(
            "checked" if col_key == "check" else "send_email") for iid in shown)
        new_state = not currently_all
        for iid in shown:
            if iid not in self._gift_roster:
                continue
            if col_key == "check":
                self._gift_roster[iid]["checked"] = new_state
                self._gift_roster[iid]["amount"] = (
                    parse_amount_from_text(self.var_gift_budget.get()) if new_state else 0.0
                )
            else:
                self._gift_roster[iid]["send_email"] = new_state
        self._apply_gift_filter()
        self._save_gift_roster_to_db(silent=True)

    def _update_gift_contributed_count(self):
        total = len(self._gift_roster) if hasattr(self, "_gift_roster") else 0
        contributed = sum(1 for info in self._gift_roster.values() if info["checked"]) if hasattr(self, "_gift_roster") else 0
        total_amount = sum(info.get("amount", 0.0) for info in self._gift_roster.values()) if hasattr(self, "_gift_roster") else 0.0
        shown = len(self.tree_gift.get_children())
        if shown != total:
            self.var_gift_contributed_count.set(f"{contributed} / {total}  (showing {shown}/{total} due to search)")
        else:
            self.var_gift_contributed_count.set(f"{contributed} / {total}")
        if hasattr(self, "var_gift_total_amount"):
            self.var_gift_total_amount.set(f"{total_amount:,.0f}")
        self._update_gift_reminder_button_label()
        self._update_gift_report_button_label()

    def _update_gift_report_button_label(self):
        """Cập nhật số người trên nút "📧 Send report to N selected people"
        — đếm theo cột "Send email" (self._gift_roster[...]["send_email"]),
        KHÔNG phải cột "Contributed" — 2 khái niệm độc lập, xem
        _build_tab_gift()."""
        if not hasattr(self, "lbl_gift_report_send"):
            return
        count = sum(1 for info in getattr(self, "_gift_roster", {}).values() if info.get("send_email"))
        self.lbl_gift_report_send.config(text=f"📧 Send report to {count} selected people")

    def _save_gift_roster_to_db(self, silent=True):
        """Lưu self._gift_roster vào database (bảng gift_contributions) —
        gọi TỰ ĐỘNG sau mọi tick/bỏ tick, "Check all"/"Uncheck all", hoặc
        import từ file. Đây là hàm PERSIST THẬT SỰ trong kiến trúc mới
        (thay cho việc ghi Excel liên tục trước đây)."""
        event_id = self.var_event_id.get().strip()
        if not event_id or not getattr(self, "_gift_roster", None):
            return
        try:
            db.save_gift_roster(event_id, self._gift_roster, self.history_path.get())
        except Exception:
            if not silent:
                raise
            pass  # best-effort silent auto-save

    def _load_gift_roster_from_db(self, event_id):
        """Đọc gift roster đã lưu trong database cho event_id, dùng bởi
        '⬅ Load setup from selected event' ở Tab 1. Trả về True nếu có dữ
        liệu để nạp."""
        try:
            roster = db.load_gift_roster(event_id, self.history_path.get())
        except Exception:
            return False
        if not roster:
            return False
        self._gift_roster = roster
        self._apply_gift_filter()
        return True

    def _export_gift_contribution_list(self, silent=False):
        """Exports the FULL self._gift_roster (not just rows currently shown
        due to search) to Excel — CHỈ khi bấm nút "📊 Export to Excel"
        (không còn tự động ghi liên tục — dữ liệu thật sự sống trong
        database, xem _save_gift_roster_to_db()). Writes EVERYONE, including
        people who haven't contributed yet (the 'Contributed' column =
        Yes/No), so it's easy to compare total invited vs. total
        contributed."""
        event_id = self.var_event_id.get().strip()
        if not event_id:
            if not silent:
                messagebox.showwarning("Missing Event ID", "Enter an Event ID on Tab 1 before exporting the file.")
            return
        if not getattr(self, "_gift_roster", None):
            if not silent:
                messagebox.showwarning("List is empty",
                                        "No one in the list yet — click '🔄 Reload list from Tab 2' first.")
            return

        out_path = filedialog.asksaveasfilename(
            title="Export Gift Contribution List to Excel",
            defaultextension=".xlsx",
            initialfile=f"Gift_Contribution_List_{event_id}.xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not out_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Gift Contribution"
            headers = ["No.", "Name", "Email", "Amount", "Contributed"]
            for i, h in enumerate(headers, start=1):
                c = ws.cell(row=1, column=i, value=h)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="003366")
                c.alignment = Alignment(horizontal="center")
            row_idx = 2
            contributed_count = 0
            total_amount = 0.0
            for i, (email, info) in enumerate(self._gift_roster.items(), start=1):
                amount = info.get("amount", 0.0)
                if info["checked"]:
                    contributed_count += 1
                    total_amount += amount
                ws.cell(row=row_idx, column=1, value=i)
                ws.cell(row=row_idx, column=2, value=info["name"])
                ws.cell(row=row_idx, column=3, value=email)
                ws.cell(row=row_idx, column=4, value=amount or None)
                ws.cell(row=row_idx, column=5, value="Yes" if info["checked"] else "No")
                row_idx += 1

            # Grand total row, right below the table (one blank row for
            # readability) — so the exported file always carries the total
            # amount collected alongside the per-person breakdown. The label
            # goes in the NAME column (not Email) so read_gift_contribution_rows()
            # correctly skips this row on reload (it keys off the Email
            # column being non-empty to recognize a real person's row).
            total_row = row_idx + 1
            label_cell = ws.cell(row=total_row, column=2, value="TOTAL COLLECTED:")
            label_cell.font = Font(bold=True)
            label_cell.alignment = Alignment(horizontal="right")
            total_cell = ws.cell(row=total_row, column=4, value=total_amount)
            total_cell.font = Font(bold=True)
            total_cell.number_format = "#,##0"

            ws.column_dimensions["A"].width = 6
            ws.column_dimensions["B"].width = 28
            ws.column_dimensions["C"].width = 32
            ws.column_dimensions["D"].width = 14
            ws.column_dimensions["E"].width = 14
            wb.save(out_path)
        except Exception as e:
            if not silent:
                messagebox.showerror("Error", f"Couldn't export the file:\n{e}")
            return

        if not silent:
            total = len(self._gift_roster)
            messagebox.showinfo(
                "File exported",
                f"Saved the contribution list ({contributed_count}/{total} people contributed, "
                f"total collected: {total_amount:,.0f}) to:\n{out_path}"
            )

    def _build_tab_calendar(self):
        f = self.tab_calendar.body
        pad = {"padx": 10, "pady": 6}

        # MỚI: giờ lấy CẢ "Yes" LẪN "Maybe" (trước đây chỉ lấy "Yes") — vì
        # người trả lời "Maybe" (chưa chắc chắn) vẫn nên được mời Calendar
        # Invite chính thức, để họ có lịch trên Outlook phòng khi sau đó họ
        # quyết định tham dự — không nên loại họ ra chỉ vì chưa chốt.
        ttk.Label(f, text='People who voted "Yes" or "Maybe" (auto-pulled from Tab 4):',
                  font=("Arial", 10, "bold")).grid(row=0, column=0, columnspan=2, sticky="w", **pad)
        self.list_yes = tk.Listbox(f, width=60, height=10)
        yes_vscroll = ttk.Scrollbar(f, orient="vertical", command=self.list_yes.yview)
        self.list_yes.configure(yscrollcommand=yes_vscroll.set)
        self.list_yes.grid(row=1, column=0, sticky="w", padx=10)
        yes_vscroll.grid(row=1, column=1, sticky="nsw")
        ttk.Button(f, text="🔄 Refresh Yes/Maybe list", command=self._refresh_calendar_yes_list)\
            .grid(row=2, column=0, sticky="w", **pad)

        # MỚI: đã BỎ dòng "Event date / time (from Tab 1)" ở đây theo yêu
        # cầu — thông tin đó vẫn được dùng NGẦM khi gửi Calendar Invite
        # (xem _send_calendar(), luôn lấy trực tiếp từ Tab 1), chỉ không
        # còn hiển thị lặp lại thành 1 dòng riêng ở Tab 5 nữa. Muốn xem/đổi
        # giờ, quay lại Tab 1.

        # ── customizable appointment body ──
        appt_header = ttk.Frame(f)
        appt_header.grid(row=6, column=0, columnspan=2, sticky="w", **pad)
        ttk.Label(appt_header, text="Appointment body (customize if needed):", font=("Arial", 9, "bold"))\
            .pack(side="left")
        ttk.Label(appt_header, text="   Language:").pack(side="left")
        self.combo_calendar_lang = ttk.Combobox(
            appt_header, width=26, state="readonly",
            values=[LANG_LABELS["en"], LANG_LABELS["ja"], LANG_LABELS["vi"], LANG_LABELS["bilingual"]],
        )
        self.combo_calendar_lang.current(0)  # default: English
        self.combo_calendar_lang.pack(side="left", padx=(4, 0))
        # Đổi ngôn ngữ -> tự thay lại nội dung mặc định ngay (giống combo_reminder_lang
        # ở Tab 4) — nếu bạn đã tự gõ tay nội dung riêng, đổi ngôn ngữ sẽ THAY THẾ
        # bằng bản mặc định của ngôn ngữ mới (không cộng dồn/giữ lại bản cũ).
        self.combo_calendar_lang.bind("<<ComboboxSelected>>", lambda e: self._apply_calendar_body_lang())

        appt_container, self.txt_appt_body = make_scrollable_text(f, width=100, height=4)
        default_body = build_calendar_body("en", *self._calendar_body_args())
        self.txt_appt_body.insert("1.0", default_body)
        self.var_appt_body_default = default_body  # Keep track of default for reset
        appt_container.grid(row=7, column=0, columnspan=2, sticky="w", padx=10, pady=4)

        ttk.Button(f, text="📅 Send Calendar Invite to the list above", command=self._send_calendar)\
            .grid(row=8, column=0, sticky="w", **pad)

        ttk.Separator(f).grid(row=9, column=0, columnspan=3, sticky="ew", padx=10, pady=10)

        # ── Attendance & Payment tracking (post-event) ──
        # Lists everyone who voted "Yes" or "Maybe" (same population as the
        # Calendar Invite list above), so you can mark who ACTUALLY showed
        # up, who's exempt from paying, and how much they paid, after the
        # event happens. Double-click "Name"/"Vote"/"Amount" to type a new
        # value; "Actual Attend" opens a Yes/No dropdown instead of free
        # text; single-click the "Free" column to toggle it. Every edit is
        # auto-saved straight to Attendance_Payment_{EventID}.xlsx in the
        # Input folder — no manual export step needed (see
        # _save_attendance_sheet_to_file()).
        ttk.Label(f, text="Attendance & Payment tracking (fill in after the event):",
                  font=("Arial", 10, "bold")).grid(row=10, column=0, columnspan=2, sticky="w", **pad)
        self._make_wrapping_label(
            f, text="Double-click \"Name\"/\"Vote\"/\"Amount\" to edit; \"Actual Attend\" opens a Yes/No "
                    "dropdown; click a row's \"Free\" cell to toggle exemption. Setting \"Actual Attend\" to "
                    "\"Yes\" auto-fills Amount from Tab 1's \"Expected event budget\" (0 if marked Free) — "
                    "you can still overwrite Amount by hand afterward. Select rows and use Ctrl+C / Ctrl+V "
                    "to copy/paste to or from Excel, or Delete to clear a row's tracking fields. Every "
                    "change is saved automatically to the Attendance_Payment_{EventID}.xlsx file.",
            font=("Arial", 8, "italic")).grid(row=11, column=0, columnspan=3, sticky="w", padx=10)

        attend_btns = ttk.Frame(f)
        attend_btns.grid(row=12, column=0, columnspan=3, sticky="w", **pad)
        ttk.Button(attend_btns, text="🔄 Refresh list from Tab 4", command=self._refresh_attendance_list)\
            .pack(side="left", padx=(0, 6))
        # MỚI: dữ liệu giờ tự động lưu vào database sau MỌI thay đổi (xem
        # _commit_attendance_edit(), _on_attendance_free_click(),
        # _on_attendance_delete_key()) — nút này giờ chỉ để xuất ra 1 file
        # Excel khi cần báo cáo/gửi người khác, không phải nơi lưu trữ chính.
        ttk.Button(attend_btns, text="📊 Export to Excel", command=self._export_attendance_to_excel)\
            .pack(side="left", padx=6)

        attend_cols = ("no", "name", "vote", "actual_attend", "free", "amount")
        attend_container, self.tree_attendance = make_scrollable_treeview(f, columns=attend_cols, height=14)
        attend_headers = ["No.", "Name", "Vote", "Actual Attend", "Free", "Amount"]
        attend_widths = [40, 240, 70, 110, 60, 100]
        for c, h, w in zip(attend_cols, attend_headers, attend_widths):
            self.tree_attendance.heading(c, text=h)
            anchor = "center" if c in ("no", "vote", "actual_attend", "free") else ("e" if c == "amount" else "w")
            self.tree_attendance.column(c, width=w, anchor=anchor)
        attend_container.grid(row=13, column=0, columnspan=3, sticky="w", padx=10, pady=6)
        # "Actual Attend" opens a readonly Yes/No dropdown (see
        # _begin_cell_edit_combobox()); Name/Vote/Amount get a normal
        # type-in-place Entry; "No." and "Free" are not double-click
        # editable ("Free" toggles on single-click instead — see
        # _on_attendance_free_click(), bound separately below so the two
        # bindings don't conflict).
        self.tree_attendance.bind("<Double-1>", self._on_attendance_tree_double_click)
        self.tree_attendance.bind("<Button-1>", self._on_attendance_free_click)
        self._enable_treeview_copy_paste(
            self.tree_attendance, on_commit=self._commit_attendance_edit,
            editable_cols={"name", "vote", "actual_attend", "free", "amount"})
        # Select rows and press Delete/Backspace to clear that row's
        # tracking fields (Actual Attend/Free/Amount) back to blank — Name
        # and Vote are left untouched since they're identity data pulled
        # from Tab 4, not something you'd normally want to blank out by
        # accident. See _on_attendance_delete_key().
        self.tree_attendance.bind("<Delete>", self._on_attendance_delete_key)
        self.tree_attendance.bind("<BackSpace>", self._on_attendance_delete_key)

        totals_frame = ttk.Frame(f)
        totals_frame.grid(row=14, column=0, columnspan=3, sticky="w", **pad)
        ttk.Label(totals_frame, text="Total actual attend:").pack(side="left")
        self.var_total_actual_attend = tk.StringVar(value="0")
        ttk.Label(totals_frame, textvariable=self.var_total_actual_attend, font=("Arial", 10, "bold"))\
            .pack(side="left", padx=(4, 24))
        ttk.Label(totals_frame, text="Total collected amount:").pack(side="left")
        self.var_total_collected_amount = tk.StringVar(value="0")
        ttk.Label(totals_frame, textvariable=self.var_total_collected_amount, font=("Arial", 10, "bold"))\
            .pack(side="left", padx=(4, 0))

        # MỚI: "Amount paid" — số tiền THỰC TẾ đã trả (vd đặt cọc/thanh
        # toán trước cho nhà hàng, chuyển khoản cho Guest of Honor...),
        # nhập tay vì con số này KHÔNG thể tự suy ra được từ bảng
        # Attendance. "Remaining amount" tự tính = Total collected amount −
        # Amount paid, cập nhật ngay khi gõ (self.var_amount_paid.trace_add
        # bên dưới) hoặc khi bảng Attendance đổi (_update_attendance_totals
        # gọi lại cùng công thức). Giá trị Amount paid được lưu theo từng
        # Event ID trong database (cột "AmountPaid" ở bảng events, xem
        # db.py) — nạp lại tự động bởi "Load setup from selected event".
        paid_frame = ttk.Frame(f)
        paid_frame.grid(row=15, column=0, columnspan=3, sticky="w", **pad)
        ttk.Label(paid_frame, text="Amount paid:").pack(side="left")
        self.var_amount_paid = tk.StringVar(value="0")
        ttk.Entry(paid_frame, textvariable=self.var_amount_paid, width=14)\
            .pack(side="left", padx=(4, 24))
        ttk.Label(paid_frame, text="Remaining amount:").pack(side="left")
        self.var_remaining_amount = tk.StringVar(value="0")
        ttk.Label(paid_frame, textvariable=self.var_remaining_amount, font=("Arial", 10, "bold"))\
            .pack(side="left", padx=(4, 0))
        # Ghi chú nhỏ để không ai nhầm đây là số tự tính từ bảng.
        ttk.Label(paid_frame, text="  (type the amount actually paid out; "
                                    "remaining = Total collected − Amount paid)",
                  font=("Arial", 8, "italic")).pack(side="left")
        self.var_amount_paid.trace_add("write", lambda *a: self._on_amount_paid_changed())

        # ── Thank You email (post-event, MỚI) ──
        # Soạn sẵn nội dung email cảm ơn mọi người đã tham gia, cùng cách
        # làm với Appointment body ở trên: đa ngôn ngữ (English/Japanese/
        # Vietnamese/Bilingual), có thể sửa tay trước khi gửi, tự refresh
        # theo Tab 1 khi CHƯA hand-edit (xem _refresh_thankyou_body_display()
        # / _apply_thankyou_body_lang(), cùng cơ chế với
        # _refresh_calendar_datetime_display()). Người nhận = những ai
        # "Actual Attend" = Yes trong bảng ở trên (xem _send_thank_you_email()).
        ttk.Separator(f).grid(row=16, column=0, columnspan=3, sticky="ew", padx=10, pady=10)

        ttk.Label(f, text="Thank You email (send after the event):",
                  font=("Arial", 10, "bold")).grid(row=17, column=0, columnspan=2, sticky="w", **pad)
        self._make_wrapping_label(
            f, text="Sent to everyone marked \"Actual Attend\" = Yes above. Automatically attaches "
                    "the Attendance & Payment Excel report (with everyone's attendance/cost info), "
                    "and the Calendar Invite for this event if it can be found in your Calendar "
                    "(best-effort — see the confirmation message after sending).",
            font=("Arial", 8, "italic")).grid(row=18, column=0, columnspan=3, sticky="w", padx=10)

        thankyou_header = ttk.Frame(f)
        thankyou_header.grid(row=19, column=0, columnspan=2, sticky="w", **pad)
        ttk.Label(thankyou_header, text="   Language:").pack(side="left")
        self.combo_thankyou_lang = ttk.Combobox(
            thankyou_header, width=26, state="readonly",
            values=[LANG_LABELS["en"], LANG_LABELS["ja"], LANG_LABELS["vi"], LANG_LABELS["bilingual"]],
        )
        self.combo_thankyou_lang.current(0)  # default: English
        self.combo_thankyou_lang.pack(side="left", padx=(4, 0))
        self.combo_thankyou_lang.bind("<<ComboboxSelected>>", lambda e: self._apply_thankyou_body_lang())

        thankyou_container, self.txt_thankyou_body = make_scrollable_text(f, width=100, height=6)
        default_thankyou_body = build_thankyou_body("en", *self._thankyou_body_args())
        self.txt_thankyou_body.insert("1.0", default_thankyou_body)
        self.var_thankyou_body_default = default_thankyou_body  # Keep track of default for reset
        thankyou_container.grid(row=20, column=0, columnspan=2, sticky="w", padx=10, pady=4)

        ttk.Button(f, text="📧 Send Thank You email to confirmed attendees",
                   command=self._send_thank_you_email).grid(row=21, column=0, sticky="w", **pad)

    def _refresh_attendance_list(self):
        """Rebuilds self._attendance_roster (the source of truth for the
        Attendance & Payment table) from the Yes/Maybe voters scanned on
        Tab 4 (self.tree_responses) — same population as the Calendar
        Invite list above. Existing "Actual Attend"/"Free"/"Amount" edits
        for people already in the roster are KEPT (so re-scanning Tab 4 or
        switching back to this tab doesn't wipe out attendance already
        marked); brand-new voters default to Actual Attend = "Yes" (with
        Amount auto-filled to match), per the requirement that Yes is the
        default rather than blank."""
        if not hasattr(self, "tree_responses"):
            return
        if not hasattr(self, "_attendance_roster"):
            self._attendance_roster = {}  # email -> {"name","vote","actual_attend","free","amount"}
        new_roster = {}
        newly_added = []
        for iid in self.tree_responses.get_children():
            _manual, name, email, vote, _received = self.tree_responses.item(iid, "values")
            if vote not in ("Yes", "Maybe"):
                continue
            is_new = email not in self._attendance_roster
            prev = self._attendance_roster.get(email, {})
            new_roster[email] = {
                "name": name,
                "vote": vote,
                "actual_attend": prev.get("actual_attend", "Yes"),
                "free": prev.get("free", False),
                "amount": prev.get("amount", 0.0),
            }
            if is_new:
                newly_added.append(email)
        self._attendance_roster = new_roster
        # Only auto-sync Amount for brand-new people (matching their
        # default "Yes") — anyone already in the roster keeps whatever
        # Amount they already had, even if it's 0 by manual choice.
        for email in newly_added:
            self._sync_attendance_amount(email)
        self._render_attendance_tree()

    def _render_attendance_tree(self):
        if not hasattr(self, "tree_attendance") or not hasattr(self, "_attendance_roster"):
            return
        self.tree_attendance.delete(*self.tree_attendance.get_children())
        for i, (email, info) in enumerate(self._attendance_roster.items(), start=1):
            amount = info.get("amount", 0.0)
            amount_display = f"{amount:,.0f}" if amount else ""
            free_display = "✅" if info.get("free") else "⬜"
            self.tree_attendance.insert(
                "", "end", iid=email,
                values=(i, info["name"], info["vote"], info.get("actual_attend", ""),
                        free_display, amount_display))
        self._update_attendance_totals()

    def _sync_attendance_amount(self, email):
        """Recomputes one person's Amount from their current Actual
        Attend / Free state: Free always wins (Amount forced to 0,
        regardless of attendance); otherwise Amount is auto-pulled from
        Tab 1's "Expected event budget" if Actual Attend is "Yes", or 0
        otherwise. Called after editing/pasting/toggling either field —
        any Amount typed in by hand afterward stays until Actual Attend or
        Free changes again."""
        info = self._attendance_roster[email]
        if info.get("free"):
            info["amount"] = 0.0
        elif (info.get("actual_attend") or "").strip().lower() == "yes":
            info["amount"] = parse_amount_from_text(self.var_budget.get())
        else:
            info["amount"] = 0.0

    def _commit_attendance_edit(self, row_id, col_name, new_value):
        email = row_id
        if email not in self._attendance_roster:
            return
        info = self._attendance_roster[email]
        new_value = new_value.strip()
        if col_name == "name":
            info["name"] = new_value
        elif col_name == "vote":
            info["vote"] = new_value
        elif col_name == "actual_attend":
            info["actual_attend"] = new_value
            self._sync_attendance_amount(email)
        elif col_name == "free":
            # Reached via paste (Ctrl+V), not the single-click toggle —
            # accepts a few common truthy spellings so pasting a column
            # copied from Excel (Yes/TRUE/1/✅) works as expected.
            info["free"] = new_value.lower() in ("yes", "true", "1", "✅", "x")
            self._sync_attendance_amount(email)
        elif col_name == "amount":
            info["amount"] = parse_amount_from_text(new_value)
        self._render_attendance_tree()
        # MỚI: auto-save ngay xuống Attendance_Payment_{EventID}.xlsx sau
        # MỌI thay đổi (double-click edit lẫn paste), không cần bấm nút
        # riêng — best-effort, lỗi (nếu có, vd file đang mở ở chỗ khác) bị
        # bỏ qua âm thầm để không làm gián đoạn thao tác chỉnh sửa bình
        # thường (xem docstring _save_attendance_sheet_to_file()).
        self._save_attendance_sheet_to_file(silent=True)

    def _on_attendance_tree_double_click(self, event):
        """Double-click dispatcher for the Attendance & Payment table:
        "Actual Attend" opens a readonly Yes/No dropdown (see
        _begin_cell_edit_combobox()); "Name"/"Vote"/"Amount" open a normal
        type-in-place Entry (see _begin_cell_edit()); "No." and "Free" are
        not editable via double-click ("Free" toggles on single-click
        instead — see _on_attendance_free_click())."""
        tree = self.tree_attendance
        region = tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        col = tree.identify_column(event.x)
        row_id = tree.identify_row(event.y)
        if not row_id or not col:
            return
        columns = tree["columns"]
        try:
            col_index = int(col.replace("#", "")) - 1
        except ValueError:
            return
        if col_index < 0 or col_index >= len(columns):
            return
        col_name = columns[col_index]
        if col_name == "actual_attend":
            self._begin_cell_edit_combobox(tree, row_id, col_name, ["Yes", "No"], self._commit_attendance_edit)
        elif col_name in ("name", "vote", "amount"):
            self._begin_cell_edit(tree, row_id, col_name, self._commit_attendance_edit)
        # "no" and "free" columns: not double-click editable, ignored here

    def _on_attendance_free_click(self, event):
        """Single-click toggle for the "Free" column (simulated checkbox,
        same ✅/⬜ pattern as the Gift Contribution tab's "Contributed"
        column) — marking someone Free forces their Amount to 0 regardless
        of Actual Attend, for people exempt from the contribution."""
        region = self.tree_attendance.identify("region", event.x, event.y)
        if region != "cell":
            return
        col = self.tree_attendance.identify_column(event.x)
        row_id = self.tree_attendance.identify_row(event.y)
        if not row_id or col != "#5":  # "#5" = the "free" column (no,name,vote,actual_attend,free,amount)
            return
        email = row_id
        if email not in self._attendance_roster:
            return
        info = self._attendance_roster[email]
        info["free"] = not info.get("free", False)
        self._sync_attendance_amount(email)
        self._render_attendance_tree()
        self._save_attendance_sheet_to_file(silent=True)

    def _on_attendance_delete_key(self, event):
        """Delete/Backspace on selected row(s) — clears that row's
        tracking fields (Actual Attend → blank, Free → off, Amount → 0)
        back to an untouched state. Name and Vote are left as-is (they're
        identity data pulled from Tab 4, not meant to be blanked by a
        stray Delete press); to change those, double-click and type a new
        value, or leave them blank by hand via that same edit box.
        Treeview only supports selecting whole ROWS (not individual
        cells), so this clears the row's editable tracking fields as a
        group rather than a single cell — the closest practical match to
        "delete a cell's content" the widget allows."""
        changed = False
        for row_id in self.tree_attendance.selection():
            if row_id not in self._attendance_roster:
                continue
            info = self._attendance_roster[row_id]
            info["actual_attend"] = ""
            info["free"] = False
            info["amount"] = 0.0
            changed = True
        if changed:
            self._render_attendance_tree()
            self._save_attendance_sheet_to_file(silent=True)
        return "break"

    def _update_attendance_totals(self):
        total_attend = sum(
            1 for info in self._attendance_roster.values()
            if (info.get("actual_attend") or "").strip().lower() == "yes"
        )
        total_amount = sum(info.get("amount", 0.0) for info in self._attendance_roster.values())
        self.var_total_actual_attend.set(str(total_attend))
        self.var_total_collected_amount.set(f"{total_amount:,.0f}")
        self._refresh_remaining_amount()

    def _refresh_remaining_amount(self):
        """Remaining amount = Total collected amount − Amount paid. Called
        both when the Attendance table changes (Total collected amount
        moves) and when the user types into "Amount paid" directly (see
        the trace_add on self.var_amount_paid) — either side changing
        should immediately update this."""
        if not hasattr(self, "var_amount_paid") or not hasattr(self, "var_remaining_amount"):
            return
        total_collected = parse_amount_from_text(self.var_total_collected_amount.get())
        amount_paid = parse_amount_from_text(self.var_amount_paid.get())
        self.var_remaining_amount.set(f"{(total_collected - amount_paid):,.0f}")

    def _on_amount_paid_changed(self):
        """Recomputes Remaining amount and auto-saves "Amount paid" to the
        database (events table, "AmountPaid" column) as the user types —
        same auto-save-on-every-edit pattern as the rest of the Attendance
        & Payment tab, best-effort/silent so a stray keystroke before an
        Event ID exists doesn't pop up an error."""
        self._refresh_remaining_amount()
        event_id = self.var_event_id.get().strip()
        if not event_id:
            return
        try:
            db.save_event_record(
                {"EventID": event_id, "AmountPaid": self.var_amount_paid.get()},
                self.history_path.get())
        except Exception:
            pass  # best-effort silent auto-save, same as _save_attendance_sheet_to_file()

    # ── Attendance & Payment / Responded result — lưu vào database ──
    # MỚI: đã đổi từ file Excel Attendance_Payment_{EventID}.xlsx (2 sheet)
    # sang lưu THẲNG vào database (bảng attendance + responses trong
    # db.py) — auto-save mỗi khi có thay đổi, không tự ghi Excel liên tục
    # nữa. Muốn có file Excel để báo cáo/gửi người khác, dùng nút
    # "📊 Export to Excel" (xem _export_attendance_to_excel()). Cả 2 đều
    # được đọc lại tự động bởi '⬅ Load setup from selected event' trên
    # Tab 1, không cần quét lại Outlook để xem lại dữ liệu cũ.

    def _save_attendance_sheet_to_file(self, silent=True):
        """Lưu bảng Attendance & Payment hiện tại (self._attendance_roster)
        vào database — gọi TỰ ĐỘNG sau mọi thay đổi (double-click, toggle
        Free, paste, Delete). Tên hàm giữ nguyên như cũ (dù giờ không còn
        ghi "file" Excel nữa) để không phải sửa lại các nơi đang gọi nó."""
        event_id = self.var_event_id.get().strip()
        if not event_id or not getattr(self, "_attendance_roster", None):
            return
        try:
            db.save_attendance_roster(event_id, self._attendance_roster, self.history_path.get())
        except Exception:
            if not silent:
                raise
            pass  # best-effort silent auto-save

    def _load_attendance_sheet_from_file(self, event_id):
        """Đọc bảng Attendance & Payment đã lưu trong database cho
        event_id, nạp vào self._attendance_roster và vẽ lại — dùng bởi
        '⬅ Load setup from selected event'. Trả về True nếu có dữ liệu để
        nạp, False nếu chưa từng lưu gì cho sự kiện này (không phải lỗi)."""
        try:
            roster = db.load_attendance_roster(event_id, self.history_path.get())
        except Exception:
            return False
        if not roster:
            return False
        self._attendance_roster = roster
        self._render_attendance_tree()
        return True

    def _save_responded_result_to_file(self, silent=True):
        """Lưu bảng "Responded / not yet responded" hiện tại của Tab 4
        (self.responses) vào database — gọi tự động ngay sau mỗi lần Scan
        Inbox thành công, và lại lần nữa khi bấm '🗂 Save event', để
        '⬅ Load setup from selected event' khôi phục được kết quả quét gần
        nhất mà KHÔNG cần quét lại Outlook (chỉ bấm '📨 Scan Inbox for Vote
        results' mới thực sự quét lại — xem _collect_responses())."""
        event_id = self.var_event_id.get().strip()
        if not event_id:
            return
        try:
            db.save_responses(event_id, self.responses, self._last_scan_time, self.history_path.get())
        except Exception:
            if not silent:
                raise
            pass  # best-effort silent auto-save

    def _load_responded_result_from_file(self, event_id):
        """Đọc kết quả Responded đã lưu trong database cho event_id, nạp
        vào self.responses và vẽ lại bảng Tab 4 — dùng bởi '⬅ Load setup
        from selected event'. Trả về True nếu có dữ liệu để nạp."""
        if not hasattr(self, "tree_responses"):
            return False
        try:
            responses, last_scan = db.load_responses(event_id, self.history_path.get())
        except Exception:
            return False
        if not responses:
            return False
        self.responses = responses
        self._last_scanned_event_id = event_id
        self._last_scan_time = last_scan or datetime.now()
        roster = self._build_effective_roster() if hasattr(self, "_build_effective_roster") else self.recipients
        self._refresh_response_tree(0, roster)
        self._update_scan_status_banner()
        return True

    def _build_attendance_workbook(self):
        """Dựng 1 openpyxl Workbook cho báo cáo Attendance & Payment —
        TÁCH RIÊNG từ _export_attendance_to_excel() (giữ nguyên layout cũ)
        để dùng chung ở 2 nơi: nút "📊 Export to Excel" (lưu ra file người
        dùng chọn) VÀ file đính kèm tự động của email cảm ơn (lưu ra file
        tạm, xem _send_thank_you_email()) — tránh code trùng lặp giữa 2
        chỗ. MỚI: thêm 2 dòng "Amount paid"/"Remaining amount" (tính năng
        #3) vào cuối báo cáo, ngay dưới "Total collected amount"."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance & Payment"
        headers = ["No.", "Name", "Email", "Vote", "Actual Attend", "Free", "Amount"]
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="003366")
            c.alignment = Alignment(horizontal="center")
        row_idx = 2
        total_attend = 0
        total_amount = 0.0
        for i, (email, info) in enumerate(self._attendance_roster.items(), start=1):
            attend = (info.get("actual_attend") or "").strip()
            amount = info.get("amount", 0.0)
            if attend.lower() == "yes":
                total_attend += 1
            total_amount += amount
            ws.cell(row=row_idx, column=1, value=i)
            ws.cell(row=row_idx, column=2, value=info["name"])
            ws.cell(row=row_idx, column=3, value=email)
            ws.cell(row=row_idx, column=4, value=info.get("vote", ""))
            ws.cell(row=row_idx, column=5, value=attend)
            ws.cell(row=row_idx, column=6, value="Yes" if info.get("free") else "No")
            ws.cell(row=row_idx, column=7, value=amount or None)
            row_idx += 1
        total_row = row_idx + 1
        ws.cell(row=total_row, column=2, value="Total actual attend:").font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=total_attend).font = Font(bold=True)
        amount_row = total_row + 1
        ws.cell(row=amount_row, column=2, value="Total collected amount:").font = Font(bold=True)
        c_amt = ws.cell(row=amount_row, column=7, value=total_amount)
        c_amt.font = Font(bold=True)
        c_amt.number_format = "#,##0"
        # MỚI: "Amount paid"/"Remaining amount" — đọc trực tiếp từ 2
        # StringVar trên UI (đã tính sẵn bởi _refresh_remaining_amount()),
        # thay vì tính lại ở đây, để LUÔN khớp đúng những gì đang hiển thị
        # trên Tab 5 lúc xuất báo cáo.
        amount_paid = parse_amount_from_text(getattr(self, "var_amount_paid", tk.StringVar(value="0")).get())
        remaining = parse_amount_from_text(getattr(self, "var_remaining_amount", tk.StringVar(value="0")).get())
        paid_row = amount_row + 1
        ws.cell(row=paid_row, column=2, value="Amount paid:").font = Font(bold=True)
        c_paid = ws.cell(row=paid_row, column=7, value=amount_paid)
        c_paid.font = Font(bold=True)
        c_paid.number_format = "#,##0"
        remaining_row = paid_row + 1
        ws.cell(row=remaining_row, column=2, value="Remaining amount:").font = Font(bold=True)
        c_rem = ws.cell(row=remaining_row, column=7, value=remaining)
        c_rem.font = Font(bold=True)
        c_rem.number_format = "#,##0"
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 26
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 14
        return wb

    def _export_attendance_to_excel(self):
        """Xuất bảng Attendance & Payment ra 1 file Excel — CHỈ khi bấm nút
        "📊 Export to Excel" (thay cho nút "Open folder"/"Export Excel
        report" cũ), vì giờ dữ liệu thật sự sống trong database, không tự
        ghi Excel liên tục nữa."""
        event_id = self.var_event_id.get().strip()
        if not event_id:
            messagebox.showwarning("Missing Event ID", "Enter an Event ID on Tab 1 first.")
            return
        if not getattr(self, "_attendance_roster", None):
            messagebox.showwarning("List is empty",
                                    "No one in the list yet — click '🔄 Refresh list from Tab 4' first.")
            return
        out_path = filedialog.asksaveasfilename(
            title="Export Attendance & Payment to Excel",
            defaultextension=".xlsx",
            initialfile=f"Attendance_Payment_{event_id}.xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not out_path:
            return
        try:
            wb = self._build_attendance_workbook()
            wb.save(out_path)
        except Exception as e:
            messagebox.showerror("Error", f"Couldn't export the report:\n{e}")
            return
        messagebox.showinfo("Exported", f"Saved the attendance & payment report to:\n{out_path}")

    def _calendar_body_args(self):
        """Lấy đúng (event_name, event_date, location, budget) hiện tại từ
        Tab 1 — dùng để build nội dung Appointment body mặc định (Tab 5),
        tránh lặp lại 4 dòng self.var_xxx.get() ở nhiều nơi gọi
        build_calendar_body()."""
        return (
            self.var_event_name.get(),
            get_date_str(self.date_event),
            self.var_location.get(),
            self.var_budget.get(),
        )

    def _refresh_calendar_datetime_display(self):
        """Tự làm mới nội dung Appointment body mặc định theo Event Name/
        Date/Location/Budget mới nhất từ Tab 1 — NHƯNG CHỈ khi người dùng
        CHƯA tự tay sửa nó (tức nội dung đang hiển thị vẫn khớp y hệt bản
        mặc định đã sinh lần gần nhất — self.var_appt_body_default). Nếu
        khác (đã hand-edit), GIỮ NGUYÊN, không ghi đè mất công sửa tay.
        (Trước đây hàm này còn cập nhật 1 dòng chữ hiển thị Event date/
        time riêng ở Tab 5 — dòng đó đã được BỎ theo yêu cầu, xem
        _build_tab_calendar(); phần refresh Appointment body vẫn giữ lại
        vì vẫn cần thiết.)"""
        if hasattr(self, "txt_appt_body") and hasattr(self, "combo_calendar_lang"):
            current_text = self.txt_appt_body.get("1.0", "end").strip()
            if current_text == (getattr(self, "var_appt_body_default", "") or "").strip():
                self._apply_calendar_body_lang()

    def _apply_calendar_body_lang(self):
        lang_label = self.combo_calendar_lang.get() or LANG_LABELS["en"]
        lang_code = LANG_LABEL_TO_CODE.get(lang_label, "en")
        new_body = build_calendar_body(lang_code, *self._calendar_body_args())
        self.var_appt_body_default = new_body
        self.txt_appt_body.delete("1.0", "end")
        self.txt_appt_body.insert("1.0", new_body)

    def _thankyou_body_args(self):
        """Lấy đúng (event_name, event_date, location, total_attend,
        total_collected, amount_paid, remaining_amount) hiện tại từ Tab 1
        + bảng Attendance & Payment (Tab 5) — dùng để build nội dung email
        cảm ơn mặc định (build_thankyou_body()). Các số liệu Attendance
        luôn đọc TRỰC TIẾP từ 3 StringVar hiển thị trên UI (đã được
        _update_attendance_totals()/_refresh_remaining_amount() giữ luôn
        cập nhật), tránh phải tính lại từ self._attendance_roster ở đây."""
        return (
            self.var_event_name.get(),
            get_date_str(self.date_event),
            self.var_location.get(),
            getattr(self, "var_total_actual_attend", tk.StringVar(value="0")).get(),
            getattr(self, "var_total_collected_amount", tk.StringVar(value="0")).get(),
            getattr(self, "var_amount_paid", tk.StringVar(value="0")).get(),
            getattr(self, "var_remaining_amount", tk.StringVar(value="0")).get(),
        )

    def _refresh_thankyou_body_display(self):
        """Tự làm mới nội dung email cảm ơn mặc định theo Tab 1/Attendance
        mới nhất — CÙNG quy tắc "chỉ ghi đè nếu chưa hand-edit" như
        _refresh_calendar_datetime_display() ở trên, để không mất công sửa
        tay khi chuyển qua chuyển lại giữa các tab."""
        if hasattr(self, "txt_thankyou_body") and hasattr(self, "combo_thankyou_lang"):
            current_text = self.txt_thankyou_body.get("1.0", "end").strip()
            if current_text == (getattr(self, "var_thankyou_body_default", "") or "").strip():
                self._apply_thankyou_body_lang()

    def _apply_thankyou_body_lang(self):
        lang_label = self.combo_thankyou_lang.get() or LANG_LABELS["en"]
        lang_code = LANG_LABEL_TO_CODE.get(lang_label, "en")
        new_body = build_thankyou_body(lang_code, *self._thankyou_body_args())
        self.var_thankyou_body_default = new_body
        self.txt_thankyou_body.delete("1.0", "end")
        self.txt_thankyou_body.insert("1.0", new_body)

    def _refresh_calendar_yes_list(self):
        # MỚI: lấy CẢ "Yes" LẪN "Maybe" (trước đây CHỈ lấy "Yes") — biến
        # self._yes_emails vẫn giữ nguyên TÊN cũ (để không phải sửa lại các
        # chỗ khác đang dùng, vd _send_calendar()) nhưng giờ nội dung thực
        # sự là "Yes + Maybe".
        self.list_yes.delete(0, "end")
        self._yes_emails = []
        for iid in self.tree_responses.get_children():
            _manual, name, email, vote, _ = self.tree_responses.item(iid, "values")
            if vote in ("Yes", "Maybe"):
                self.list_yes.insert("end", f"{name} <{email}> [{vote}]")
                self._yes_emails.append(email)

    def _send_calendar(self):
        if not getattr(self, "_yes_emails", None):
            messagebox.showwarning("No Yes/Maybe votes", "No one has voted Yes or Maybe yet, or responses haven't been scanned on Tab 4.")
            return
        try:
            hh, mm = self.var_start_time.get().split(":")
            eh, em = self.var_end_time.get().split(":")
        except Exception:
            messagebox.showerror("Invalid time format", "Enter time as HH:MM, e.g. 18:00")
            return

        base = get_date_obj(self.date_event)
        start_dt = base.replace(hour=int(hh), minute=int(mm))
        end_dt = base.replace(hour=int(eh), minute=int(em))
        subject = self.var_event_name.get()
        location = self.var_location.get()
        # Use custom appointment body from Tab 5
        body = self.txt_appt_body.get("1.0", "end").strip()
        if not body:
            body = self.var_appt_body_default

        event_id = self.var_event_id.get().strip()

        # MỚI: chỉ tìm email UPDATE INVITE để đính kèm nếu RSVP_History.xlsx
        # CÓ ghi nhận đã từng gửi update invite cho đúng Event ID này (cột
        # UpdateInviteDate không trống) — đúng yêu cầu "chỉ dựa vào thông
        # tin trong History để quyết định đính kèm gì", thay vì cứ quét
        # Sent Items xem có khớp Subject hay không.
        include_update = False
        try:
            for rec in db.load_history(self.history_path.get()):
                if rec.get("EventID") == event_id:
                    include_update = bool(rec.get("UpdateInviteDate"))
                    break
        except Exception:
            pass  # best-effort — không đọc được History thì coi như chưa có update invite

        def worker():
            try:
                # send_calendar_invite() giờ CHỈ tìm + đính kèm ĐÚNG 2 loại
                # email dựa theo thông tin History: email mời GỐC (luôn tìm)
                # và email UPDATE INVITE (chỉ tìm nếu include_update=True ở
                # trên) — trong CÙNG 1 phiên COM (tránh bug reference hỏng
                # đã gặp trước đây — xem docstring trong outlook_com.py).
                # KHÔNG còn đính kèm nhầm email nhắc nhở hay email reply
                # Yes/No/RE (xem docstring _find_all_sent_invite_mails()).
                # attached_count = SỐ email đính kèm thành công.
                appt, attached_count = outlook_com.send_calendar_invite(
                    self._yes_emails, subject, location, start_dt, end_dt, body,
                    attach_event_id=event_id or None,
                    include_update=include_update,
                )
                self._calendar_sent_flag = "Sent"
                if attached_count:
                    attach_note = (f"\n\n📎 Found and attached {attached_count} email(s) "
                                    f"(original invite" + (" + update invite" if include_update else "")
                                    + f") related to this event.")
                elif event_id:
                    attach_note = (
                        "\n\n⚠️ Could not find/attach any confirmation email in Sent "
                        "Items (it may have been moved/deleted, sent from a different account, "
                        "or this EventID has no matching email) — the invite was still created "
                        "without an attachment."
                    )
                else:
                    attach_note = ""
                self.after(0, lambda: messagebox.showinfo(
                    "Done",
                    f"Meeting invite opened for {len(self._yes_emails)} people. "
                    "Review it and click Send in Outlook." + attach_note))
            except Exception as e:
                err_msg = str(e)
                self.after(0, lambda: messagebox.showerror("Error", f"Could not send the meeting invite:\n{err_msg}"))

        threading.Thread(target=worker, daemon=True).start()

    def _send_thank_you_email(self):
        """MỚI: gửi email cảm ơn sau sự kiện tới những ai "Actual Attend" =
        Yes trong bảng Attendance & Payment. Tự đính kèm: (1) báo cáo Excel
        Attendance & Payment (đầy đủ tên/vote/actual attend/free/amount +
        Amount paid/Remaining amount, xem _build_attendance_workbook()),
        và (2) Calendar Invite của sự kiện, tìm best-effort trong folder
        Calendar theo đúng Subject = Event Name (xem
        outlook_com.send_thankyou_email() / outlook_com._find_calendar_
        invite_appointment() — CÙNG file/convention với các nút gửi khác
        của app, không còn tách riêng như bản trước khi có outlook_com.py)."""
        actual_attendees = [
            (info.get("name") or email, email)
            for email, info in getattr(self, "_attendance_roster", {}).items()
            if (info.get("actual_attend") or "").strip().lower() == "yes"
        ]
        if not actual_attendees:
            messagebox.showwarning(
                "No confirmed attendees",
                "No one is marked \"Actual Attend\" = Yes in the table above yet. "
                "Fill in Actual Attend for the people who showed up first."
            )
            return

        body = self.txt_thankyou_body.get("1.0", "end").strip()
        if not body:
            messagebox.showwarning(
                "Empty content",
                "Change the language dropdown to regenerate the text, or type the content by hand "
                "before sending.")
            return

        if not getattr(self, "_attendance_roster", None):
            messagebox.showwarning("List is empty",
                                    "No one in the Attendance & Payment table yet — click "
                                    "'🔄 Refresh list from Tab 4' first.")
            return

        event_id = self.var_event_id.get().strip()
        event_name = self.var_event_name.get()
        lang_label = self.combo_thankyou_lang.get() or LANG_LABELS["en"]
        lang_code = LANG_LABEL_TO_CODE.get(lang_label, "en")
        subject = build_thankyou_subject(lang_code, event_id, event_name)

        # Chuẩn bị file Excel đính kèm TRƯỚC (trong main thread, vì cần
        # đọc self._attendance_roster/các StringVar UI — an toàn hơn đọc
        # chúng từ worker thread), lưu vào 1 file tạm — không phải nơi lưu
        # trữ chính (dữ liệu thật vẫn ở database), chỉ để đính kèm email.
        try:
            wb = self._build_attendance_workbook()
            excel_path = os.path.join(
                tempfile.gettempdir(), f"Attendance_Payment_{event_id or 'event'}.xlsx")
            wb.save(excel_path)
        except Exception as e:
            messagebox.showerror("Error", f"Couldn't prepare the attendance Excel attachment:\n{e}")
            return

        attendee_count = len(actual_attendees)

        def worker():
            try:
                mail, calendar_attached = outlook_com.send_thankyou_email(
                    actual_attendees, subject, body,
                    excel_path=excel_path, event_name=event_name,
                )
                if calendar_attached:
                    attach_note = "\n\n📎 Attached the attendance report and the Calendar Invite."
                else:
                    attach_note = (
                        "\n\n📎 Attached the attendance report.\n\n"
                        "⚠️ Couldn't find the Calendar Invite in your Calendar folder to attach "
                        "(it may not have been sent yet, or the meeting Subject no longer matches "
                        "the Event Name on Tab 1) — the email was still created without it."
                    )
                self.after(0, lambda: messagebox.showinfo(
                    "Done",
                    f"Thank-you email opened for {attendee_count} confirmed attendee(s). "
                    "Review it, then click Send in Outlook." + attach_note))
            except Exception as e:
                err_msg = str(e)
                self.after(0, lambda: messagebox.showerror(
                    "Error", f"Could not create the thank-you email:\n{err_msg}"))

        threading.Thread(target=worker, daemon=True).start()

    def _build_tab_history(self):
        f = self.tab_history.body
        top = ttk.Frame(f)
        top.pack(fill="x", padx=10, pady=8)
        ttk.Label(top, text="Database file:").pack(side="left")
        ttk.Entry(top, textvariable=self.history_path, width=40).pack(side="left", padx=6)
        ttk.Button(top, text="📂 Browse...", command=self._browse_history_file).pack(side="left", padx=4)
        ttk.Button(top, text="🔄 Reload", command=self._refresh_history_tree).pack(side="left", padx=4)
        # Double-click any cell below to edit it directly, then click this
        # button to write your changes back to the database (see
        # _save_history_edits()). Nothing is written until you click Save —
        # double-clicking only edits what's shown on screen.
        ttk.Button(top, text="💾 Save changes", command=self._save_history_edits)\
            .pack(side="left", padx=4)
        # MỚI: kiến trúc lưu trữ đã đổi sang SQLite (rsvp_data.db) — bảng
        # này giờ đọc/ghi thẳng vào DB, KHÔNG còn tự động đọc/ghi
        # RSVP_History.xlsx nữa. Muốn có file Excel để báo cáo/gửi người
        # khác thì bấm nút này để xuất TOÀN BỘ bảng hiện tại ra 1 file
        # .xlsx mới — Excel giờ chỉ còn là "ảnh chụp" xuất ra khi cần, không
        # phải nơi lưu trữ chính nữa.
        ttk.Button(top, text="📊 Export to Excel", command=self._export_history_to_excel)\
            .pack(side="left", padx=4)

        self._make_wrapping_label(f, text="💡 Double-click any cell below to edit it in place (or select rows and use "
                          "Ctrl+C / Ctrl+V to copy/paste to or from Excel), then click "
                          "'💾 Save changes' to write your edits back to the database. "
                          "(Renaming the Event ID column is supported — it moves the row under the new ID "
                          "instead of creating a duplicate.) Use '📊 Export to Excel' any time you need a "
                          "report file to share or archive — the database itself is the working copy.",
                  font=("Arial", 8, "italic")).pack(anchor="w", padx=10, pady=(0, 4))

        # BUG ĐÃ SỬA: trước đây bảng này chỉ hiện 1 phần cột (thiếu Deadline,
        # Location, OrganizerNote, RecipientFile, UpdateInviteDate,
        # ReportFile, CalendarSent, ReminderSent — 8/24 cột có trong
        # RSVP_History.xlsx bị bỏ sót). Giờ hiện ĐẦY ĐỦ toàn bộ, đúng thứ tự
        # với db.EVENT_COLUMNS — cuộn ngang (thanh cuộn dưới cùng cửa sổ) để
        # xem hết nếu màn hình không đủ rộng. Đã thêm "LastReminderSentDate"
        # (log lại lần gần nhất bấm "Gửi email nhắc nhở" ở Tab 4) và 7 cột
        # MỚI của tính năng "Event mode": EventMode/Organizer/GuestOfHonor/
        # GiftBudget/StartTime/EndTime/GiftDeadline.
        cols = ("EventID", "EventName", "EventDate", "Deadline", "Location", "Budget",
                 "EmailLanguage", "OrganizerNote", "RecipientFile", "SentDate", "UpdateInviteDate",
                 "TotalInvited", "Yes", "No", "Maybe", "NoResponse",
                 "ReportFile", "CalendarSent",
                 "ActualAttendees", "CostPerPerson", "TotalIncome", "TotalExpense", "Balance",
                 "ReminderSent", "LastReminderSentDate",
                 "EventMode", "Organizer", "GuestOfHonor", "GiftBudget", "GiftDeadline", "StartTime", "EndTime")
        tree_container, self.tree_history = make_scrollable_treeview(f, columns=cols, height=18)
        widths = [110, 170, 85, 85, 110, 90, 140, 220, 220, 120, 130,
                  65, 45, 45, 55, 70,
                  180, 90,
                  85, 85, 80, 80, 75,
                  120, 140,
                  90, 140, 140, 100, 90, 65, 65]
        headers = ["Event ID", "Event Name", "Date", "Deadline", "Location", "Budget",
                   "Language", "Organizer Note", "Recipient File", "Sent Date", "Update Invite Date",
                   "Invited", "Yes", "No", "Maybe", "No Resp.",
                   "Report File", "Calendar Sent",
                   "Actual Att.", "Cost/Person", "Income", "Expense", "Balance",
                   "Reminder Sent", "Last Reminder Sent",
                   "Event Mode", "Organizer", "Guest of Honor", "Gift Budget", "Gift Deadline", "Start", "End"]
        for c, h, w in zip(cols, headers, widths):
            self.tree_history.heading(c, text=h)
            self.tree_history.column(c, width=w)
        tree_container.pack(fill="both", expand=True, padx=10, pady=6)
        # Every column is editable via double-click (editable_cols=None ->
        # no restriction) — see _commit_history_edit(). This only updates
        # what's shown in the table; nothing is saved to disk until
        # '💾 Save changes' is clicked.
        self.tree_history.bind(
            "<Double-1>",
            lambda e: self._on_editable_tree_double_click(
                self.tree_history, e, None, self._commit_history_edit))
        # Select rows and use Ctrl+C / Ctrl+V to copy/paste to or from
        # Excel — pasted cells go through the same _commit_history_edit()
        # used by double-click, so it only edits what's on screen; nothing
        # is written to the database until '💾 Save changes' is clicked.
        self._enable_treeview_copy_paste(self.tree_history, on_commit=self._commit_history_edit)

        self._refresh_history_tree()

    def _browse_history_file(self):
        path = filedialog.askopenfilename(filetypes=[("SQLite database", "*.db"), ("All files", "*.*")])
        if path:
            self.history_path.set(path)
            self._refresh_history_tree()

    def _refresh_history_tree(self):
        self.tree_history.delete(*self.tree_history.get_children())
        try:
            records = db.load_history(self.history_path.get())
        except Exception:
            records = []
        for rec in records:
            # Use the ORIGINAL Event ID as this row's Treeview iid — even if
            # the user later edits the displayed Event ID cell, the iid
            # stays fixed, so _save_history_edits() can still tell which
            # row on disk each edited row corresponds to (needed to support
            # renaming the Event ID without creating a duplicate row).
            base_iid = (rec.get("EventID") or "").strip() or "(blank)"
            iid = base_iid
            suffix = 2
            while self.tree_history.exists(iid):
                iid = f"{base_iid}__{suffix}"
                suffix += 1
            self.tree_history.insert("", "end", iid=iid, values=(
                rec.get("EventID"), rec.get("EventName"), rec.get("EventDate"),
                rec.get("Deadline"), rec.get("Location"), rec.get("Budget"),
                rec.get("EmailLanguage"), rec.get("OrganizerNote"), rec.get("RecipientFile"),
                rec.get("SentDate"), rec.get("UpdateInviteDate"),
                rec.get("TotalInvited"), rec.get("Yes"), rec.get("No"), rec.get("Maybe"), rec.get("NoResponse"),
                rec.get("ReportFile"), rec.get("CalendarSent"),
                rec.get("ActualAttendees"), rec.get("CostPerPerson"),
                rec.get("TotalIncome"), rec.get("TotalExpense"), rec.get("Balance"),
                rec.get("ReminderSent"), rec.get("LastReminderSentDate"),
                rec.get("EventMode"), rec.get("Organizer"), rec.get("GuestOfHonor"),
                rec.get("GiftBudget"), rec.get("GiftDeadline"), rec.get("StartTime"), rec.get("EndTime"),
            ))

    def _commit_history_edit(self, row_id, col_name, new_value):
        """Called after double-click editing a cell on Tab 7 — only updates
        what's displayed in the Treeview (in memory). Nothing touches the
        database until _save_history_edits() runs."""
        if not self.tree_history.exists(row_id):
            return
        self.tree_history.set(row_id, col_name, new_value)

    def _delete_history_row_by_event_id(self, event_id, path):
        """Deletes the row whose EventID matches event_id — used only to
        clean up the OLD row after a rename (see _save_history_edits()).
        Thin wrapper around db.delete_event() so the caller doesn't need to
        know it's now backed by SQLite instead of an Excel row."""
        if not event_id:
            return False
        try:
            return db.delete_event(event_id, path)
        except Exception:
            return False

    def _save_history_edits(self):
        """Writes every row currently shown in the Tab 7 table back to the
        database — this is what actually persists any double-click edits
        made above. Uses db.save_event_record() for each row (matches/
        updates by EventID, same as the rest of the app), so it correctly
        UPDATES existing rows rather than duplicating them. If a row's
        Event ID was changed (renamed), the row is saved under the NEW
        Event ID and the OLD row is then removed, so renaming works
        cleanly instead of leaving a stale duplicate behind."""
        if not hasattr(self, "tree_history"):
            return
        cols = self.tree_history["columns"]
        path = self.history_path.get()
        saved = 0
        renamed = 0
        errors = []
        for row_id in self.tree_history.get_children():
            values = self.tree_history.item(row_id, "values")
            record = dict(zip(cols, values))
            new_id = (record.get("EventID") or "").strip()
            if not new_id:
                continue  # no Event ID to key off of — can't safely save this row
            try:
                db.save_event_record(record, path)
                saved += 1
                original_id = row_id  # Treeview iid was set to the ORIGINAL EventID on load
                if original_id and original_id != new_id:
                    if self._delete_history_row_by_event_id(original_id, path):
                        renamed += 1
            except Exception as e:
                errors.append(f"{new_id}: {e}")

        self._refresh_history_tree()
        if errors:
            messagebox.showerror(
                "Some rows failed to save",
                f"Saved {saved} row(s), but {len(errors)} failed:\n\n" + "\n".join(errors[:10])
            )
        else:
            note = f"\n\n({renamed} Event ID rename(s) applied.)" if renamed else ""
            messagebox.showinfo("Saved", f"Saved {saved} row(s) to the database.{note}")

    def _export_history_to_excel(self):
        """Exports every row currently shown in the Tab 7 table to a new
        RSVP_History.xlsx-style Excel file — on demand only, since the
        database is now the actual working copy (see db.py). Prompts for a
        save location so it doesn't silently overwrite an old Excel file
        left over from before the SQLite migration."""
        if not hasattr(self, "tree_history") or not self.tree_history.get_children():
            messagebox.showwarning("Nothing to export", "There's no data in the table to export yet.")
            return
        out_path = filedialog.asksaveasfilename(
            title="Export History to Excel",
            defaultextension=".xlsx",
            initialfile="RSVP_History_export.xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not out_path:
            return
        try:
            cols = self.tree_history["columns"]
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "History"
            for i, c in enumerate(cols, start=1):
                cell = ws.cell(row=1, column=i, value=c)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="003366")
                cell.alignment = Alignment(horizontal="center")
            for r, row_id in enumerate(self.tree_history.get_children(), start=2):
                values = self.tree_history.item(row_id, "values")
                for c, v in enumerate(values, start=1):
                    ws.cell(row=r, column=c, value=v)
            for i in range(1, len(cols) + 1):
                ws.column_dimensions[get_column_letter(i)].width = 16
            wb.save(out_path)
        except Exception as e:
            messagebox.showerror("Error", f"Couldn't export to Excel:\n{e}")
            return
        messagebox.showinfo("Exported", f"Saved a snapshot of the History table to:\n{out_path}")


if __name__ == "__main__":
    app = RSVPApp()
    app.mainloop()
