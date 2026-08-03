"""
db.py — Lớp lưu trữ nội bộ trung tâm cho Outlook RSVP Tool (SQLite)
=============================================================================
File mặc định: rsvp_data.db (nằm cùng thư mục với app)

MỤC ĐÍCH (thay thế kiến trúc cũ):
    Trước đây, dữ liệu "sống" trực tiếp trong nhiều file Excel
    (RSVP_History.xlsx, Participant_List_{EventID}.xlsx,
    Gift_Contribution_List_{EventID}.xlsx, Attendance_Payment_{EventID}.xlsx)
    — mỗi lần user tick/sửa 1 ô là phải mở lại workbook, ghi định dạng, lưu
    xuống đĩa. Vừa chậm, vừa dễ xung đột nếu file đang mở trong Excel/
    OneDrive, vừa tạo nhiều file rải rác.

    Giờ TOÀN BỘ dữ liệu (Events/History, Recipients, Gift Contribution,
    Attendance & Payment, Responded results) sống trong 1 file SQLite DUY
    NHẤT (rsvp_data.db). Excel CHỈ còn được tạo ra khi user chủ động bấm
    "Export to Excel" để báo cáo/gửi cho người khác — không còn tự động
    ghi Excel liên tục.

    File Excel CŨ (nếu có, từ trước khi chuyển sang kiến trúc này) được TỰ
    ĐỘNG NHẬP 1 LẦN vào DB khi app khởi động lần đầu — xem
    migrate_from_excel_if_needed() — để không mất dữ liệu các sự kiện cũ.

THIẾT KẾ BẢNG:
    events              — 1 dòng / sự kiện, y hệt 31 cột RSVP_History.xlsx cũ
                           + LastScanTime (thời điểm Scan Inbox gần nhất)
    recipients          — danh sách người nhận (Tab 2), theo (EventID, Email)
    gift_contributions  — theo dõi quyên góp quà (Tab 6), theo (EventID, Email)
    attendance          — theo dõi tham dự & đóng góp (Tab 5), theo (EventID, Email)
    responses           — kết quả Scan Inbox đã quét (Tab 4), theo (EventID, Email)

API được thiết kế GIỐNG HỆT history.py cũ ở những hàm dùng chung
(save_event_record/load_history) để rsvp_app.py cần sửa ít nhất có thể.
"""
import os
import sqlite3
from datetime import datetime

DB_FILE_DEFAULT = "rsvp_data.db"

# Đúng thứ tự 31 cột đã dùng trong Tab 7 (Event History) của rsvp_app.py —
# xem cols = (...) trong _build_tab_history(). Giữ nguyên tên/thứ tự để
# tương thích ngược với mọi chỗ đang build dict theo các key này.
EVENT_COLUMNS = [
    "EventID", "EventName", "EventDate", "Deadline", "Location", "Budget",
    "EmailLanguage", "OrganizerNote", "RecipientFile", "SentDate", "UpdateInviteDate",
    "TotalInvited", "Yes", "No", "Maybe", "NoResponse",
    "ReportFile", "CalendarSent",
    "ActualAttendees", "CostPerPerson", "TotalIncome", "TotalExpense", "Balance",
    "ReminderSent", "LastReminderSentDate",
    "EventMode", "Organizer", "GuestOfHonor", "GiftBudget", "GiftDeadline", "StartTime", "EndTime",
    # MỚI: "Amount paid" nhập tay ở Tab 5 (Attendance & Payment) — số tiền
    # THỰC TẾ đã trả/chuyển cho ai đó (vd trả trước cho nhà hàng), khác với
    # "TotalCollectedAmount" (tổng cộng dồn từ cột Amount của từng người
    # trong bảng Attendance). "Remaining amount" = TotalCollectedAmount -
    # AmountPaid, tự tính ở UI, KHÔNG lưu riêng (tính lại mỗi lần cần).
    "AmountPaid",
]

_SCHEMA = """
CREATE TABLE IF NOT EXISTS events (
    EventID TEXT PRIMARY KEY,
    EventName TEXT, EventDate TEXT, Deadline TEXT, Location TEXT, Budget TEXT,
    EmailLanguage TEXT, OrganizerNote TEXT, RecipientFile TEXT, SentDate TEXT, UpdateInviteDate TEXT,
    TotalInvited TEXT, Yes TEXT, No TEXT, Maybe TEXT, NoResponse TEXT,
    ReportFile TEXT, CalendarSent TEXT,
    ActualAttendees TEXT, CostPerPerson TEXT, TotalIncome TEXT, TotalExpense TEXT, Balance TEXT,
    ReminderSent TEXT, LastReminderSentDate TEXT,
    EventMode TEXT, Organizer TEXT, GuestOfHonor TEXT, GiftBudget TEXT, GiftDeadline TEXT,
    StartTime TEXT, EndTime TEXT,
    AmountPaid TEXT,
    LastScanTime TEXT,
    RowOrder INTEGER
);
CREATE TABLE IF NOT EXISTS recipients (
    EventID TEXT NOT NULL, Email TEXT NOT NULL, Name TEXT,
    PRIMARY KEY (EventID, Email)
);
CREATE TABLE IF NOT EXISTS gift_contributions (
    EventID TEXT NOT NULL, Email TEXT NOT NULL, Name TEXT,
    Checked INTEGER DEFAULT 0, Amount REAL DEFAULT 0,
    SendEmail INTEGER DEFAULT 0,
    PRIMARY KEY (EventID, Email)
);
CREATE TABLE IF NOT EXISTS attendance (
    EventID TEXT NOT NULL, Email TEXT NOT NULL, Name TEXT, Vote TEXT,
    ActualAttend TEXT, Free INTEGER DEFAULT 0, Amount REAL DEFAULT 0,
    PRIMARY KEY (EventID, Email)
);
CREATE TABLE IF NOT EXISTS responses (
    EventID TEXT NOT NULL, Email TEXT NOT NULL, Name TEXT, Vote TEXT, ReceivedAt TEXT,
    Manual INTEGER DEFAULT 0,
    PRIMARY KEY (EventID, Email)
);
"""


def get_connection(path=DB_FILE_DEFAULT):
    """Mở kết nối SQLite, tự tạo file + schema nếu chưa có. Mỗi lời gọi mở
    1 connection MỚI rồi đóng lại ngay sau khi dùng xong (xem các hàm bên
    dưới) — đơn giản, an toàn cho 1 app single-user chạy trên máy cá nhân,
    không cần connection pool phức tạp."""
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    conn.executescript(_SCHEMA)
    # MỚI: "CREATE TABLE IF NOT EXISTS" ở trên chỉ tạo bảng MỚI đúng schema
    # đầy đủ — nếu file rsvp_data.db đã tồn tại TỪ TRƯỚC (tạo bởi bản code
    # cũ, chưa có cột AmountPaid), bảng "events" vẫn giữ nguyên các cột CŨ,
    # KHÔNG tự thêm cột mới. Cần ALTER TABLE riêng, best-effort: nếu cột đã
    # có sẵn (DB mới tạo từ _SCHEMA ở trên đã có rồi) thì sqlite3 báo lỗi
    # "duplicate column name" — bỏ qua lỗi đó là an toàn.
    try:
        conn.execute("ALTER TABLE events ADD COLUMN AmountPaid TEXT")
        conn.commit()
    except sqlite3.OperationalError:
        pass  # cột đã tồn tại — không cần làm gì thêm
    try:
        conn.execute("ALTER TABLE responses ADD COLUMN Manual INTEGER DEFAULT 0")
        conn.commit()
    except sqlite3.OperationalError:
        pass  # cột đã tồn tại — không cần làm gì thêm
    try:
        conn.execute("ALTER TABLE gift_contributions ADD COLUMN SendEmail INTEGER DEFAULT 0")
        conn.commit()
    except sqlite3.OperationalError:
        pass  # cột đã tồn tại — không cần làm gì thêm
    return conn


# ══════════════════════════════════════════════════════════════════════════
# EVENTS (thay RSVP_History.xlsx)
# ══════════════════════════════════════════════════════════════════════════

def save_event_record(record: dict, path=DB_FILE_DEFAULT):
    """Ghi/cập nhật 1 sự kiện — API GIỐNG HỆT history.py cũ: nếu EventID đã
    tồn tại thì UPDATE (chỉ đè các cột CÓ trong record, giữ nguyên các cột
    khác — merge-preserve, không phải ghi đè toàn bộ dòng), nếu chưa có thì
    INSERT dòng mới. record: dict với các key trong EVENT_COLUMNS (thiếu
    key nào thì giữ nguyên giá trị cũ nếu đang UPDATE, hoặc để trống nếu là
    dòng MỚI)."""
    event_id = record.get("EventID")
    if not event_id:
        raise ValueError("record cần có 'EventID'")
    conn = get_connection(path)
    try:
        cur = conn.execute("SELECT * FROM events WHERE EventID = ?", (event_id,))
        existing = cur.fetchone()
        if existing is None:
            record.setdefault("SentDate", datetime.now().strftime("%Y-%m-%d %H:%M"))
            cur = conn.execute("SELECT COALESCE(MAX(RowOrder), 0) + 1 FROM events")
            row_order = cur.fetchone()[0]
            cols = list(EVENT_COLUMNS) + ["RowOrder"]
            vals = [record.get(c, "") for c in EVENT_COLUMNS] + [row_order]
            placeholders = ", ".join("?" * len(cols))
            conn.execute(f"INSERT INTO events ({', '.join(cols)}) VALUES ({placeholders})", vals)
        else:
            # merge-preserve: chỉ cập nhật cột nào THỰC SỰ có mặt trong
            # record — cột không được truyền vào giữ nguyên giá trị cũ
            # (đúng hành vi history.py cũ, xem PROJECT_CONTEXT).
            set_cols = [c for c in EVENT_COLUMNS if c in record and c != "EventID"]
            if set_cols:
                set_clause = ", ".join(f"{c} = ?" for c in set_cols)
                vals = [record[c] for c in set_cols] + [event_id]
                conn.execute(f"UPDATE events SET {set_clause} WHERE EventID = ?", vals)
        conn.commit()
    finally:
        conn.close()
    return path


def load_history(path=DB_FILE_DEFAULT):
    """Trả về list[dict] — mỗi dict là 1 sự kiện đã lưu (mới nhất ở cuối,
    theo đúng thứ tự đã tạo — RowOrder), CÙNG ĐỊNH DẠNG với history.py cũ
    để mọi nơi đang dùng list[dict] của các cột EVENT_COLUMNS không cần
    sửa gì thêm."""
    conn = get_connection(path)
    try:
        rows = conn.execute("SELECT * FROM events ORDER BY RowOrder ASC").fetchall()
    finally:
        conn.close()
    return [{c: row[c] for c in EVENT_COLUMNS} for row in rows]


def delete_event(event_id, path=DB_FILE_DEFAULT):
    """Xoá HẲN 1 sự kiện khỏi bảng events — dùng khi ĐỔI TÊN Event ID ở Tab
    7 (Event History): lưu dòng mới dưới ID mới rồi xoá dòng cũ, tránh bị
    trùng lặp (xem rsvp_app.py: _save_history_edits())."""
    conn = get_connection(path)
    try:
        cur = conn.execute("DELETE FROM events WHERE EventID = ?", (event_id,))
        conn.commit()
        return cur.rowcount > 0
    finally:
        conn.close()


def set_last_scan_time(event_id, when: datetime, path=DB_FILE_DEFAULT):
    conn = get_connection(path)
    try:
        conn.execute(
            "UPDATE events SET LastScanTime = ? WHERE EventID = ?",
            (when.strftime("%Y-%m-%d %H:%M") if when else None, event_id),
        )
        conn.commit()
    finally:
        conn.close()


def get_last_scan_time(event_id, path=DB_FILE_DEFAULT):
    conn = get_connection(path)
    try:
        row = conn.execute("SELECT LastScanTime FROM events WHERE EventID = ?", (event_id,)).fetchone()
    finally:
        conn.close()
    if not row or not row["LastScanTime"]:
        return None
    try:
        return datetime.strptime(row["LastScanTime"], "%Y-%m-%d %H:%M")
    except Exception:
        return None


# ══════════════════════════════════════════════════════════════════════════
# RECIPIENTS (thay Participant_List_{EventID}.xlsx)
# ══════════════════════════════════════════════════════════════════════════

def save_recipients(event_id, recipients, path=DB_FILE_DEFAULT):
    """Ghi ĐÈ TOÀN BỘ danh sách người nhận của 1 sự kiện — recipients:
    list[(name, email)]. Dùng DELETE + INSERT lại từ đầu (đơn giản, đúng
    ngữ nghĩa "đây là danh sách hiện tại", không cần diff phức tạp — Tab 2
    luôn thao tác trên TOÀN BỘ danh sách cùng lúc, không sửa từng dòng lẻ
    tẻ như Gift/Attendance)."""
    conn = get_connection(path)
    try:
        conn.execute("DELETE FROM recipients WHERE EventID = ?", (event_id,))
        conn.executemany(
            "INSERT OR REPLACE INTO recipients (EventID, Email, Name) VALUES (?, ?, ?)",
            [(event_id, email, name) for name, email in recipients],
        )
        conn.commit()
    finally:
        conn.close()


def load_recipients(event_id, path=DB_FILE_DEFAULT):
    """Trả về list[(name, email)] — thứ tự theo lúc lưu (rowid)."""
    conn = get_connection(path)
    try:
        rows = conn.execute(
            "SELECT Name, Email FROM recipients WHERE EventID = ? ORDER BY rowid ASC", (event_id,)
        ).fetchall()
    finally:
        conn.close()
    return [(row["Name"], row["Email"]) for row in rows]


# ══════════════════════════════════════════════════════════════════════════
# GIFT CONTRIBUTIONS (thay Gift_Contribution_List_{EventID}.xlsx)
# ══════════════════════════════════════════════════════════════════════════

def save_gift_roster(event_id, roster: dict, path=DB_FILE_DEFAULT):
    """roster: dict[email -> {"name","checked","amount","send_email"}] — ghi
    đè toàn bộ (giống save_recipients, đơn giản và đủ nhanh cho quy mô vài
    chục người/sự kiện — auto-save mỗi lần tick vẫn nhanh vì SQLite ghi 1
    file nhỏ, không phải định dạng lại cả workbook Excel như trước).
    MỚI: "send_email" (mặc định False nếu thiếu key, tương thích ngược với
    dữ liệu cũ chưa có cột này) — đánh dấu người này được CHỌN để nhận email
    báo cáo quyên góp (Tab 6, xem _send_gift_report_email() trong
    rsvp_app.py), độc lập với "checked" (đã đóng góp hay chưa)."""
    conn = get_connection(path)
    try:
        conn.execute("DELETE FROM gift_contributions WHERE EventID = ?", (event_id,))
        conn.executemany(
            "INSERT OR REPLACE INTO gift_contributions (EventID, Email, Name, Checked, Amount, SendEmail) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            [(event_id, email, info.get("name", email), 1 if info.get("checked") else 0,
              info.get("amount", 0.0), 1 if info.get("send_email") else 0) for email, info in roster.items()],
        )
        conn.commit()
    finally:
        conn.close()


def load_gift_roster(event_id, path=DB_FILE_DEFAULT):
    conn = get_connection(path)
    try:
        rows = conn.execute(
            "SELECT Email, Name, Checked, Amount, SendEmail FROM gift_contributions "
            "WHERE EventID = ? ORDER BY rowid ASC", (event_id,)
        ).fetchall()
    finally:
        conn.close()
    return {
        row["Email"]: {
            "name": row["Name"], "checked": bool(row["Checked"]), "amount": row["Amount"] or 0.0,
            "send_email": bool(row["SendEmail"]),
        }
        for row in rows
    }


# ══════════════════════════════════════════════════════════════════════════
# ATTENDANCE & PAYMENT (thay Attendance_Payment_{EventID}.xlsx, sheet 1)
# ══════════════════════════════════════════════════════════════════════════

def save_attendance_roster(event_id, roster: dict, path=DB_FILE_DEFAULT):
    """roster: dict[email -> {"name","vote","actual_attend","free","amount"}]"""
    conn = get_connection(path)
    try:
        conn.execute("DELETE FROM attendance WHERE EventID = ?", (event_id,))
        conn.executemany(
            "INSERT OR REPLACE INTO attendance "
            "(EventID, Email, Name, Vote, ActualAttend, Free, Amount) VALUES (?, ?, ?, ?, ?, ?, ?)",
            [(event_id, email, info.get("name", email), info.get("vote", ""),
              info.get("actual_attend", ""), 1 if info.get("free") else 0,
              info.get("amount", 0.0)) for email, info in roster.items()],
        )
        conn.commit()
    finally:
        conn.close()


def load_attendance_roster(event_id, path=DB_FILE_DEFAULT):
    conn = get_connection(path)
    try:
        rows = conn.execute(
            "SELECT Email, Name, Vote, ActualAttend, Free, Amount FROM attendance "
            "WHERE EventID = ? ORDER BY rowid ASC", (event_id,)
        ).fetchall()
    finally:
        conn.close()
    return {
        row["Email"]: {
            "name": row["Name"], "vote": row["Vote"], "actual_attend": row["ActualAttend"],
            "free": bool(row["Free"]), "amount": row["Amount"] or 0.0,
        }
        for row in rows
    }


# ══════════════════════════════════════════════════════════════════════════
# RESPONSES / RESPONDED RESULT (thay Attendance_Payment_{EventID}.xlsx, sheet 2)
# ══════════════════════════════════════════════════════════════════════════

def save_responses(event_id, responses: dict, last_scan_time=None, path=DB_FILE_DEFAULT):
    """responses: dict[email.lower() -> {"name","vote","received": datetime|None,
    "manual": bool}] — đúng định dạng self.responses trong rsvp_app.py.
    MỚI: "manual" (mặc định False nếu thiếu key, để tương thích ngược với
    dữ liệu cũ chưa có cờ này) — đánh dấu phiếu vote này được SỬA TAY trên
    Tab 4 (xem _commit_response_vote_edit() trong rsvp_app.py), thay vì đọc
    được từ 1 email vote thật qua Scan Inbox. Cờ này được _collect_responses()
    dùng để KHÔNG cho lần Scan Inbox tiếp theo ghi đè mất phiếu sửa tay, trừ
    khi tìm thấy 1 email vote MỚI HƠN thời điểm sửa tay."""
    conn = get_connection(path)
    try:
        conn.execute("DELETE FROM responses WHERE EventID = ?", (event_id,))
        rows = []
        for email, info in responses.items():
            received = info.get("received")
            received_str = received.strftime("%Y-%m-%d %H:%M") if received else ""
            rows.append((event_id, email, info.get("name", ""), info.get("vote", ""),
                         received_str, 1 if info.get("manual") else 0))
        conn.executemany(
            "INSERT OR REPLACE INTO responses (EventID, Email, Name, Vote, ReceivedAt, Manual) "
            "VALUES (?, ?, ?, ?, ?, ?)", rows,
        )
        if last_scan_time is not None:
            conn.execute(
                "UPDATE events SET LastScanTime = ? WHERE EventID = ?",
                (last_scan_time.strftime("%Y-%m-%d %H:%M"), event_id),
            )
        conn.commit()
    finally:
        conn.close()


def load_responses(event_id, path=DB_FILE_DEFAULT):
    """Trả về (responses_dict, last_scan_time) — responses_dict đúng định
    dạng self.responses (key = email.lower()), CÓ kèm "manual" (bool)."""
    conn = get_connection(path)
    try:
        rows = conn.execute(
            "SELECT Email, Name, Vote, ReceivedAt, Manual FROM responses WHERE EventID = ? ORDER BY rowid ASC",
            (event_id,)
        ).fetchall()
    finally:
        conn.close()
    responses = {}
    for row in rows:
        received_dt = None
        if row["ReceivedAt"]:
            try:
                received_dt = datetime.strptime(row["ReceivedAt"], "%Y-%m-%d %H:%M")
            except Exception:
                received_dt = None
        responses[row["Email"].lower()] = {
            "name": row["Name"], "vote": row["Vote"], "received": received_dt,
            "manual": bool(row["Manual"]),
        }
    return responses, get_last_scan_time(event_id, path)


# ══════════════════════════════════════════════════════════════════════════
# MIGRATION — nhập 1 LẦN dữ liệu từ các file Excel CŨ (nếu có) vào DB, để
# không mất lịch sử các sự kiện đã lưu TRƯỚC KHI chuyển sang kiến trúc này.
# ══════════════════════════════════════════════════════════════════════════

def migrate_from_excel_if_needed(db_path=DB_FILE_DEFAULT, history_xlsx="RSVP_History.xlsx", search_dir=None):
    """Chỉ chạy MỘT LẦN DUY NHẤT: nếu DB đã có ít nhất 1 sự kiện rồi (tức
    đã từng migrate hoặc đã dùng DB từ đầu), KHÔNG làm gì cả — an toàn khi
    gọi lại nhiều lần (vd mỗi lần mở app). Nếu DB đang trống VÀ tìm thấy
    RSVP_History.xlsx cũ, đọc toàn bộ sự kiện từ đó, rồi với MỖI Event ID,
    tìm thêm Participant_List_{EventID}.xlsx / Gift_Contribution_List_
    {EventID}.xlsx / Attendance_Payment_{EventID}.xlsx (sheet "Attendance &
    Payment" + "Responded result") CÙNG THƯ MỤC để nhập nốt Recipients/
    Gift/Attendance/Responses tương ứng — best-effort, thiếu file nào thì
    bỏ qua file đó, không chặn việc nhập các phần còn lại.

    Trả về (migrated: bool, event_count: int, notes: list[str]) để app
    hiển thị cho user biết đã nhập được những gì."""
    notes = []
    conn = get_connection(db_path)
    try:
        existing_count = conn.execute("SELECT COUNT(*) FROM events").fetchone()[0]
    finally:
        conn.close()
    if existing_count > 0:
        return False, 0, notes  # DB đã có dữ liệu -> không migrate lại, tránh nhân đôi

    if not os.path.exists(history_xlsx):
        return False, 0, notes  # không có gì để nhập — coi như bắt đầu mới hoàn toàn

    try:
        import openpyxl
    except ImportError:
        notes.append("openpyxl chưa cài — không migrate được.")
        return False, 0, notes

    search_dir = search_dir or os.path.dirname(os.path.abspath(history_xlsx)) or "."

    try:
        wb = openpyxl.load_workbook(history_xlsx, data_only=True)
        ws = wb["History"] if "History" in wb.sheetnames else wb.active
        # File cũ có thể chỉ 24 cột (bản trước khi thêm Event mode/Gift) —
        # zip với EVENT_COLUMNS và lấy None cho cột thiếu, giống hệt cách
        # history.py cũ tự "migrate" file thiếu cột.
        header_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            rec = dict(zip(header_row, row))
            rows.append(rec)
    except Exception as e:
        notes.append(f"Không đọc được {history_xlsx}: {e}")
        return False, 0, notes

    imported = 0
    for rec in rows:
        event_id = rec.get("EventID")
        if not event_id:
            continue
        clean_rec = {c: rec.get(c, "") for c in EVENT_COLUMNS}
        clean_rec["EventID"] = event_id
        save_event_record(clean_rec, db_path)
        imported += 1

        # Recipients — từ Participant_List_{EventID}.xlsx (đường dẫn ghi
        # trong cột RecipientFile, hoặc suy ra theo tên chuẩn nếu cột trống).
        # BUG ĐÃ SỬA: file này dùng ĐÚNG format của _load_recipients() /
        # _export_recipients_to_excel() trong rsvp_app.py — sheet "DanhSach"
        # (hoặc active nếu không có), HEADER Ở DÒNG 3 ("Họ tên"/"Email"), dữ
        # liệu bắt đầu DÒNG 4 — KHÔNG PHẢI header ở dòng 1 như bản migration
        # trước đó lỡ giả định (khiến không tìm thấy cột nào, bỏ qua hết
        # người nhận — đúng như lỗi "Recipients trống sau khi Load setup").
        # Cũng bỏ qua đúng dòng "(ví dụ)"/"(example)" của file mẫu TEMPLATE,
        # y hệt _load_recipients().
        recipient_path = rec.get("RecipientFile") or os.path.join(
            search_dir, f"Participant_List_{event_id}.xlsx")
        if recipient_path and os.path.exists(recipient_path):
            try:
                rwb = openpyxl.load_workbook(recipient_path, data_only=True)
                rws = rwb["DanhSach"] if "DanhSach" in rwb.sheetnames else rwb.active
                people = []
                for r in rws.iter_rows(min_row=4, values_only=True):
                    name, email = (r + (None, None))[:2] if r else (None, None)
                    if not email or "@" not in str(email):
                        continue
                    if "(ví dụ)" in str(name or "") or "(example)" in str(name or "").lower():
                        continue
                    people.append((str(name or "").strip(), str(email).strip()))
                if people:
                    save_recipients(event_id, people, db_path)
            except Exception:
                notes.append(f"[{event_id}] không nhập được danh sách người nhận.")

        # Gift Contribution — từ Gift_Contribution_List_{EventID}.xlsx
        gift_path = os.path.join(search_dir, f"Gift_Contribution_List_{event_id}.xlsx")
        if os.path.exists(gift_path):
            try:
                gwb = openpyxl.load_workbook(gift_path, data_only=True)
                gws = gwb.active
                gheader_row = next(gws.iter_rows(min_row=1, max_row=1), ())
                gheader = {(str(c.value).strip() if c.value else ""): i for i, c in enumerate(gheader_row)}
                i_name, i_email = gheader.get("Name"), gheader.get("Email")
                i_amount, i_contrib = gheader.get("Amount"), gheader.get("Contributed")
                roster = {}
                if i_email is not None:
                    for r in gws.iter_rows(min_row=2, values_only=True):
                        if not r or i_email >= len(r) or not r[i_email]:
                            continue
                        email = str(r[i_email]).strip()
                        roster[email] = {
                            "name": str(r[i_name]).strip() if i_name is not None and r[i_name] else email,
                            "checked": (str(r[i_contrib]).strip().lower() == "yes") if i_contrib is not None and i_contrib < len(r) and r[i_contrib] is not None else False,
                            "amount": float(r[i_amount]) if i_amount is not None and i_amount < len(r) and isinstance(r[i_amount], (int, float)) else 0.0,
                        }
                if roster:
                    save_gift_roster(event_id, roster, db_path)
            except Exception:
                notes.append(f"[{event_id}] không nhập được danh sách quyên góp quà.")

        # Attendance & Payment + Responded result — từ
        # Attendance_Payment_{EventID}.xlsx (2 sheet)
        attn_path = os.path.join(search_dir, f"Attendance_Payment_{event_id}.xlsx")
        if os.path.exists(attn_path):
            try:
                awb = openpyxl.load_workbook(attn_path, data_only=True)
                if "Attendance & Payment" in awb.sheetnames:
                    aws = awb["Attendance & Payment"]
                    aheader_row = next(aws.iter_rows(min_row=1, max_row=1), ())
                    aheader = {(str(c.value).strip() if c.value else ""): i for i, c in enumerate(aheader_row)}
                    i_email = aheader.get("Email")
                    roster = {}
                    if i_email is not None:
                        for r in aws.iter_rows(min_row=2, values_only=True):
                            if not r or i_email >= len(r) or not r[i_email]:
                                continue
                            email = str(r[i_email]).strip()
                            g = lambda key: (r[aheader[key]] if key in aheader and aheader[key] < len(r) else None)
                            amount = g("Amount")
                            roster[email] = {
                                "name": str(g("Name")).strip() if g("Name") else email,
                                "vote": str(g("Vote")).strip() if g("Vote") else "",
                                "actual_attend": str(g("Actual Attend")).strip() if g("Actual Attend") else "",
                                "free": str(g("Free")).strip().lower() == "yes" if g("Free") else False,
                                "amount": float(amount) if isinstance(amount, (int, float)) else 0.0,
                            }
                    if roster:
                        save_attendance_roster(event_id, roster, db_path)
                if "Responded result" in awb.sheetnames:
                    rws2 = awb["Responded result"]
                    rheader_row = next(rws2.iter_rows(min_row=1, max_row=1), ())
                    rheader = {(str(c.value).strip() if c.value else ""): i for i, c in enumerate(rheader_row)}
                    i_email = rheader.get("Email")
                    responses = {}
                    if i_email is not None:
                        for r in rws2.iter_rows(min_row=2, values_only=True):
                            if not r or i_email >= len(r) or not r[i_email]:
                                continue
                            email = str(r[i_email]).strip()
                            g = lambda key: (r[rheader[key]] if key in rheader and rheader[key] < len(r) else None)
                            received_raw = g("Received At")
                            received_dt = None
                            if received_raw:
                                try:
                                    received_dt = datetime.strptime(str(received_raw), "%Y-%m-%d %H:%M")
                                except Exception:
                                    received_dt = None
                            responses[email.lower()] = {
                                "name": str(g("Name")).strip() if g("Name") else "",
                                "vote": str(g("Vote")).strip() if g("Vote") else "",
                                "received": received_dt,
                            }
                    if responses:
                        save_responses(event_id, responses, None, db_path)
            except Exception:
                notes.append(f"[{event_id}] không nhập được dữ liệu Attendance/Responded.")

    return True, imported, notes
