"""
history.py — Lưu & đọc lịch sử các sự kiện RSVP vào 1 file Excel duy nhất
=============================================================================
File mặc định: RSVP_History.xlsx (nằm cùng thư mục với app)

Mục đích:
    - Mỗi lần hoàn tất 1 sự kiện (đã gửi mời, thu thập xong phản hồi),
      lưu lại 1 dòng vào đây để tracking lâu dài.
    - Lần sau tạo sự kiện mới, có thể "Nạp từ sự kiện cũ" để lấy lại
      các thông tin cấu hình (tên, địa điểm, ghi chú...) làm mẫu, đỡ gõ lại.
"""
import json
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HISTORY_FILE_DEFAULT = "RSVP_History.xlsx"
FIXED_OVERRIDES_FILE_DEFAULT = "fixed_wording_overrides.json"
PROMPT_OVERRIDES_FILE_DEFAULT = "prompt_overrides.json"


def load_fixed_overrides(path=FIXED_OVERRIDES_FILE_DEFAULT):
    """Đọc các bản 'câu văn mặc định' đã được người dùng tự chỉnh & lưu lại cho FIXED
    part (áp dụng cho mọi sự kiện sau này, cho tới khi người dùng Reset hoặc sửa lại)."""
    if not os.path.exists(path):
        return {"en": "", "ja": "", "vi": ""}
    try:
        with open(path, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        return {"en": data.get("en", ""), "ja": data.get("ja", ""), "vi": data.get("vi", "")}
    except Exception:
        return {"en": "", "ja": "", "vi": ""}


def save_fixed_overrides(overrides: dict, path=FIXED_OVERRIDES_FILE_DEFAULT):
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(overrides, fh, ensure_ascii=False, indent=2)
    return path


def load_prompt_overrides(path=PROMPT_OVERRIDES_FILE_DEFAULT):
    """Đọc custom prompt template được người dùng tùy chỉnh.
    Trả về dict {"single": "...", "bilingual": "..."} hoặc rỗng nếu chưa lưu."""
    if not os.path.exists(path):
        return {"single": "", "bilingual": ""}
    try:
        with open(path, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        return {
            "single": data.get("single", ""),
            "bilingual": data.get("bilingual", "")
        }
    except Exception:
        return {"single": "", "bilingual": ""}


def save_prompt_overrides(overrides: dict, path=PROMPT_OVERRIDES_FILE_DEFAULT):
    """Lưu custom prompt template được người dùng tùy chỉnh."""
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(overrides, fh, ensure_ascii=False, indent=2)
    return path

COLUMNS = [
    "EventID", "EventName", "EventDate", "Deadline", "Location", "Budget",
    "EmailLanguage", "OrganizerNote", "RecipientFile", "SentDate",
    "UpdateInviteDate",  # thời điểm gần nhất 1 email "Send update invite" được
                          # gửi cho sự kiện này (khác với SentDate = lần gửi
                          # invite ĐẦU TIÊN) — để trống nếu chưa từng gửi update.
    "TotalInvited", "Yes", "No", "Maybe", "NoResponse",
    "ReportFile", "CalendarSent",
    "ActualAttendees", "CostPerPerson", "TotalIncome", "TotalExpense", "Balance",
    "ReminderSent",  # datetime lần gần nhất email nhắc nhở TỰ ĐỘNG (Task
                      # Scheduler) đã được gửi cho sự kiện này — dùng để
                      # tránh gửi nhắc nhở trùng lặp mỗi ngày sau deadline.
                      # Xem send_scheduled_reminders.py.
    "LastReminderSentDate",  # datetime lần gần nhất người dùng bấm "Gửi
                      # email nhắc nhở" TỪ TAB 4 (thủ công) — ghi lại NGAY
                      # khi bấm gửi, dù ở chế độ "Gửi ngay" hay "Mở Outlook
                      # để review" (giống cách SentDate của invite gốc được
                      # ghi ngay lúc gửi/mở, xem _send_invite()). Cột này CHỈ
                      # để xem lại lịch sử ở Tab 6 — KHÔNG dùng để chống gửi
                      # trùng (đó là việc riêng của cột "ReminderSent" ở
                      # trên, chỉ set khi auto_send=True chắc chắn đã gửi).

    # ── MỚI: tính năng "Event mode" (Event / Gift / Event + Gift) ──
    "EventMode",      # "Event" | "Gift" | "Event + Gift" — chọn ở Tab 1,
                      # quyết định sự kiện này có kèm quyên góp mua quà tặng
                      # hay không (xem Tab 3 "Send Gift Contribution Notice").
    "Organizer",      # Người phụ trách nhận đóng góp (hiện trong email Gift).
    "GuestOfHonor",   # Người được tặng quà (vd người sắp nghỉ việc).
    "GiftBudget",     # Ngân sách dự kiến CHO QUÀ TẶNG — khác với cột "Budget"
                      # (ngân sách sự kiện nói chung, hiển thị là "Expected
                      # event budget" trên UI từ nay để phân biệt rõ 2 khoản).
    "StartTime",      # Giờ bắt đầu sự kiện (HH:MM) — chuyển từ Tab 5 lên
                      # Tab 1 (ngay dưới Event Date), lưu vào History để
                      # Tab 5 (Send Meeting) dùng lại khi tạo Calendar Invite.
    "EndTime",        # Giờ kết thúc sự kiện (HH:MM) — tương tự StartTime.
    "GiftDeadline",   # Hạn ĐÓNG GÓP QUÀ TẶNG — khác với cột "Deadline" (hạn
                      # RSVP có tham dự hay không). Dùng cho email "Send Gift
                      # Contribution Notice" ở Tab 3.
]

# Độ rộng cột tương ứng 1-1 với COLUMNS (dùng cả khi tạo file mới lẫn khi
# migrate thêm cột vào file cũ — xem _migrate_history_columns()).
COLUMN_WIDTHS = [16, 28, 14, 12, 20, 16, 14, 30, 26, 16, 16, 12, 8, 8, 8, 12, 26, 12,
                  16, 14, 14, 14, 12, 16, 18,
                  14, 18, 18, 16, 10, 10, 14]


def _migrate_history_columns(path):
    """Đảm bảo file History HIỆN CÓ có đủ tất cả các cột trong COLUMNS, kể cả
    khi file đó được tạo bằng phiên bản CODE CŨ (thiếu cột mới hơn, vd
    'UpdateInviteDate') — bằng cách CHÈN cột còn thiếu vào file, thay vì để
    load_history()/save_event_record() cứ đọc/ghi theo VỊ TRÍ CỐ ĐỊNH trong
    COLUMNS trong khi file thật lại có ÍT CỘT HƠN.

    ⚠️ Đây là lỗi NGHIÊM TRỌNG nếu không migrate: vì load_history() dùng
    `zip(COLUMNS, row_values)` để gán tên cột theo VỊ TRÍ, một khi COLUMNS
    (trong code) có thêm cột ở GIỮA danh sách (không phải cuối) mà file thật
    chưa có, mọi cột SAU vị trí đó sẽ bị đọc/ghi NHẦM SANG TÊN CỘT KHÁC — vd
    dữ liệu cột "TotalInvited" cũ bị đọc nhầm thành "UpdateInviteDate", "Yes"
    bị đọc nhầm thành "TotalInvited", v.v. Hàm này tránh việc đó bằng cách
    chèn đúng 1 cột trống mới vào ĐÚNG VỊ TRÍ mong muốn (ngay sau SentDate),
    dùng insert_cols() của openpyxl — thao tác này tự dịch toàn bộ dữ liệu
    của các cột phía sau sang phải 1 cột cho MỌI dòng, giữ nguyên tương ứng
    dữ liệu-tên cột, không làm lệch/mất dữ liệu các dòng đã có.

    An toàn để gọi lại nhiều lần (không làm gì nếu file đã đủ cột)."""
    wb = openpyxl.load_workbook(path)
    ws = wb["History"] if "History" in wb.sheetnames else wb.active
    header = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    changed = False

    def _style_header_cell(c, text):
        c.value = text
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="003366")
        c.alignment = Alignment(horizontal="center")

    # "UpdateInviteDate": chèn NGAY SAU "SentDate" nếu thiếu — đúng vị trí
    # trong COLUMNS — thay vì thêm ở cuối, để khớp layout với file mới tạo
    # (và khớp đúng yêu cầu "ngay phía bên phải của SentDate").
    if "UpdateInviteDate" not in header and "SentDate" in header:
        sent_pos = header.index("SentDate") + 1  # 1-based column index
        insert_at = sent_pos + 1
        ws.insert_cols(insert_at)
        _style_header_cell(ws.cell(row=1, column=insert_at), "UpdateInviteDate")
        header.insert(insert_at - 1, "UpdateInviteDate")
        changed = True

    # Lưới an toàn cho các cột KHÁC có thể thiếu (vd 1 file rất cũ chưa từng
    # có "ReminderSent") -> thêm vào CUỐI, không ảnh hưởng vị trí cột đã có.
    for col in COLUMNS:
        if col not in header:
            new_idx = ws.max_column + 1
            _style_header_cell(ws.cell(row=1, column=new_idx), col)
            header.append(col)
            changed = True

    if changed:
        # Áp lại độ rộng cột theo đúng COLUMNS/COLUMN_WIDTHS sau khi
        # insert_cols() (openpyxl không tự dịch chuyển column_dimensions của
        # các cột bị đẩy sang phải, nên set lại cho gọn thay vì để lệch).
        for i, col in enumerate(COLUMNS):
            if col in header:
                idx = header.index(col) + 1
                w = COLUMN_WIDTHS[i] if i < len(COLUMN_WIDTHS) else 14
                ws.column_dimensions[get_column_letter(idx)].width = w
        wb.save(path)
    return changed


def ensure_history_file(path=HISTORY_FILE_DEFAULT):
    if os.path.exists(path):
        _migrate_history_columns(path)  # tự thêm cột thiếu nếu file được tạo bằng bản code cũ hơn
        return path
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "History"
    for i, col in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=i, value=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="003366")
        c.alignment = Alignment(horizontal="center")
    # ⚠️ BUG ĐÃ SỬA: dùng chr(65 + i) để suy ra ký tự cột ("A", "B", ...) chỉ
    # đúng với TỐI ĐA 26 cột — khi COLUMNS vượt quá 26 (giờ đã 31 cột, sau
    # khi thêm các cột "Event mode"/Gift), chr(65+26) = '[' (không phải "AA")
    # → ws.column_dimensions['['] khiến openpyxl báo lỗi "not a valid column
    # name" ngay khi LƯU file — tức là ensure_history_file() (được gọi ở đầu
    # MỌI thao tác đọc/ghi History) sập hoàn toàn với file History MỚI TẠO
    # (chưa tồn tại). Sửa bằng get_column_letter() (đã import sẵn ở đầu file)
    # — hàm chuẩn của openpyxl, tự xử lý đúng cả các cột sau Z (AA, AB, ...).
    widths = COLUMN_WIDTHS
    for i in range(len(COLUMNS)):
        col_letter = get_column_letter(i + 1)
        w = widths[i] if i < len(widths) else 14
        ws.column_dimensions[col_letter].width = w
    wb.save(path)
    return path


def load_history(path=HISTORY_FILE_DEFAULT):
    """Trả về list[dict] — mỗi dict là 1 sự kiện đã lưu (mới nhất ở cuối)."""
    ensure_history_file(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["History"] if "History" in wb.sheetnames else wb.active
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        rec = dict(zip(COLUMNS, list(row) + [None] * (len(COLUMNS) - len(row))))
        records.append(rec)
    return records


def save_event_record(record: dict, path=HISTORY_FILE_DEFAULT):
    """
    Ghi thêm 1 dòng mới vào lịch sử. Nếu EventID đã tồn tại, CẬP NHẬT dòng đó
    thay vì tạo trùng (vd: chạy lại collect_responses nhiều lần cho cùng sự kiện).
    record: dict với các key trong COLUMNS.

    Với dòng MỚI: key nào thiếu trong `record` thì để trống ("") — như cũ.
    Với dòng ĐÃ TỒN TẠI (đang UPDATE): key nào thiếu trong `record` thì GIỮ
    NGUYÊN giá trị cũ đang có trên dòng đó, KHÔNG xoá trắng — vì có những cột
    được ghi bởi 1 nơi khác với nơi gọi save_event_record() lần này, ví dụ
    cột "ReminderSent" được send_scheduled_reminders.py (chạy nền qua Windows
    Task Scheduler) tự set, nhưng rsvp_app.py's _save_to_history() không hề
    biết tới field này — nếu không giữ nguyên, mỗi lần người dùng bấm "🗂 Save
    event to History" ở Tab 4 sẽ VÔ TÌNH xoá mất dấu "đã nhắc nhở", khiến
    script nền gửi nhắc nhở TRÙNG LẶP vào lần chạy kế tiếp.
    """
    ensure_history_file(path)
    wb = openpyxl.load_workbook(path)
    ws = wb["History"] if "History" in wb.sheetnames else wb.active

    event_id = record.get("EventID")
    target_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == event_id:
            target_row = r
            break
    is_new_row = target_row is None
    if is_new_row:
        target_row = ws.max_row + 1 if ws.cell(row=2, column=1).value else 2

    # ⚠️ BUG ĐÃ SỬA: trước đây dòng dưới đây chạy VÔ ĐIỀU KIỆN (cả khi
    # is_new_row=False), khiến bất kỳ lệnh save_event_record() nào KHÔNG kèm
    # "SentDate" (vd _push_recipient_file_to_history() chỉ gửi EventID +
    # RecipientFile khi Tab 2 "Load list" / Tab 1 "Load setup from selected
    # event") đều bị tự nhét datetime.now() vào record["SentDate"] — làm cho
    # nhánh "elif not is_new_row: giữ nguyên giá trị cũ" bên dưới KHÔNG BAO
    # GIỜ được chạy cho cột SentDate nữa (vì lúc đó "SentDate" đã có sẵn
    # trong record). Hậu quả: mỗi lần bấm "Load setup from selected event"
    # (hoặc Tab 2 "Load list") cho 1 sự kiện ĐÃ CÓ trong History, SentDate cũ
    # (kể cả khi người dùng vừa tự sửa tay đúng lại) bị ghi đè thành GIỜ HIỆN
    # TẠI. Giờ chỉ set mặc định SentDate = now() khi thực sự là DÒNG MỚI.
    if is_new_row:
        record.setdefault("SentDate", datetime.now().strftime("%Y-%m-%d %H:%M"))
    for i, col in enumerate(COLUMNS, start=1):
        if col in record:
            val = record[col]
        elif not is_new_row:
            val = ws.cell(row=target_row, column=i).value  # giữ nguyên giá trị cũ
        else:
            val = ""
        ws.cell(row=target_row, column=i, value=val)

    wb.save(path)
    return path
